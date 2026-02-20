from __future__ import annotations

import base64
from contextlib import contextmanager
from copy import copy
from datetime import date, datetime
import hashlib
import json
import os
from pathlib import Path
import re
import socket
import shutil
import sys
import tempfile
import time
import traceback
from typing import Any
from urllib.error import URLError
from urllib.parse import quote
from urllib.request import urlopen
from uuid import uuid4

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils.exceptions import InvalidFileException


DEFAULT_WORKBOOK_PATH = Path.home() / "Downloads" / "IT POs.xlsx"
ASSETS_DIR = Path(__file__).resolve().parent / "assets"
APP_ICON_PATH = ASSETS_DIR / "potrol-icon.svg"
APP_LOGO_PATH = ASSETS_DIR / "potrol-logo.svg"
DEFAULT_HEADERS = [
    "PO Number",
    "Date",
    "Vendor/Store",
    "Department",
    "Location",
    "Items Being Purchased",
    "Price Per Item",
    "Quantity",
    "Sub Total",
    "Shipping Cost",
    "Sales Tax",
    "Grand Total",
]
SUPPORTED_WORKBOOK_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}
REQUIRED_HEADERS: list[str] = []
DEFAULT_SHEET_NAME = "PO Log"
DEFAULT_LOCATION_OPTIONS = ["GLN", "MID", "AUR", "SNT", "CRN", "PHX", "LEB", "CAN"]
LOCATION_ALIAS_MAP: dict[str, str] = {
    "CNR": "CRN",
    "GLENPOOL": "GLN",
    "MIDDLEBURY": "MID",
    "MN": "MN",
    "MIN": "MN",
    "MINNESOTA": "MN",
    "NM": "NM",
    "NEWMEXICO": "NM",
    "CO": "CO",
    "COLORADO": "CO",
    "IN": "IN",
    "COON": "CRN",
    "RAPIDS": "CRN",
}
DEFAULT_DEPARTMENT_OPTIONS = sorted(
    [
        "IT",
        "Accounting",
        "Sales",
        "Executive",
        "Warehouse",
        "Shipping",
        "Remote worker",
        "HR",
        "Supply Chain",
        "Other",
    ],
    key=lambda value: value.casefold(),
)
LOCATION_CONFIG_PATH = Path.home() / ".potrol_locations.json"
APP_SETTINGS_PATH = Path.home() / ".potrol_settings.json"
APP_DRAFTS_PATH = Path.home() / ".potrol_drafts.json"
APP_RUNTIME_LOG_PATH = Path.home() / ".potrol_runtime.log"
PO_PREFIX = "IT"
PO_START_NUMBER = 579
PURCHASE_REASON_COLUMN_INDEX = 10
LIVE_PO_REFRESH_INTERVAL_SECONDS = 5
WORKBOOK_LOCK_TIMEOUT_SECONDS = 12.0
WORKBOOK_LOCK_STALE_SECONDS = 120.0
WORKBOOK_SYNC_INTERVAL_SECONDS = 5
PO_RESERVATION_STALE_SECONDS = 900.0
PO_RESERVATION_SYNC_SECONDS = 8.0
DRAFT_AUTOSAVE_MIN_SECONDS = 1.0
APP_VERSION = "2026.02.19.1"
WORKBOOK_OPEN_RETRY_COUNT = 3
WORKBOOK_OPEN_RETRY_DELAY_SECONDS = 0.35
PO_SEQUENCE_CACHE_MAX_KEYS = 32
PO_SCAN_EMPTY_STREAK_BREAK = 12000
PO_SCAN_HARD_ROW_LIMIT = 350000
DEFAULT_BACKUP_KEEP_LATEST = 1
MIN_BACKUP_KEEP_LATEST = 1
MAX_BACKUP_KEEP_LATEST = 25
MAX_RUNTIME_LOG_LINES = 1200
DEFAULT_EDITOR_PAGE_SIZE = 100
DEFAULT_EDITOR_SEARCH_SCAN_LIMIT = 10000
DEFAULT_THEME_NAME = "Sky"
THEME_PRESETS: dict[str, dict[str, str]] = {
    "Sky": {
        "bg_start": "#f3f6ff",
        "bg_end": "#f7f9ff",
        "surface": "#ffffff",
        "surface_soft": "#eef2ff",
        "border": "#d2d9e8",
        "outline": "#c5ccda",
        "text": "#1b1b1f",
        "muted": "#47464f",
        "accent": "#0b57d0",
        "accent_strong": "#0842a0",
    },
    "Slate": {
        "bg_start": "#f3f2f7",
        "bg_end": "#f8f7fb",
        "surface": "#ffffff",
        "surface_soft": "#ececf3",
        "border": "#d0d0db",
        "outline": "#c3c3d0",
        "text": "#1c1b1f",
        "muted": "#49454f",
        "accent": "#4f378b",
        "accent_strong": "#3a2867",
    },
    "Forest": {
        "bg_start": "#edf6ef",
        "bg_end": "#f5faf4",
        "surface": "#ffffff",
        "surface_soft": "#e8f4ea",
        "border": "#c8dcc9",
        "outline": "#b8cdb9",
        "text": "#1a1c19",
        "muted": "#3f4b3c",
        "accent": "#386a20",
        "accent_strong": "#285216",
    },
    "High Contrast": {
        "bg_start": "#f4f4f4",
        "bg_end": "#ffffff",
        "surface": "#ffffff",
        "surface_soft": "#f0f1f3",
        "border": "#80848d",
        "outline": "#646873",
        "text": "#11131a",
        "muted": "#232833",
        "accent": "#005ac1",
        "accent_strong": "#004391",
    },
    "Ocean": {
        "bg_start": "#e9f7fb",
        "bg_end": "#f4fcff",
        "surface": "#ffffff",
        "surface_soft": "#e8f5fb",
        "border": "#c1d9e5",
        "outline": "#a9c8d7",
        "text": "#132028",
        "muted": "#415863",
        "accent": "#006c84",
        "accent_strong": "#005267",
    },
    "Sand": {
        "bg_start": "#fbf6ef",
        "bg_end": "#fffaf4",
        "surface": "#ffffff",
        "surface_soft": "#f8efe2",
        "border": "#e3d4be",
        "outline": "#d2c0a8",
        "text": "#2b2418",
        "muted": "#5f5647",
        "accent": "#8a5a00",
        "accent_strong": "#6b4600",
    },
    "Rose": {
        "bg_start": "#fff1f5",
        "bg_end": "#fff8fb",
        "surface": "#ffffff",
        "surface_soft": "#ffeaf1",
        "border": "#e7c7d5",
        "outline": "#d8afbf",
        "text": "#2a1621",
        "muted": "#654454",
        "accent": "#a23a68",
        "accent_strong": "#7f2c51",
    },
    "Mint": {
        "bg_start": "#ecfbf6",
        "bg_end": "#f7fffb",
        "surface": "#ffffff",
        "surface_soft": "#e4f7ef",
        "border": "#bfdfce",
        "outline": "#a5ccb8",
        "text": "#15221d",
        "muted": "#3f5c50",
        "accent": "#0f766e",
        "accent_strong": "#0b5d56",
    },
    "Indigo": {
        "bg_start": "#f2f3ff",
        "bg_end": "#fafaff",
        "surface": "#ffffff",
        "surface_soft": "#eceeff",
        "border": "#cfd3ee",
        "outline": "#b5badf",
        "text": "#1d1f2f",
        "muted": "#4f5370",
        "accent": "#4456c7",
        "accent_strong": "#3343a6",
    },
    "Vibrant Coral": {
        "bg_start": "#fff0ec",
        "bg_end": "#fff8f6",
        "surface": "#ffffff",
        "surface_soft": "#ffe8df",
        "border": "#efc7b8",
        "outline": "#e4b6a4",
        "text": "#2e1712",
        "muted": "#6f4a42",
        "accent": "#ff5a3c",
        "accent_strong": "#d94329",
    },
    "Vibrant Violet": {
        "bg_start": "#f7f0ff",
        "bg_end": "#fcf7ff",
        "surface": "#ffffff",
        "surface_soft": "#f1e5ff",
        "border": "#d9c6ef",
        "outline": "#c7afe3",
        "text": "#23162f",
        "muted": "#57426c",
        "accent": "#8b3dff",
        "accent_strong": "#6e2dd0",
    },
    "Vibrant Teal": {
        "bg_start": "#eafcfb",
        "bg_end": "#f5ffff",
        "surface": "#ffffff",
        "surface_soft": "#dcf7f4",
        "border": "#b8e2de",
        "outline": "#9ecfca",
        "text": "#132826",
        "muted": "#3d605d",
        "accent": "#00a896",
        "accent_strong": "#008576",
    },
    "Neon Lime": {
        "bg_start": "#f7ffd9",
        "bg_end": "#feffef",
        "surface": "#ffffff",
        "surface_soft": "#eeffb9",
        "border": "#d8e6a0",
        "outline": "#bfce83",
        "text": "#1a2200",
        "muted": "#4a5a0d",
        "accent": "#8eff00",
        "accent_strong": "#60c900",
    },
    "Neon Pink": {
        "bg_start": "#ffe7f6",
        "bg_end": "#fff3fb",
        "surface": "#ffffff",
        "surface_soft": "#ffd6f0",
        "border": "#ebb5d9",
        "outline": "#db9cc7",
        "text": "#2a0f24",
        "muted": "#6c3a62",
        "accent": "#ff00a8",
        "accent_strong": "#d10087",
    },
    "Neon Cyan": {
        "bg_start": "#e9ffff",
        "bg_end": "#f3ffff",
        "surface": "#ffffff",
        "surface_soft": "#d6fcff",
        "border": "#b4e3e8",
        "outline": "#9bcfd5",
        "text": "#05252a",
        "muted": "#2d5a60",
        "accent": "#00e5ff",
        "accent_strong": "#00b8cc",
    },
    "Midnight Aurora": {
        "color_scheme": "dark",
        "bg_start": "#090d1a",
        "bg_end": "#141a2f",
        "surface": "#171f36",
        "surface_soft": "#212b47",
        "border": "#334166",
        "outline": "#44547d",
        "text": "#eaf0ff",
        "muted": "#afbadb",
        "accent": "#5aa9ff",
        "accent_strong": "#2d8dff",
        "placeholder": "#96a5cf",
        "disabled_text": "#9aa9d1",
    },
    "Ubuntu": {
        "color_scheme": "dark",
        "bg_start": "#140f1b",
        "bg_end": "#25162d",
        "surface": "#2a1b33",
        "surface_soft": "#372344",
        "border": "#59406a",
        "outline": "#715186",
        "text": "#f9edff",
        "muted": "#d0bbdd",
        "accent": "#ff7a5c",
        "accent_strong": "#ff5530",
        "placeholder": "#c0a5cf",
        "disabled_text": "#c7afd3",
    },
    "E-Ink": {
        "bg_start": "#f7f7f7",
        "bg_end": "#ffffff",
        "surface": "#ffffff",
        "surface_soft": "#f1f1f1",
        "border": "#b5b5b5",
        "outline": "#9d9d9d",
        "text": "#111111",
        "muted": "#333333",
        "accent": "#1a1a1a",
        "accent_strong": "#000000",
    },
    "Solaris Punch": {
        "bg_start": "#fff1d9",
        "bg_end": "#fff9ee",
        "surface": "#ffffff",
        "surface_soft": "#ffe9c7",
        "border": "#e7c28f",
        "outline": "#d8ae74",
        "text": "#2d1d08",
        "muted": "#70563a",
        "accent": "#ff7a00",
        "accent_strong": "#d85f00",
    },
    "Lagoon Pop": {
        "bg_start": "#e6fff9",
        "bg_end": "#f4fffd",
        "surface": "#ffffff",
        "surface_soft": "#d7fff2",
        "border": "#a6e7d3",
        "outline": "#8cd6c0",
        "text": "#0f2923",
        "muted": "#3c6158",
        "accent": "#00b894",
        "accent_strong": "#009976",
    },
    "Aurora Bloom": {
        "bg_start": "#f2ecff",
        "bg_end": "#fbf8ff",
        "surface": "#ffffff",
        "surface_soft": "#ece2ff",
        "border": "#cdb8f0",
        "outline": "#b59de2",
        "text": "#23173b",
        "muted": "#5b4a7d",
        "accent": "#a03dff",
        "accent_strong": "#7f2fd0",
    },
    "Nebula Circuit": {
        "color_scheme": "dark",
        "bg_start": "#060b16",
        "bg_end": "#101a30",
        "surface": "#15213b",
        "surface_soft": "#1d2d4f",
        "border": "#2c4474",
        "outline": "#3a588f",
        "text": "#e8f1ff",
        "muted": "#a9bbdb",
        "accent": "#00d4ff",
        "accent_strong": "#00a7cc",
        "placeholder": "#95b5d2",
        "disabled_text": "#8ea4c5",
    },
    "Crimson Void": {
        "color_scheme": "dark",
        "bg_start": "#13070d",
        "bg_end": "#25111a",
        "surface": "#2d1622",
        "surface_soft": "#3a1f2d",
        "border": "#5a3043",
        "outline": "#734058",
        "text": "#ffecf4",
        "muted": "#d6b3c3",
        "accent": "#ff3f81",
        "accent_strong": "#d82c68",
        "placeholder": "#c89ab0",
        "disabled_text": "#c8a8b7",
    },
    "Cobalt Sunrise": {
        "bg_start": "#e8f2ff",
        "bg_end": "#fff5e8",
        "surface": "#ffffff",
        "surface_soft": "#edf3ff",
        "border": "#c6d3e6",
        "outline": "#b3c3d8",
        "text": "#1b1f2a",
        "muted": "#4e5b74",
        "accent": "#2563eb",
        "accent_strong": "#1d4ed8",
    },
    "Mojave Bloom": {
        "bg_start": "#fff1de",
        "bg_end": "#fffaf0",
        "surface": "#ffffff",
        "surface_soft": "#ffe9cc",
        "border": "#e7c08d",
        "outline": "#d5ad76",
        "text": "#2b1d0f",
        "muted": "#6f563a",
        "accent": "#d97706",
        "accent_strong": "#b45309",
    },
    "Electric Tropic": {
        "bg_start": "#e8fff8",
        "bg_end": "#f8f6ff",
        "surface": "#ffffff",
        "surface_soft": "#e3fbf4",
        "border": "#b9e6d9",
        "outline": "#9fd3c6",
        "text": "#102621",
        "muted": "#3f625a",
        "accent": "#008c7a",
        "accent_strong": "#00695c",
    },
    "Blacklight Arcade": {
        "color_scheme": "dark",
        "bg_start": "#0b0716",
        "bg_end": "#1a1233",
        "surface": "#1c1538",
        "surface_soft": "#2a1f4d",
        "border": "#4b3b78",
        "outline": "#605093",
        "text": "#f3ecff",
        "muted": "#c8b9ea",
        "accent": "#b55dff",
        "accent_strong": "#8d35f2",
        "placeholder": "#ad9ad8",
        "disabled_text": "#b6a7de",
    },
    "Abyss Voltage": {
        "color_scheme": "dark",
        "bg_start": "#04121b",
        "bg_end": "#0d2333",
        "surface": "#122d40",
        "surface_soft": "#1a3a52",
        "border": "#2b5575",
        "outline": "#3a6d91",
        "text": "#e8f7ff",
        "muted": "#b5d5e8",
        "accent": "#00c2ff",
        "accent_strong": "#0097cf",
        "placeholder": "#8eb6cd",
        "disabled_text": "#9bc1d7",
    },
}
THEME_NAME_ALIASES: dict[str, str] = {
    "Obsidian Ember": "Ubuntu",
    "Obsidion Ember": "Ubuntu",
}
WORKBOOK_PATH_STATE_KEY = "state::workbook_path"
BACKUP_DIR_STATE_KEY = "state::backup_dir"
BACKUP_KEEP_LATEST_STATE_KEY = "state::backup_keep_latest"
SHEET_SELECT_STATE_KEY = "state::selected_sheet_name"
SHEET_SELECT_WORKBOOK_STATE_KEY = "state::selected_sheet_name_workbook"
SETTINGS_TAB_STATE_KEY = "state::settings_tab_selector"
SETTINGS_TAB_PENDING_STATE_KEY = "state::settings_tab_selector_pending"
THEME_STATE_KEY = "state::theme_name"
ENTRY_FORM_RESET_KEY_PREFIX = "state::entry_form_reset"
UPDATE_MANIFEST_URL_STATE_KEY = "state::update_manifest_url"
OPEN_SETTINGS_ONCE_STATE_KEY = "state::open_settings_once"
DEFAULT_UPDATE_MANIFEST_URL = ""
DESKTOP_MODE_ENV_VAR = "POTROL_DESKTOP_MODE"
BROWSER_MODE_OVERRIDE_ENV_VAR = "POTROL_ALLOW_BROWSER_MODE"

SESSION_ID = uuid4().hex
SESSION_OWNER = (
    f"{os.environ.get('USERNAME', 'user')}@"
    f"{os.environ.get('COMPUTERNAME', socket.gethostname())}"
)
_PO_SEQUENCE_CACHE: dict[tuple[str, str, str, str], set[int]] = {}


def sanitize_headers(raw_headers: list[Any]) -> list[str]:
    headers: list[str] = []
    seen: set[str] = set()
    for index, value in enumerate(raw_headers, start=1):
        header = str(value).strip() if value is not None else ""
        if not header:
            header = f"Column {index}"

        base = header
        suffix = 2
        while header in seen:
            header = f"{base} ({suffix})"
            suffix += 1

        seen.add(header)
        headers.append(header)
    return headers


def normalize_cell_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if value is None:
        return ""
    return value


def normalize_editor_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime().strftime("%Y-%m-%d %H:%M")
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def has_non_empty_editor_value(value: Any) -> bool:
    normalized = normalize_editor_cell_value(value)
    if normalized is None:
        return False
    if isinstance(normalized, str):
        return bool(normalized.strip())
    return True


def header_is_id(header: str) -> bool:
    lowered = header.lower()
    return any(
        token in lowered
        for token in (" id", "id ", "po number", "po #", "po#", "record number")
    ) or lowered in {"id", "po"}


def header_is_timestamp(header: str) -> bool:
    lowered = header.lower()
    return any(token in lowered for token in ("created", "timestamp", "entered", "entry"))


def header_is_date_like(header: str) -> bool:
    lowered = header.lower()
    return "date" in lowered or header_is_timestamp(header)


def field_key(sheet_name: str, header: str, scope: str = "") -> str:
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", header).strip("_").lower()
    scope_token = str(scope).strip()
    if scope_token:
        return f"field::{scope_token}::{sheet_name}::{slug}"
    return f"field::{sheet_name}::{slug}"


def ensure_required_headers(headers: list[str]) -> list[str]:
    merged_headers = headers.copy()
    existing = {header.casefold() for header in headers}
    for required_header in REQUIRED_HEADERS:
        if required_header.casefold() not in existing:
            merged_headers.append(required_header)
    return merged_headers


def normalize_location_code(value: str) -> str:
    return re.sub(r"[^A-Z0-9_-]", "", value.strip().upper())


def build_location_alias_lookup(location_options: list[str] | None = None) -> dict[str, str]:
    alias_lookup: dict[str, str] = {}
    for raw_code in (location_options or []):
        normalized_code = normalize_location_code(str(raw_code))
        if normalized_code:
            alias_lookup[normalized_code] = normalized_code
    for default_code in DEFAULT_LOCATION_OPTIONS:
        normalized_default = normalize_location_code(default_code)
        if normalized_default and normalized_default not in alias_lookup:
            alias_lookup[normalized_default] = normalized_default

    for alias_token, target_token in LOCATION_ALIAS_MAP.items():
        normalized_alias = normalize_location_code(alias_token)
        normalized_target = normalize_location_code(target_token)
        if not normalized_alias or not normalized_target:
            continue
        alias_lookup[normalized_alias] = alias_lookup.get(normalized_target, normalized_target)
    return alias_lookup


def extract_location_code(
    raw_location_value: Any,
    raw_department_loc_value: Any,
    location_options: list[str] | None = None,
) -> str:
    alias_lookup = build_location_alias_lookup(location_options)

    def parse_location_text(raw_value: Any) -> str:
        text = str(raw_value).strip()
        if not text:
            return ""
        normalized_text = normalize_location_code(text)
        if normalized_text and normalized_text in alias_lookup:
            return alias_lookup[normalized_text]

        tokens = [
            normalize_location_code(token)
            for token in re.split(r"[^A-Za-z0-9]+", text.upper())
            if str(token).strip()
        ]
        for token in tokens:
            if token in alias_lookup:
                return alias_lookup[token]
        return ""

    location_code = parse_location_text(raw_location_value)
    if location_code:
        return location_code
    return parse_location_text(raw_department_loc_value)


def validate_workbook_input(raw_value: str, resolved_path: Path) -> str | None:
    workbook_text = str(raw_value).strip()
    if not workbook_text:
        return "Workbook path cannot be blank."

    lowered_text = workbook_text.lower()
    if lowered_text.startswith("http://") or lowered_text.startswith("https://"):
        return (
            "Web links are not supported as workbook paths. "
            "Use a local or network Excel file path instead."
        )

    suffix = resolved_path.suffix.lower()
    if suffix == ".url":
        return (
            "Internet shortcuts (.url) are not supported. "
            "Select the actual Excel workbook from a local or network folder "
            "(.xlsx/.xlsm/.xltx/.xltm)."
        )

    if suffix and suffix not in SUPPORTED_WORKBOOK_EXTENSIONS:
        allowed_text = ", ".join(sorted(SUPPORTED_WORKBOOK_EXTENSIONS))
        return f"Unsupported workbook file type `{suffix}`. Use one of: {allowed_text}."

    if suffix == "":
        return (
            "Workbook path must include a workbook filename and extension "
            "(.xlsx/.xlsm/.xltx/.xltm)."
        )

    if resolved_path.exists() and resolved_path.is_dir():
        return "Workbook path points to a folder. Select an Excel workbook file instead."

    return None


def path_key(path: Path) -> str:
    try:
        return str(path.expanduser().resolve()).casefold()
    except Exception:
        return str(path).casefold()


def is_network_path(path: Path) -> bool:
    text = str(path).strip()
    if text.startswith("\\\\") or text.startswith("//"):
        return True

    if os.name != "nt":
        return False

    drive = Path(text).drive
    if not drive:
        return False

    try:
        import ctypes

        drive_root = f"{drive}\\"
        drive_type = ctypes.windll.kernel32.GetDriveTypeW(drive_root)
        # DRIVE_REMOTE == 4
        return int(drive_type) == 4
    except Exception:
        return False


def copy_file_with_retry(
    source: Path,
    destination: Path,
    retries: int = 3,
    delay_seconds: float = 0.2,
) -> None:
    attempts = max(1, int(retries))
    last_error: Exception | None = None
    for attempt in range(attempts):
        try:
            destination.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, destination)
            return
        except Exception as exc:
            last_error = exc
            if attempt < (attempts - 1):
                time.sleep(max(0.01, float(delay_seconds)))
    if last_error is not None:
        raise last_error


def open_workbook_with_retry(
    path: Path | str,
    *,
    read_only: bool,
    data_only: bool,
) -> Any:
    workbook_path = Path(path).expanduser()
    attempts = max(1, int(WORKBOOK_OPEN_RETRY_COUNT))
    last_error: Exception | None = None

    for attempt in range(attempts):
        try:
            return load_workbook(
                str(workbook_path),
                read_only=read_only,
                data_only=data_only,
            )
        except InvalidFileException:
            raise
        except Exception as exc:
            last_error = exc
            if attempt < (attempts - 1):
                time.sleep(max(0.05, float(WORKBOOK_OPEN_RETRY_DELAY_SECONDS)))

    if last_error is not None:
        raise last_error
    raise FileNotFoundError(f"Could not open workbook: {workbook_path}")


def hex_to_rgb_triplet(value: str, fallback: str = "11, 87, 208") -> str:
    text = str(value).strip().lstrip("#")
    if len(text) == 3:
        text = "".join(char * 2 for char in text)
    if len(text) != 6 or any(char not in "0123456789abcdefABCDEF" for char in text):
        return fallback

    red = int(text[0:2], 16)
    green = int(text[2:4], 16)
    blue = int(text[4:6], 16)
    return f"{red}, {green}, {blue}"


def contrast_text_color(background_hex: str, light: str = "#ffffff", dark: str = "#111111") -> str:
    text = str(background_hex).strip().lstrip("#")
    if len(text) == 3:
        text = "".join(char * 2 for char in text)
    if len(text) != 6 or any(char not in "0123456789abcdefABCDEF" for char in text):
        return light
    red = int(text[0:2], 16)
    green = int(text[2:4], 16)
    blue = int(text[4:6], 16)
    luminance = (0.2126 * red + 0.7152 * green + 0.0722 * blue) / 255.0
    return dark if luminance >= 0.62 else light


def hex_luminance(value: str, fallback: float = 0.5) -> float:
    text = str(value).strip().lstrip("#")
    if len(text) == 3:
        text = "".join(char * 2 for char in text)
    if len(text) != 6 or any(char not in "0123456789abcdefABCDEF" for char in text):
        return fallback
    red = int(text[0:2], 16)
    green = int(text[2:4], 16)
    blue = int(text[4:6], 16)
    return (0.2126 * red + 0.7152 * green + 0.0722 * blue) / 255.0


def resolve_theme_palette(theme_name: str) -> dict[str, str]:
    fallback_light = dict(THEME_PRESETS.get("Sky", {}))
    fallback_dark = dict(THEME_PRESETS.get("Midnight Aurora", {}))
    raw_theme = dict(THEME_PRESETS.get(theme_name, fallback_light))
    raw_scheme = str(raw_theme.get("color_scheme", "light")).strip().lower()
    scheme = "dark" if raw_scheme == "dark" else "light"

    resolved = dict(fallback_dark if scheme == "dark" else fallback_light)
    resolved.update(raw_theme)
    resolved["color_scheme"] = scheme

    if scheme == "light":
        if hex_luminance(resolved.get("bg_start", ""), 1.0) < 0.58:
            resolved["bg_start"] = fallback_light.get("bg_start", "#f3f6ff")
        if hex_luminance(resolved.get("bg_end", ""), 1.0) < 0.60:
            resolved["bg_end"] = fallback_light.get("bg_end", "#f7f9ff")
        if hex_luminance(resolved.get("surface", ""), 1.0) < 0.68:
            resolved["surface"] = "#ffffff"
        if hex_luminance(resolved.get("surface_soft", ""), 1.0) < 0.58:
            resolved["surface_soft"] = fallback_light.get("surface_soft", "#eef2ff")
        if hex_luminance(resolved.get("text", ""), 0.0) > 0.46:
            resolved["text"] = fallback_light.get("text", "#1b1b1f")
        if hex_luminance(resolved.get("muted", ""), 0.0) > 0.58:
            resolved["muted"] = fallback_light.get("muted", "#47464f")
    else:
        if hex_luminance(resolved.get("bg_start", ""), 0.0) > 0.30:
            resolved["bg_start"] = fallback_dark.get("bg_start", "#090d1a")
        if hex_luminance(resolved.get("bg_end", ""), 0.0) > 0.34:
            resolved["bg_end"] = fallback_dark.get("bg_end", "#141a2f")
        if hex_luminance(resolved.get("surface", ""), 0.0) > 0.42:
            resolved["surface"] = fallback_dark.get("surface", "#171f36")
        if hex_luminance(resolved.get("surface_soft", ""), 0.0) > 0.48:
            resolved["surface_soft"] = fallback_dark.get("surface_soft", "#212b47")
        if hex_luminance(resolved.get("text", ""), 1.0) < 0.64:
            resolved["text"] = fallback_dark.get("text", "#eaf0ff")
        if hex_luminance(resolved.get("muted", ""), 1.0) < 0.45:
            resolved["muted"] = fallback_dark.get("muted", "#afbadb")

    resolved["outline"] = str(resolved.get("outline", resolved.get("border", "#d2d9e8")))
    if not str(resolved.get("placeholder", "")).strip():
        resolved["placeholder"] = str(resolved.get("muted", "#6e7280"))
    if not str(resolved.get("disabled_text", "")).strip():
        resolved["disabled_text"] = str(resolved.get("muted", "#6e7280"))
    return {key: str(value) for key, value in resolved.items()}


def get_query_param_text(name: str) -> str:
    try:
        raw_value: Any = st.query_params.get(name, "")
    except Exception:
        return ""
    if isinstance(raw_value, list):
        if not raw_value:
            return ""
        raw_value = raw_value[0]
    return str(raw_value).strip()


def clear_query_param(name: str) -> None:
    try:
        if name in st.query_params:
            del st.query_params[name]
    except Exception:
        pass


def canonical_theme_name(theme_name: str) -> str:
    raw_theme_name = str(theme_name).strip()
    if not raw_theme_name:
        return ""

    if raw_theme_name in THEME_NAME_ALIASES:
        return THEME_NAME_ALIASES[raw_theme_name]
    if raw_theme_name in THEME_PRESETS:
        return raw_theme_name

    normalized_name = raw_theme_name.casefold()
    for old_name, mapped_name in THEME_NAME_ALIASES.items():
        if old_name.casefold() == normalized_name:
            return mapped_name

    for preset_name in THEME_PRESETS:
        if preset_name.casefold() == normalized_name:
            return preset_name

    return raw_theme_name


def normalize_backup_keep_latest(value: Any, default: int = DEFAULT_BACKUP_KEEP_LATEST) -> int:
    try:
        parsed = int(str(value).strip())
    except Exception:
        parsed = int(default)
    parsed = max(MIN_BACKUP_KEEP_LATEST, parsed)
    parsed = min(MAX_BACKUP_KEEP_LATEST, parsed)
    return parsed


def append_runtime_log(level: str, context: str, message: str) -> None:
    level_text = str(level).strip().upper() or "INFO"
    context_text = str(context).strip() or "runtime"
    message_text = str(message).replace("\r", " ").replace("\n", " ").strip()
    if not message_text:
        message_text = "(no details)"

    line = f"{datetime.now().isoformat(timespec='seconds')} [{level_text}] {context_text} :: {message_text}\n"

    try:
        APP_RUNTIME_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        with APP_RUNTIME_LOG_PATH.open("a", encoding="utf-8") as handle:
            handle.write(line)
    except Exception:
        return

    try:
        lines = APP_RUNTIME_LOG_PATH.read_text(encoding="utf-8", errors="replace").splitlines()
        if len(lines) > MAX_RUNTIME_LOG_LINES:
            trimmed_lines = lines[-MAX_RUNTIME_LOG_LINES:]
            APP_RUNTIME_LOG_PATH.write_text("\n".join(trimmed_lines) + "\n", encoding="utf-8")
    except Exception:
        pass


def log_runtime_error(context: str, exc: Exception) -> None:
    error_summary = f"{exc.__class__.__name__}: {exc}"
    trace_text = traceback.format_exc().strip()
    if trace_text and trace_text != "NoneType: None":
        error_summary = f"{error_summary} | traceback={trace_text}"
    append_runtime_log("ERROR", context, error_summary)


def read_runtime_log_tail(max_lines: int = 120) -> list[str]:
    bounded_lines = max(1, int(max_lines))
    try:
        if not APP_RUNTIME_LOG_PATH.exists():
            return []
        log_lines = APP_RUNTIME_LOG_PATH.read_text(encoding="utf-8", errors="replace").splitlines()
        return log_lines[-bounded_lines:]
    except Exception:
        return []


def get_runtime_log_line_count() -> int:
    try:
        if not APP_RUNTIME_LOG_PATH.exists():
            return 0
        return len(APP_RUNTIME_LOG_PATH.read_text(encoding="utf-8", errors="replace").splitlines())
    except Exception:
        return 0


def clear_runtime_log() -> None:
    try:
        APP_RUNTIME_LOG_PATH.unlink(missing_ok=True)
    except Exception:
        pass


def render_logo_image(path: Path, width: int = 360, palette: dict[str, str] | None = None) -> None:
    if not path.exists():
        return

    suffix = path.suffix.lower()
    if suffix == ".svg":
        mime_type = "image/svg+xml"
    elif suffix in {".jpg", ".jpeg"}:
        mime_type = "image/jpeg"
    else:
        mime_type = "image/png"

    try:
        raw_bytes = path.read_bytes()
        if suffix == ".svg" and palette is not None:
            svg_text = raw_bytes.decode("utf-8")
            replacement_map = {
                "#0b67c2": palette.get("accent", "#0b67c2"),
                "#14b8a6": palette.get("accent_strong", "#14b8a6"),
                "#0f172a": palette.get("text", "#0f172a"),
                "#476581": palette.get("muted", "#476581"),
                "#8ba4bf": palette.get("border", "#8ba4bf"),
                "#d9e4f2": palette.get("surface_soft", "#d9e4f2"),
                "#f8fafc": palette.get("surface", "#f8fafc"),
            }
            for source_color, target_color in replacement_map.items():
                svg_text = re.sub(
                    re.escape(source_color),
                    target_color,
                    svg_text,
                    flags=re.IGNORECASE,
                )
            raw_bytes = svg_text.encode("utf-8")

        encoded = base64.b64encode(raw_bytes).decode("ascii")
        st.markdown(
            f"""
            <div class="potrol-logo-wrap">
                <img src="data:{mime_type};base64,{encoded}" alt="POtrol" style="width:{int(width)}px; max-width:100%; height:auto; display:block;" />
            </div>
            """,
            unsafe_allow_html=True,
        )
    except Exception:
        st.image(str(path), use_container_width=False, width=width)


def load_location_options() -> list[str]:
    if LOCATION_CONFIG_PATH.exists():
        try:
            saved = json.loads(LOCATION_CONFIG_PATH.read_text(encoding="utf-8"))
            if isinstance(saved, list):
                cleaned = [normalize_location_code(str(item)) for item in saved]
                options = [code for code in cleaned if code]
                if options:
                    return sorted(set(options))
        except Exception:
            pass
    return DEFAULT_LOCATION_OPTIONS.copy()


def save_location_options(options: list[str]) -> None:
    cleaned = sorted(set(normalize_location_code(option) for option in options if option))
    LOCATION_CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
    LOCATION_CONFIG_PATH.write_text(json.dumps(cleaned, indent=2), encoding="utf-8")


def load_app_settings() -> dict[str, Any]:
    if APP_SETTINGS_PATH.exists():
        try:
            saved = json.loads(APP_SETTINGS_PATH.read_text(encoding="utf-8"))
            if isinstance(saved, dict):
                workbook_path = str(saved.get("workbook_path", "")).strip()
                backup_dir = str(saved.get("backup_dir", "")).strip()
                theme_name = canonical_theme_name(str(saved.get("theme", "")).strip())
                update_manifest_url = str(saved.get("update_manifest_url", "")).strip()
                backup_keep_latest = normalize_backup_keep_latest(
                    saved.get("backup_keep_latest", DEFAULT_BACKUP_KEEP_LATEST)
                )
                settings: dict[str, Any] = {}
                if workbook_path:
                    settings["workbook_path"] = workbook_path
                if backup_dir:
                    settings["backup_dir"] = backup_dir
                if theme_name and theme_name in THEME_PRESETS:
                    settings["theme"] = theme_name
                if update_manifest_url:
                    settings["update_manifest_url"] = update_manifest_url
                settings["backup_keep_latest"] = backup_keep_latest
                return settings
        except Exception:
            pass
    return {}


def save_app_settings(
    workbook_path: str,
    backup_dir: str,
    theme: str | None = None,
    update_manifest_url: str | None = None,
    backup_keep_latest: int | None = None,
) -> None:
    existing_settings = load_app_settings()
    if theme is None:
        existing_theme = canonical_theme_name(str(existing_settings.get("theme", DEFAULT_THEME_NAME)).strip())
        theme_name = existing_theme if existing_theme in THEME_PRESETS else DEFAULT_THEME_NAME
    else:
        requested_theme = canonical_theme_name(str(theme).strip())
        theme_name = requested_theme if requested_theme in THEME_PRESETS else DEFAULT_THEME_NAME

    if update_manifest_url is None:
        manifest_value = str(existing_settings.get("update_manifest_url", DEFAULT_UPDATE_MANIFEST_URL)).strip()
    else:
        manifest_value = str(update_manifest_url).strip()

    if backup_keep_latest is None:
        keep_latest_value = normalize_backup_keep_latest(
            existing_settings.get("backup_keep_latest", DEFAULT_BACKUP_KEEP_LATEST)
        )
    else:
        keep_latest_value = normalize_backup_keep_latest(backup_keep_latest)

    payload = {
        "workbook_path": str(workbook_path).strip(),
        "backup_dir": str(backup_dir).strip(),
        "theme": theme_name,
        "update_manifest_url": manifest_value,
        "backup_keep_latest": keep_latest_value,
    }
    APP_SETTINGS_PATH.parent.mkdir(parents=True, exist_ok=True)
    APP_SETTINGS_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def load_json_dict(path: Path) -> dict[str, Any]:
    backup_path = path.with_suffix(f"{path.suffix}.bak")
    candidate_paths = [path, backup_path]

    for index, candidate in enumerate(candidate_paths):
        if not candidate.exists():
            continue
        try:
            loaded = json.loads(candidate.read_text(encoding="utf-8"))
            if isinstance(loaded, dict):
                # Self-heal from backup if the primary file is corrupt.
                if index == 1 and candidate != path:
                    try:
                        write_json_dict_atomic(path, loaded, keep_backup=False)
                    except Exception:
                        pass
                return loaded
        except Exception:
            continue
    return {}


def write_json_dict_atomic(
    path: Path,
    payload: dict[str, Any],
    *,
    keep_backup: bool = True,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_fd, temp_name = tempfile.mkstemp(
        prefix=f"{path.name}.",
        suffix=".tmp",
        dir=str(path.parent),
    )
    try:
        with os.fdopen(temp_fd, "w", encoding="utf-8") as temp_file:
            json.dump(payload, temp_file, indent=2)
            temp_file.flush()
            os.fsync(temp_file.fileno())
        if keep_backup and path.exists():
            backup_path = path.with_suffix(f"{path.suffix}.bak")
            try:
                copy_file_with_retry(path, backup_path, retries=2, delay_seconds=0.08)
            except Exception:
                pass
        os.replace(temp_name, path)
    finally:
        try:
            Path(temp_name).unlink(missing_ok=True)
        except Exception:
            pass


def build_draft_key(workbook_path: Path, sheet_name: str) -> str:
    normalized_path = path_key(workbook_path.expanduser())
    normalized_sheet = str(sheet_name).strip().casefold()
    return f"{normalized_path}::{normalized_sheet}"


def sanitize_draft_payload(payload: dict[str, Any]) -> dict[str, Any]:
    line_items_raw = payload.get("line_items", [])
    line_items = ensure_line_item_rows(line_items_raw)
    return {
        "vendor": str(payload.get("vendor", "")).strip(),
        "department": str(payload.get("department", "IT")).strip(),
        "location": str(payload.get("location", "")).strip(),
        "line_items": line_items,
        "shipping_cost": round(parse_float(payload.get("shipping_cost", 0.0), 0.0), 2),
        "sales_tax": round(parse_float(payload.get("sales_tax", 0.0), 0.0), 2),
        "purchase_reason": str(payload.get("purchase_reason", "")).strip(),
        "saved_at_ts": float(payload.get("saved_at_ts", time.time()) or time.time()),
    }


def load_entry_draft(workbook_path: Path, sheet_name: str) -> dict[str, Any] | None:
    draft_store = load_json_dict(APP_DRAFTS_PATH)
    draft_key = build_draft_key(workbook_path, sheet_name)
    draft_value = draft_store.get(draft_key)
    if isinstance(draft_value, dict):
        try:
            return sanitize_draft_payload(draft_value)
        except Exception:
            return None
    return None


def save_entry_draft(workbook_path: Path, sheet_name: str, payload: dict[str, Any]) -> None:
    draft_store = load_json_dict(APP_DRAFTS_PATH)
    draft_key = build_draft_key(workbook_path, sheet_name)
    draft_store[draft_key] = sanitize_draft_payload(payload)
    write_json_dict_atomic(APP_DRAFTS_PATH, draft_store)


def clear_entry_draft(workbook_path: Path, sheet_name: str) -> None:
    if not APP_DRAFTS_PATH.exists():
        return
    draft_store = load_json_dict(APP_DRAFTS_PATH)
    draft_key = build_draft_key(workbook_path, sheet_name)
    if draft_key not in draft_store:
        return
    draft_store.pop(draft_key, None)
    write_json_dict_atomic(APP_DRAFTS_PATH, draft_store)


def draft_payload_hash(payload: dict[str, Any]) -> str:
    encoded = json.dumps(payload, sort_keys=True, separators=(",", ":"), ensure_ascii=True)
    return hashlib.sha1(encoded.encode("utf-8")).hexdigest()


def parse_version_key(value: str) -> tuple[int, ...]:
    tokens = re.findall(r"\d+", str(value))
    if not tokens:
        return (0,)
    return tuple(int(token) for token in tokens)


def is_version_newer(candidate: str, current: str) -> bool:
    return parse_version_key(candidate) > parse_version_key(current)


def fetch_update_manifest(url: str) -> dict[str, str]:
    target_url = str(url).strip()
    if not target_url:
        raise ValueError("Update manifest URL is blank.")

    with urlopen(target_url, timeout=8) as response:
        charset = response.headers.get_content_charset() or "utf-8"
        body_text = response.read().decode(charset, errors="replace")

    parsed = json.loads(body_text)
    if not isinstance(parsed, dict):
        raise ValueError("Update manifest must be a JSON object.")

    version_value = str(parsed.get("version", "")).strip()
    if not version_value:
        raise ValueError("Update manifest is missing `version`.")

    download_url = str(parsed.get("download_url", "")).strip()
    notes = str(parsed.get("notes", "")).strip()
    return {
        "version": version_value,
        "download_url": download_url,
        "notes": notes,
    }


def get_workbook_signature(path: Path) -> str:
    if not path.exists():
        return ""
    try:
        stat_result = path.stat()
        return f"{stat_result.st_mtime_ns}:{stat_result.st_size}"
    except Exception:
        return ""


def normalize_header_token(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.casefold())


def find_first_header(headers: list[str], aliases: list[str]) -> str | None:
    alias_tokens = {normalize_header_token(alias) for alias in aliases}
    for header in headers:
        if normalize_header_token(str(header)) in alias_tokens:
            return header
    return None


def find_po_column_index(headers: list[str]) -> int | None:
    aliases = {"ponumber", "po", "po#"}
    for index, header in enumerate(headers):
        token = normalize_header_token(str(header))
        if token in aliases:
            return index
    return None


def find_po_column_indexes(worksheet: Any, header_rows_to_scan: int = 25) -> list[int]:
    aliases = {"ponumber", "po", "po#"}
    max_rows = max(1, min(int(header_rows_to_scan), int(worksheet.max_row or 1)))
    max_columns = max(1, int(worksheet.max_column or 1))

    indexes: set[int] = set()
    for row in worksheet.iter_rows(
        min_row=1,
        max_row=max_rows,
        min_col=1,
        max_col=max_columns,
        values_only=True,
    ):
        for column_index, value in enumerate(row):
            if value is None:
                continue
            token = normalize_header_token(str(value))
            if token in aliases:
                indexes.add(column_index)

    # Legacy sheets often keep PO values in column A without a clean header row.
    indexes.add(0)
    return sorted(indexes)


def worksheet_effective_max_row(worksheet: Any) -> int:
    fallback = max(1, int(worksheet.max_row or 1))
    try:
        dimension_text = str(worksheet.calculate_dimension())
        match = re.search(r":[$A-Z]+(\d+)$", dimension_text)
        if not match:
            return fallback
        parsed = int(match.group(1))
        if parsed <= 0:
            return fallback
        return min(parsed, fallback)
    except Exception:
        return fallback


def build_entry_schema(headers: list[str]) -> tuple[str, dict[str, str], list[str]]:
    po_header = find_first_header(headers, ["PO#", "PO #", "PO Number", "PO"])
    date_header = find_first_header(headers, ["Date"])
    vendor_header = find_first_header(headers, ["Vendor", "Vendor/Store"])
    department_header = find_first_header(headers, ["Department", "Deparment"])
    location_header = find_first_header(headers, ["Location", "Loc"])
    dept_or_loc_header = find_first_header(
        headers,
        [
            "Deparment/Loc",
            "Department/Loc",
            "Deparment",
            "Department",
            "Location",
            "Loc",
        ],
    )
    item_header = find_first_header(headers, ["Item", "Items Being Purchased", "Items"])
    price_header = find_first_header(headers, ["Price", "Price Per Item"])
    qty_header = find_first_header(headers, ["QTY", "Quantity"])
    sub_total_header = find_first_header(headers, ["Sub Total", "Subtotal"])
    grand_total_header = find_first_header(headers, ["Grand Total", "GrandTotal"])

    if all([po_header, item_header, price_header, qty_header, sub_total_header, grand_total_header]):
        if dept_or_loc_header:
            department_target = dept_or_loc_header
            location_target = dept_or_loc_header
        else:
            department_target = department_header
            location_target = location_header
        mapping: dict[str, str] = {
            "PO Number": po_header,
            "Date": date_header or "Date",
            "Vendor/Store": vendor_header or "Vendor",
            "Department": department_target or "Department",
            "Location": location_target or "Location",
            "Items Being Purchased": item_header,
            "Price Per Item": price_header,
            "Quantity": qty_header,
            "Sub Total": sub_total_header,
            "Grand Total": grand_total_header,
        }

        ordered_headers: list[str] = []
        seen: set[str] = set()
        for header in [
            po_header,
            date_header,
            vendor_header,
            department_target,
            location_target,
            item_header,
            price_header,
            qty_header,
            sub_total_header,
            grand_total_header,
        ]:
            if not header:
                continue
            key = header.casefold()
            if key in seen:
                continue
            seen.add(key)
            ordered_headers.append(header)
        return "legacy", mapping, ordered_headers

    return "default", {}, DEFAULT_HEADERS.copy()


def parse_po_number(value: Any, prefix: str) -> int | None:
    if value is None:
        return None

    normalized = str(value).strip().upper()
    pattern = rf"^\s*'?{re.escape(prefix.upper())}\s*[-_/]?\s*0*(\d+)\s*$"
    match = re.match(pattern, normalized)
    if not match:
        return None
    return int(match.group(1))


def collect_po_sequences(path: Path, prefix: str = PO_PREFIX, sheet_name: str | None = None) -> set[int]:
    if not path.exists():
        return set()

    workbook_signature = get_workbook_signature(path)
    cache_key = (
        path_key(path),
        workbook_signature,
        str(prefix).strip().upper(),
        str(sheet_name or "__all__").strip().casefold(),
    )
    cached_sequences = _PO_SEQUENCE_CACHE.get(cache_key)
    if cached_sequences is not None:
        return set(cached_sequences)

    try:
        workbook = open_workbook_with_retry(path, read_only=True, data_only=True)
    except Exception:
        return set()

    sequences: set[int] = set()
    try:
        if sheet_name and sheet_name in workbook.sheetnames:
            sheet_names = [sheet_name]
        else:
            sheet_names = workbook.sheetnames

        for current_sheet_name in sheet_names:
            worksheet = workbook[current_sheet_name]
            po_column_indexes = find_po_column_indexes(worksheet)
            effective_max_row = worksheet_effective_max_row(worksheet)
            for column_index in po_column_indexes:
                excel_col = int(column_index) + 1
                empty_streak = 0
                sequence_found_in_column = False
                rows_scanned = 0
                for row in worksheet.iter_rows(
                    min_row=1,
                    max_row=effective_max_row,
                    min_col=excel_col,
                    max_col=excel_col,
                    values_only=True,
                ):
                    rows_scanned += 1
                    cell_value = row[0] if row else None
                    sequence = parse_po_number(cell_value, prefix=prefix)
                    if sequence is not None:
                        sequences.add(sequence)
                        sequence_found_in_column = True
                        empty_streak = 0
                    else:
                        text_value = str(cell_value).strip() if cell_value is not None else ""
                        if text_value:
                            empty_streak = 0
                        else:
                            empty_streak += 1
                    if sequence_found_in_column and empty_streak >= PO_SCAN_EMPTY_STREAK_BREAK:
                        break
                    if rows_scanned >= PO_SCAN_HARD_ROW_LIMIT and empty_streak >= PO_SCAN_EMPTY_STREAK_BREAK:
                        break
    finally:
        workbook.close()

    if len(_PO_SEQUENCE_CACHE) >= PO_SEQUENCE_CACHE_MAX_KEYS:
        try:
            _PO_SEQUENCE_CACHE.pop(next(iter(_PO_SEQUENCE_CACHE)))
        except Exception:
            _PO_SEQUENCE_CACHE.clear()
    _PO_SEQUENCE_CACHE[cache_key] = set(sequences)
    return sequences


def po_number_exists(
    path: Path,
    po_number: str,
    prefix: str = PO_PREFIX,
    sheet_name: str | None = None,
) -> bool:
    sequence = parse_po_number(po_number, prefix=prefix)
    if sequence is None:
        return False
    return sequence in collect_po_sequences(path, prefix=prefix, sheet_name=sheet_name)


def get_next_po_number(
    path: Path,
    sheet_name: str | None = None,
    prefix: str = PO_PREFIX,
    starting_number: int = PO_START_NUMBER,
) -> str:
    if not path.exists():
        return f"{prefix}{starting_number}"

    sequences = collect_po_sequences(path, prefix=prefix, sheet_name=sheet_name)
    max_sequence = max(sequences) if sequences else None
    next_sequence = starting_number if max_sequence is None else max(max_sequence + 1, starting_number)
    return f"{prefix}{next_sequence}"


def get_po_reservation_path(path: Path) -> Path:
    return path.with_name(f".{path.name}.po_reservations.json")


def read_po_reservations(path: Path) -> dict[str, dict[str, Any]]:
    reservation_path = get_po_reservation_path(path)
    raw = load_json_dict(reservation_path)
    cleaned: dict[str, dict[str, Any]] = {}
    for session_id, value in raw.items():
        if not isinstance(session_id, str) or not isinstance(value, dict):
            continue
        cleaned[session_id] = value
    return cleaned


def write_po_reservations(path: Path, reservations: dict[str, dict[str, Any]]) -> None:
    reservation_path = get_po_reservation_path(path)
    write_json_dict_atomic(reservation_path, reservations)


def cleanup_po_reservations(
    reservations: dict[str, dict[str, Any]],
    existing_sequences: set[int],
    prefix: str,
    now_ts: float,
    stale_seconds: float,
) -> tuple[dict[str, dict[str, Any]], int | None]:
    max_reserved_sequence: int | None = None
    cleaned: dict[str, dict[str, Any]] = {}
    for session_id, entry in reservations.items():
        po_number = str(entry.get("po_number", "")).strip()
        sequence = parse_po_number(po_number, prefix=prefix)
        if sequence is None:
            continue

        updated_ts = float(entry.get("updated_ts", 0.0) or 0.0)
        if updated_ts <= 0:
            continue
        if now_ts - updated_ts > stale_seconds:
            continue
        if sequence in existing_sequences:
            continue

        owner = str(entry.get("owner", "")).strip()
        cleaned[session_id] = {
            "po_number": f"{prefix}{sequence}",
            "owner": owner,
            "updated_ts": updated_ts,
        }
        if max_reserved_sequence is None or sequence > max_reserved_sequence:
            max_reserved_sequence = sequence
    return cleaned, max_reserved_sequence


def reserve_session_po_number(
    path: Path,
    session_id: str,
    owner_label: str,
    sheet_name: str | None = None,
    prefix: str = PO_PREFIX,
    starting_number: int = PO_START_NUMBER,
    stale_seconds: float = PO_RESERVATION_STALE_SECONDS,
    lock_timeout_seconds: float = 2.2,
) -> str:
    if not path.exists():
        return f"{prefix}{starting_number}"

    with workbook_write_lock(path, timeout_seconds=lock_timeout_seconds):
        now_ts = time.time()
        existing_sequences = collect_po_sequences(path, prefix=prefix, sheet_name=sheet_name)
        raw_reservations = read_po_reservations(path)
        active_reservations, max_reserved_sequence = cleanup_po_reservations(
            reservations=raw_reservations,
            existing_sequences=existing_sequences,
            prefix=prefix,
            now_ts=now_ts,
            stale_seconds=stale_seconds,
        )

        session_entry = active_reservations.get(session_id, {})
        session_sequence = parse_po_number(session_entry.get("po_number"), prefix=prefix)
        max_existing = max(existing_sequences) if existing_sequences else starting_number - 1
        next_sequence = max(max_existing + 1, starting_number)
        if max_reserved_sequence is not None:
            next_sequence = max(next_sequence, max_reserved_sequence + 1)

        if session_sequence is None or session_sequence in existing_sequences:
            assigned_sequence = next_sequence
        else:
            assigned_sequence = session_sequence

        active_reservations[session_id] = {
            "po_number": f"{prefix}{assigned_sequence}",
            "owner": str(owner_label).strip(),
            "updated_ts": now_ts,
        }
        write_po_reservations(path, active_reservations)
        return f"{prefix}{assigned_sequence}"


def release_session_po_reservation(path: Path, session_id: str) -> None:
    if not path.exists():
        return
    with workbook_write_lock(path, timeout_seconds=2.2):
        reservations = read_po_reservations(path)
        if session_id not in reservations:
            return
        reservations.pop(session_id, None)
        write_po_reservations(path, reservations)


def get_active_po_reservation_count(path: Path, sheet_name: str | None = None) -> int:
    if not path.exists():
        return 0
    now_ts = time.time()
    existing_sequences = collect_po_sequences(path, prefix=PO_PREFIX, sheet_name=sheet_name)
    raw_reservations = read_po_reservations(path)
    active_reservations, _ = cleanup_po_reservations(
        reservations=raw_reservations,
        existing_sequences=existing_sequences,
        prefix=PO_PREFIX,
        now_ts=now_ts,
        stale_seconds=PO_RESERVATION_STALE_SECONDS,
    )
    return len(active_reservations)


def create_workbook(path: Path, sheet_name: str, headers: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(headers)
    workbook.save(path)


def get_sheet_names(path: Path) -> list[str]:
    workbook = open_workbook_with_retry(path, read_only=True, data_only=True)
    names = workbook.sheetnames
    workbook.close()
    return names


def choose_default_sheet_name(sheet_names: list[str], year: int) -> str:
    if not sheet_names:
        return DEFAULT_SHEET_NAME

    year_token = str(year)
    best_name = sheet_names[0]
    best_score = -1

    for name in sheet_names:
        lowered = name.casefold()
        score = 0

        if re.search(rf"\b{re.escape(year_token)}\b", lowered):
            score += 140
        elif year_token in lowered:
            score += 110

        if re.search(rf"{re.escape(year_token)}\s*[-_ ]*\s*totals?", lowered):
            score += 50

        if "totals" in lowered:
            score += 32
        elif "total" in lowered:
            score += 24

        if "grand total" in lowered:
            score += 20
        if "summary" in lowered:
            score += 8
        if "po" in lowered or "log" in lowered:
            score += 4

        if score > best_score:
            best_name = name
            best_score = score

    if best_score <= 0:
        for name in sheet_names:
            if name.casefold() == DEFAULT_SHEET_NAME.casefold():
                return name
        return sheet_names[0]

    return best_name


def browse_workbook_file(current_path: Path) -> Path | None:
    try:
        import tkinter as tk
        from tkinter import filedialog
    except Exception:
        return None

    root: Any = None
    try:
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        initial_dir = current_path.parent if current_path.parent.exists() else Path.home()
        selected = filedialog.askopenfilename(
            title="Select Workbook",
            initialdir=str(initial_dir),
            filetypes=[
                ("Excel Workbook", "*.xlsx *.xlsm *.xltx *.xltm"),
            ],
        )
        if selected:
            return Path(selected)
        return None
    finally:
        if root is not None:
            try:
                root.destroy()
            except Exception:
                pass


def browse_folder(current_path: Path) -> Path | None:
    try:
        import tkinter as tk
        from tkinter import filedialog
    except Exception:
        return None

    root: Any = None
    try:
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        initial_dir = current_path if current_path.exists() else Path.home()
        selected = filedialog.askdirectory(
            title="Select Folder",
            initialdir=str(initial_dir),
            mustexist=False,
        )
        if selected:
            return Path(selected)
        return None
    finally:
        if root is not None:
            try:
                root.destroy()
            except Exception:
                pass


@st.cache_data(show_spinner=False)
def load_sheet_data(path_str: str, sheet_name: str) -> tuple[list[str], list[dict[str, Any]], list[int]]:
    workbook = open_workbook_with_retry(Path(path_str), read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        return ensure_required_headers(DEFAULT_HEADERS.copy()), [], []

    worksheet = workbook[sheet_name]
    raw_headers = [worksheet.cell(row=1, column=column).value for column in range(1, worksheet.max_column + 1)]

    if not raw_headers or all(value is None or str(value).strip() == "" for value in raw_headers):
        headers = DEFAULT_HEADERS.copy()
    else:
        headers = sanitize_headers(raw_headers)
    headers = ensure_required_headers(headers)

    rows: list[dict[str, Any]] = []
    row_numbers: list[int] = []
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True),
        start=2,
    ):
        if row is None:
            continue
        if all(value is None or str(value).strip() == "" for value in row):
            continue

        record: dict[str, Any] = {}
        for index, header in enumerate(headers):
            value = row[index] if index < len(row) else None
            record[header] = normalize_cell_value(value)
        rows.append(record)
        row_numbers.append(row_index)

    workbook.close()
    return headers, rows, row_numbers


def build_reporting_frame_for_sheets(
    path_str: str,
    target_sheet_names: list[str] | tuple[str, ...],
    location_options: list[str] | None = None,
) -> pd.DataFrame:
    report_columns = [
        "PO Number",
        "Date",
        "Vendor/Store",
        "Department/Loc",
        "Location",
        "Total",
        "Date Parsed",
        "Month",
        "Worksheet",
    ]
    normalized_sheet_names = [str(name).strip() for name in target_sheet_names if str(name).strip()]
    if not normalized_sheet_names:
        return pd.DataFrame(columns=report_columns)

    report_frames: list[pd.DataFrame] = []
    for source_sheet_name in normalized_sheet_names:
        source_headers, source_rows, _ = load_sheet_data(path_str, source_sheet_name)
        if not source_rows:
            continue
        source_frame = pd.DataFrame(source_rows, columns=source_headers)
        sheet_report_frame = build_po_reporting_frame(
            source_frame,
            source_headers,
            location_options=location_options,
        )
        if sheet_report_frame.empty:
            continue
        scoped_report_frame = sheet_report_frame.copy()
        scoped_report_frame["Worksheet"] = source_sheet_name
        report_frames.append(scoped_report_frame)

    if not report_frames:
        return pd.DataFrame(columns=report_columns)

    combined_report_frame = pd.concat(report_frames, ignore_index=True)
    if "Worksheet" not in combined_report_frame.columns:
        combined_report_frame["Worksheet"] = ""
    return combined_report_frame.sort_values(by=["Date Parsed", "PO Number"], ascending=[False, False]).reset_index(
        drop=True
    )


def create_backup(path: Path, backup_dir: Path, keep_latest: int) -> Path | None:
    if not path.exists():
        return None

    keep_latest = max(MIN_BACKUP_KEEP_LATEST, int(keep_latest))
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
    backup_path = backup_dir / f"{path.stem}-{stamp}{path.suffix}"
    try:
        copy_file_with_retry(path, backup_path, retries=3, delay_seconds=0.15)
    except Exception:
        return None

    backups = sorted(
        backup_dir.glob(f"{path.stem}-*{path.suffix}"),
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    for old_backup in backups[keep_latest:]:
        try:
            old_backup.unlink()
        except Exception:
            continue

    return backup_path


def list_backups(path: Path, backup_dir: Path) -> list[Path]:
    if not backup_dir.exists():
        return []
    return sorted(
        backup_dir.glob(f"{path.stem}-*{path.suffix}"),
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )


def get_latest_backup(path: Path, backup_dir: Path) -> Path | None:
    backups = list_backups(path, backup_dir)
    return backups[0] if backups else None


def restore_backup(path: Path, backup_dir: Path, backup_file: Path) -> Path | None:
    if not backup_file.exists():
        return None

    try:
        resolved_backup = backup_file.resolve()
        resolved_dir = backup_dir.resolve()
        if resolved_backup.parent != resolved_dir:
            return None
    except Exception:
        return None

    backup_dir.mkdir(parents=True, exist_ok=True)
    if path.exists():
        safety_stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        safety_copy = backup_dir / f"{path.stem}-restore-safety-{safety_stamp}{path.suffix}"
        try:
            copy_file_with_retry(path, safety_copy, retries=2, delay_seconds=0.1)
        except Exception:
            pass

    temp_fd, temp_name = tempfile.mkstemp(
        prefix=f"{path.name}.restore.",
        suffix=".tmp",
        dir=str(path.parent),
    )
    os.close(temp_fd)
    temp_path = Path(temp_name)
    try:
        copy_file_with_retry(backup_file, temp_path, retries=3, delay_seconds=0.2)
        os.replace(temp_path, path)
    finally:
        try:
            temp_path.unlink(missing_ok=True)
        except Exception:
            pass
    return backup_file


def restore_latest_backup(path: Path, backup_dir: Path) -> Path | None:
    latest_backup = get_latest_backup(path, backup_dir)
    if latest_backup is None:
        return None
    return restore_backup(path, backup_dir, latest_backup)


def copy_previous_row_style(
    worksheet: Any,
    target_row: int,
    start_col: int,
    end_col: int,
) -> bool:
    source_row = target_row - 1
    if source_row < 2:
        return False

    copied_any = False
    for column in range(start_col, end_col + 1):
        source_cell = worksheet.cell(row=source_row, column=column)
        target_cell = worksheet.cell(row=target_row, column=column)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
            copied_any = True
    return copied_any


def apply_default_box_border(worksheet: Any, row_index: int, start_col: int, end_col: int) -> None:
    side = Side(border_style="thin", color="000000")
    boxed_border = Border(left=side, right=side, top=side, bottom=side)

    for column in range(start_col, end_col + 1):
        worksheet.cell(row=row_index, column=column).border = boxed_border


def apply_group_outline_border(
    worksheet: Any,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
) -> None:
    thin = Side(border_style="thin", color="000000")
    none = Side(border_style=None, color=None)

    for row_index in range(start_row, end_row + 1):
        for column_index in range(start_col, end_col + 1):
            left = thin if column_index == start_col else none
            right = thin if column_index == end_col else none
            top = thin if row_index == start_row else none
            bottom = thin if row_index == end_row else none
            worksheet.cell(row=row_index, column=column_index).border = Border(
                left=left,
                right=right,
                top=top,
                bottom=bottom,
            )


def create_line_item_row(
    item: str = "",
    unit_price: float = 0.0,
    quantity: int = 1,
) -> dict[str, Any]:
    return {
        "Row ID": uuid4().hex,
        "Item": item,
        "Price Per Item": round(parse_float(unit_price, 0.0), 2),
        "Quantity": parse_int(quantity, 1),
    }


def default_line_items() -> list[dict[str, Any]]:
    return [create_line_item_row()]


def ensure_line_item_rows(raw_items: Any) -> list[dict[str, Any]]:
    if not isinstance(raw_items, list):
        return default_line_items()

    normalized_rows: list[dict[str, Any]] = []
    for raw_item in raw_items:
        if not isinstance(raw_item, dict):
            continue
        row_id = str(raw_item.get("Row ID", "")).strip() or uuid4().hex
        normalized_rows.append(
            {
                "Row ID": row_id,
                "Item": str(raw_item.get("Item", "")),
                "Price Per Item": round(parse_float(raw_item.get("Price Per Item", 0.0), 0.0), 2),
                "Quantity": parse_int(raw_item.get("Quantity", 1), 1),
            }
        )

    return normalized_rows or default_line_items()


def parse_float(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default

    if isinstance(value, str):
        cleaned = value.strip().replace("$", "").replace(",", "")
        if not cleaned:
            return default
        try:
            return float(cleaned)
        except ValueError:
            return default

    try:
        if pd.isna(value):
            return default
    except Exception:
        pass

    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def parse_int(value: Any, default: int = 1) -> int:
    try:
        parsed = int(round(parse_float(value, default=float(default))))
    except Exception:
        parsed = default
    return parsed if parsed > 0 else default


def normalize_line_items(
    raw_items: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[str]]:
    normalized: list[dict[str, Any]] = []
    errors: list[str] = []

    for index, raw_item in enumerate(raw_items, start=1):
        item_name = str(raw_item.get("Item", "")).strip()
        unit_price = round(parse_float(raw_item.get("Price Per Item", 0.0), 0.0), 2)
        quantity = parse_int(raw_item.get("Quantity", 1), 1)

        if not item_name and unit_price == 0.0:
            continue

        if not item_name:
            errors.append(f"Line {index} is missing an item name.")
            continue

        line_total = round(unit_price * quantity, 2)
        normalized.append(
            {
                "Items Being Purchased": item_name,
                "Price Per Item": unit_price,
                "Quantity": quantity,
                "Sub Total": line_total,
            }
        )

    return normalized, errors


def get_next_id_value(worksheet: Any, column_index: int) -> str:
    numeric_values: list[int] = []
    for row_index in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row=row_index, column=column_index).value
        if isinstance(value, bool):
            continue
        if isinstance(value, (int, float)):
            numeric_values.append(int(value))
            continue
        if isinstance(value, str) and value.strip().isdigit():
            numeric_values.append(int(value.strip()))

    if not numeric_values:
        return "1"
    return str(max(numeric_values) + 1)


def row_has_values(worksheet: Any, row_index: int, column_indexes: list[int]) -> bool:
    for column_index in column_indexes:
        value = worksheet.cell(row=row_index, column=column_index).value
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return True
    return False


def find_next_write_row(
    worksheet: Any,
    column_indexes: list[int],
    header_row: int = 1,
) -> int:
    last_candidate_row = max(int(worksheet.max_row or header_row), header_row)
    while last_candidate_row > header_row and not row_has_values(
        worksheet,
        row_index=last_candidate_row,
        column_indexes=column_indexes,
    ):
        last_candidate_row -= 1
    return last_candidate_row + 1


def get_workbook_lock_path(path: Path) -> Path:
    return path.with_name(f"{path.name}.lock")


@contextmanager
def workbook_write_lock(
    path: Path,
    timeout_seconds: float = WORKBOOK_LOCK_TIMEOUT_SECONDS,
    stale_seconds: float = WORKBOOK_LOCK_STALE_SECONDS,
):
    lock_path = get_workbook_lock_path(path)
    timeout = max(float(timeout_seconds), 0.1)
    stale_after = max(float(stale_seconds), 1.0)
    deadline = time.monotonic() + timeout
    lock_fd: int | None = None

    while lock_fd is None:
        try:
            lock_fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            lock_note = f"pid={os.getpid()} at {datetime.now().isoformat(timespec='seconds')}\n"
            os.write(lock_fd, lock_note.encode("utf-8"))
        except FileExistsError:
            stale_lock = False
            try:
                lock_age = time.time() - lock_path.stat().st_mtime
                stale_lock = lock_age >= stale_after
            except Exception:
                stale_lock = False

            if stale_lock:
                try:
                    lock_path.unlink(missing_ok=True)
                    continue
                except Exception:
                    pass

            if time.monotonic() >= deadline:
                raise TimeoutError(
                    "Workbook is busy because another save is in progress. "
                    "Please wait a moment and try Save PO again."
                )
            time.sleep(0.2)

    try:
        yield
    finally:
        if lock_fd is not None:
            try:
                os.close(lock_fd)
            except Exception:
                pass
        try:
            lock_path.unlink(missing_ok=True)
        except Exception:
            pass


def append_record(
    path: Path,
    sheet_name: str,
    headers: list[str],
    values: dict[str, Any] | list[dict[str, Any]],
    backup_dir: Path,
    keep_backups: int,
    purchase_reason: str = "",
    purchase_reason_column_index: int | None = None,
) -> Path | None:
    backup_path = create_backup(path, backup_dir, keep_backups)

    workbook = open_workbook_with_retry(path, read_only=False, data_only=False)
    if sheet_name not in workbook.sheetnames:
        worksheet = workbook.create_sheet(title=sheet_name)
        worksheet.append(headers)
    else:
        worksheet = workbook[sheet_name]

    existing_raw_headers = [
        worksheet.cell(row=1, column=column).value for column in range(1, worksheet.max_column + 1)
    ]
    if not existing_raw_headers or all(value is None or str(value).strip() == "" for value in existing_raw_headers):
        for column, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=column, value=header)
        existing_headers = headers.copy()
    else:
        existing_headers = sanitize_headers(existing_raw_headers)

    header_to_column: dict[str, int] = {
        header: column for column, header in enumerate(existing_headers, start=1)
    }

    for header in headers:
        if header not in header_to_column:
            next_column = worksheet.max_column + 1
            worksheet.cell(row=1, column=next_column, value=header)
            header_to_column[header] = next_column

    values_to_write = values if isinstance(values, list) else [values]
    row_columns = [header_to_column[header] for header in headers if header in header_to_column]
    start_col = min(row_columns) if row_columns else 1
    end_col = max(row_columns) if row_columns else worksheet.max_column
    first_written_row = find_next_write_row(
        worksheet,
        column_indexes=row_columns if row_columns else [1],
        header_row=1,
    )
    last_written_row = first_written_row - 1

    for offset, row_values in enumerate(values_to_write):
        row_index = first_written_row + offset
        last_written_row = row_index
        style_copied = copy_previous_row_style(worksheet, row_index, start_col, end_col)
        for header in headers:
            raw_value = row_values.get(header, "")
            user_value = raw_value.strip() if isinstance(raw_value, str) else raw_value

            if (user_value is None or user_value == "") and header_is_timestamp(header):
                user_value = datetime.now().strftime("%Y-%m-%d %H:%M")

            column_index = header_to_column[header]
            if (
                len(values_to_write) == 1
                and (user_value is None or user_value == "")
                and header_is_id(header)
            ):
                user_value = get_next_id_value(worksheet, column_index)

            if isinstance(user_value, str):
                worksheet.cell(row=row_index, column=column_index, value=user_value or None)
            else:
                worksheet.cell(row=row_index, column=column_index, value=user_value)

        if len(values_to_write) == 1 and not style_copied:
            apply_default_box_border(worksheet, row_index, start_col, end_col)

    if len(values_to_write) > 1 and first_written_row <= last_written_row:
        apply_group_outline_border(
            worksheet=worksheet,
            start_row=first_written_row,
            end_row=last_written_row,
            start_col=start_col,
            end_col=end_col,
        )

    reason_text = str(purchase_reason).strip()
    if (
        reason_text
        and purchase_reason_column_index is not None
        and first_written_row <= last_written_row
        and purchase_reason_column_index > 0
    ):
        reason_cell = worksheet.cell(row=first_written_row, column=purchase_reason_column_index)
        if first_written_row > 1:
            source_cell = worksheet.cell(row=first_written_row - 1, column=purchase_reason_column_index)
            if source_cell.has_style:
                reason_cell._style = copy(source_cell._style)
        reason_cell.value = reason_text

    workbook.save(path)
    workbook.close()
    return backup_path


def update_sheet_rows(
    path: Path,
    sheet_name: str,
    headers: list[str],
    row_updates: list[tuple[int, dict[str, Any]]],
    backup_dir: Path,
    keep_backups: int,
    row_deletes: list[int] | None = None,
    new_rows: list[dict[str, Any]] | None = None,
) -> Path | None:
    normalized_updates: list[tuple[int, dict[str, Any]]] = []
    for row_number, row_values in row_updates:
        try:
            parsed_row_number = int(row_number)
        except Exception:
            continue
        if parsed_row_number <= 1:
            continue
        normalized_updates.append((parsed_row_number, row_values))

    normalized_deletes: list[int] = []
    for row_number in row_deletes or []:
        try:
            parsed_row_number = int(row_number)
        except Exception:
            continue
        if parsed_row_number <= 1:
            continue
        normalized_deletes.append(parsed_row_number)
    normalized_deletes = sorted(set(normalized_deletes), reverse=True)

    normalized_new_rows: list[dict[str, Any]] = []
    for raw_row in new_rows or []:
        normalized_row: dict[str, Any] = {
            header: normalize_editor_cell_value(raw_row.get(header, ""))
            for header in headers
        }
        if any(has_non_empty_editor_value(normalized_row.get(header, "")) for header in headers):
            normalized_new_rows.append(normalized_row)

    if not normalized_updates and not normalized_deletes and not normalized_new_rows:
        return None

    backup_path = create_backup(path, backup_dir, keep_backups)
    workbook = open_workbook_with_retry(path, read_only=False, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Worksheet `{sheet_name}` was not found in `{path}`.")

        worksheet = workbook[sheet_name]
        existing_raw_headers = [
            worksheet.cell(row=1, column=column).value for column in range(1, worksheet.max_column + 1)
        ]
        if not existing_raw_headers or all(value is None or str(value).strip() == "" for value in existing_raw_headers):
            for column, header in enumerate(headers, start=1):
                worksheet.cell(row=1, column=column, value=header)
            existing_headers = headers.copy()
        else:
            existing_headers = sanitize_headers(existing_raw_headers)

        header_to_column: dict[str, int] = {
            header: column for column, header in enumerate(existing_headers, start=1)
        }
        for header in headers:
            if header not in header_to_column:
                next_column = worksheet.max_column + 1
                worksheet.cell(row=1, column=next_column, value=header)
                header_to_column[header] = next_column

        rows_deleted = set(normalized_deletes)
        for row_number, row_values in normalized_updates:
            if row_number in rows_deleted:
                continue
            for header in headers:
                if header not in header_to_column:
                    continue
                column_index = header_to_column[header]
                normalized_value = normalize_editor_cell_value(row_values.get(header, ""))
                if isinstance(normalized_value, str):
                    worksheet.cell(row=row_number, column=column_index, value=normalized_value or None)
                else:
                    worksheet.cell(row=row_number, column=column_index, value=normalized_value)

        for row_number in normalized_deletes:
            if row_number <= worksheet.max_row:
                worksheet.delete_rows(row_number, 1)

        if normalized_new_rows:
            row_columns = [header_to_column[header] for header in headers if header in header_to_column]
            start_col = min(row_columns) if row_columns else 1
            end_col = max(row_columns) if row_columns else worksheet.max_column
            first_insert_row = find_next_write_row(
                worksheet,
                column_indexes=row_columns if row_columns else [1],
                header_row=1,
            )
            for offset, row_values in enumerate(normalized_new_rows):
                row_index = first_insert_row + offset
                style_copied = copy_previous_row_style(worksheet, row_index, start_col, end_col)
                for header in headers:
                    if header not in header_to_column:
                        continue
                    column_index = header_to_column[header]
                    cell_value = row_values.get(header, "")
                    if (cell_value is None or cell_value == "") and header_is_timestamp(header):
                        cell_value = datetime.now().strftime("%Y-%m-%d %H:%M")
                    if (cell_value is None or cell_value == "") and header_is_id(header):
                        cell_value = get_next_id_value(worksheet, column_index)
                    if isinstance(cell_value, str):
                        worksheet.cell(row=row_index, column=column_index, value=cell_value or None)
                    else:
                        worksheet.cell(row=row_index, column=column_index, value=cell_value)
                if not style_copied:
                    apply_default_box_border(worksheet, row_index, start_col, end_col)

        workbook.save(path)
    finally:
        workbook.close()

    return backup_path


def filter_records(frame: pd.DataFrame, query: str) -> pd.DataFrame:
    filtered, _, _ = filter_records_lazy(frame, query, max_scan_rows=len(frame))
    return filtered


def filter_records_lazy(
    frame: pd.DataFrame,
    query: str,
    max_scan_rows: int = DEFAULT_EDITOR_SEARCH_SCAN_LIMIT,
) -> tuple[pd.DataFrame, bool, int]:
    if not query.strip():
        return frame, False, len(frame)

    bounded_scan_rows = max(1, int(max_scan_rows))
    lowered = query.strip().lower()
    truncated = len(frame) > bounded_scan_rows
    candidate_frame = frame.tail(bounded_scan_rows) if truncated else frame

    mask = pd.Series(False, index=candidate_frame.index)
    for column_name in candidate_frame.columns:
        column_values = candidate_frame[column_name].astype(str)
        mask = mask | column_values.str.lower().str.contains(lowered, na=False)

    return candidate_frame[mask], truncated, len(candidate_frame)


def first_non_empty(series: pd.Series) -> str:
    for value in series:
        text = str(value).strip()
        if text:
            return text
    return ""


def build_po_reporting_frame(
    frame: pd.DataFrame,
    headers: list[str],
    location_options: list[str] | None = None,
) -> pd.DataFrame:
    if frame.empty:
        return pd.DataFrame(
            columns=["PO Number", "Date", "Vendor/Store", "Department/Loc", "Location", "Total"]
        )

    po_header = find_first_header(headers, ["PO Number", "PO#", "PO #", "PO"])
    date_header = find_first_header(headers, ["Date"])
    vendor_header = find_first_header(headers, ["Vendor/Store", "Vendor"])
    dept_loc_header = find_first_header(headers, ["Department/Loc", "Deparment/Loc"])
    department_header = find_first_header(headers, ["Department", "Deparment"])
    location_header = find_first_header(headers, ["Location", "Loc"])
    grand_total_header = find_first_header(headers, ["Grand Total", "GrandTotal"])
    sub_total_header = find_first_header(headers, ["Sub Total", "Subtotal"])
    price_header = find_first_header(headers, ["Price Per Item", "Unit Price", "Price"])
    quantity_header = find_first_header(headers, ["Quantity", "Qty", "QTY"])

    working = frame.copy()
    if po_header and po_header in working.columns:
        working["__po"] = working[po_header].astype(str).str.strip().replace("", pd.NA).ffill().fillna("")
    else:
        return pd.DataFrame(
            columns=["PO Number", "Date", "Vendor/Store", "Department/Loc", "Location", "Total"]
        )

    working = working[working["__po"] != ""].copy()
    if working.empty:
        return pd.DataFrame(
            columns=["PO Number", "Date", "Vendor/Store", "Department/Loc", "Location", "Total"]
        )

    if grand_total_header and grand_total_header in working.columns:
        working["__grand_total"] = working[grand_total_header].apply(lambda value: parse_float(value, 0.0))
    else:
        working["__grand_total"] = 0.0
    if sub_total_header and sub_total_header in working.columns:
        working["__sub_total"] = working[sub_total_header].apply(lambda value: parse_float(value, 0.0))
    else:
        working["__sub_total"] = 0.0
    if price_header and price_header in working.columns:
        working["__fallback_price"] = working[price_header].apply(lambda value: parse_float(value, 0.0))
    else:
        working["__fallback_price"] = 0.0
    if quantity_header and quantity_header in working.columns:
        working["__fallback_qty"] = working[quantity_header].apply(lambda value: parse_float(value, 1.0))
    else:
        working["__fallback_qty"] = 1.0
    working["__fallback_qty"] = working["__fallback_qty"].apply(lambda value: value if value > 0 else 1.0)
    working["__fallback_line_total"] = (working["__fallback_price"] * working["__fallback_qty"]).round(2)

    grouped_rows: list[dict[str, Any]] = []
    for po_number, group in working.groupby("__po", dropna=True):
        max_grand = float(group["__grand_total"].max())
        sum_sub_total = float(group["__sub_total"].sum())
        fallback_group_total = float(group["__fallback_line_total"].sum())
        if max_grand > 0:
            total_value = round(max_grand, 2)
        elif sum_sub_total > 0:
            total_value = round(sum_sub_total, 2)
        elif fallback_group_total > 0:
            total_value = round(fallback_group_total, 2)
        else:
            total_value = 0.0

        department_loc_text = ""
        if dept_loc_header and dept_loc_header in group.columns:
            department_loc_text = first_non_empty(group[dept_loc_header])
        if not department_loc_text:
            department_value = first_non_empty(group[department_header]) if department_header in group.columns else ""
            location_value = first_non_empty(group[location_header]) if location_header in group.columns else ""
            if location_value and department_value:
                department_loc_text = f"{location_value}/{department_value}"
            else:
                department_loc_text = location_value or department_value

        location_value = extract_location_code(
            raw_location_value=(first_non_empty(group[location_header]) if location_header in group.columns else ""),
            raw_department_loc_value=department_loc_text,
            location_options=location_options,
        )

        grouped_rows.append(
            {
                "PO Number": str(po_number).strip(),
                "Date": first_non_empty(group[date_header]) if date_header in group.columns else "",
                "Vendor/Store": first_non_empty(group[vendor_header]) if vendor_header in group.columns else "",
                "Department/Loc": department_loc_text,
                "Location": location_value,
                "Total": total_value,
            }
        )

    reporting_frame = pd.DataFrame(grouped_rows)
    if reporting_frame.empty:
        return reporting_frame

    reporting_frame["Date Parsed"] = pd.to_datetime(reporting_frame["Date"], errors="coerce")
    reporting_frame["Month"] = reporting_frame["Date Parsed"].dt.to_period("M").astype(str)
    reporting_frame.loc[reporting_frame["Month"] == "NaT", "Month"] = "Unknown"
    return reporting_frame.sort_values(by=["Date Parsed", "PO Number"], ascending=[False, False]).reset_index(
        drop=True
    )


def build_diagnostics_payload(
    workbook_path: Path,
    sheet_name: str,
    theme_name: str,
    update_manifest_url: str,
    backup_keep_latest: int,
) -> dict[str, Any]:
    workbook_exists = workbook_path.exists()
    workbook_writable = False
    if workbook_exists:
        try:
            with open(workbook_path, "ab"):
                workbook_writable = True
        except Exception:
            workbook_writable = False

    payload = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "app_version": APP_VERSION,
        "python": sys.version.split(" ")[0],
        "executable": sys.executable,
        "session_owner": SESSION_OWNER,
        "session_id": SESSION_ID,
        "workbook_path": str(workbook_path),
        "workbook_exists": workbook_exists,
        "workbook_writable": workbook_writable,
        "active_sheet": sheet_name,
        "theme": theme_name,
        "update_manifest_url": str(update_manifest_url).strip(),
        "backup_keep_latest": int(backup_keep_latest),
        "lock_file": str(get_workbook_lock_path(workbook_path)),
        "lock_file_exists": get_workbook_lock_path(workbook_path).exists(),
        "reservation_file": str(get_po_reservation_path(workbook_path)),
        "active_reservations": get_active_po_reservation_count(workbook_path, sheet_name=sheet_name),
        "runtime_log_path": str(APP_RUNTIME_LOG_PATH),
        "runtime_log_exists": APP_RUNTIME_LOG_PATH.exists(),
        "runtime_log_lines": get_runtime_log_line_count(),
    }
    return payload


def main() -> None:
    desktop_mode_enabled = os.environ.get(DESKTOP_MODE_ENV_VAR, "").strip() == "1"
    browser_mode_override = os.environ.get(BROWSER_MODE_OVERRIDE_ENV_VAR, "").strip() == "1"
    if not desktop_mode_enabled and not browser_mode_override:
        st.set_page_config(
            page_title="POtrol",
            page_icon=":ledger:",
            layout="wide",
            initial_sidebar_state="collapsed",
        )
        st.error("POtrol browser mode is disabled. Launch `POtrol.exe` or `potrol_launcher.py`.")
        st.info("For local development only, set `POTROL_ALLOW_BROWSER_MODE=1` before running Streamlit.")
        st.stop()

    page_icon: str = ":ledger:"
    if APP_ICON_PATH.exists():
        page_icon = str(APP_ICON_PATH)

    st.set_page_config(
        page_title="POtrol",
        page_icon=page_icon,
        layout="wide",
        initial_sidebar_state="collapsed",
    )

    app_settings = load_app_settings()
    if THEME_STATE_KEY not in st.session_state:
        saved_theme_name = canonical_theme_name(str(app_settings.get("theme", "")).strip())
        st.session_state[THEME_STATE_KEY] = (
            saved_theme_name if saved_theme_name in THEME_PRESETS else DEFAULT_THEME_NAME
        )
    else:
        st.session_state[THEME_STATE_KEY] = canonical_theme_name(
            str(st.session_state.get(THEME_STATE_KEY, "")).strip()
        )
    if st.session_state[THEME_STATE_KEY] not in THEME_PRESETS:
        st.session_state[THEME_STATE_KEY] = DEFAULT_THEME_NAME
    theme_palette = resolve_theme_palette(st.session_state[THEME_STATE_KEY])
    theme_outline = theme_palette.get("outline", theme_palette["border"])
    accent_rgb = hex_to_rgb_triplet(theme_palette["accent"], "11, 87, 208")
    accent_strong_rgb = hex_to_rgb_triplet(theme_palette["accent_strong"], "8, 66, 160")
    theme_color_scheme = str(theme_palette.get("color_scheme", "light")).strip().lower()
    if theme_color_scheme not in {"light", "dark"}:
        theme_color_scheme = "light"
    theme_placeholder = str(theme_palette.get("placeholder", theme_palette["muted"])).strip()
    theme_disabled_text = str(theme_palette.get("disabled_text", theme_palette["muted"])).strip()

    st.markdown(
        """
        <style>
            :root {
                color-scheme: light;
                --potrol-bg-start: #f3f6ff;
                --potrol-bg-end: #f7f9ff;
                --potrol-surface: #ffffff;
                --potrol-surface-soft: #eef2ff;
                --potrol-border: #d2d9e8;
                --potrol-outline: #c5ccda;
                --potrol-text: #1b1b1f;
                --potrol-muted: #47464f;
                --potrol-accent: #0b57d0;
                --potrol-accent-strong: #0842a0;
                --potrol-accent-rgb: 11, 87, 208;
                --potrol-accent-strong-rgb: 8, 66, 160;
                --potrol-disabled-bg: #f3f5f8;
                --potrol-placeholder: #6e7280;
                --potrol-disabled-text: #394150;
                --potrol-radius: 16px;
                --potrol-radius-lg: 24px;
                --potrol-field-height: 2.9rem;
                --potrol-field-border-width: 1px;
                --potrol-control-radius: var(--potrol-radius);
                --potrol-control-border-width: var(--potrol-field-border-width);
                --potrol-control-border-color: var(--potrol-outline);
                --primary-color: var(--potrol-accent);
                --secondary-background-color: var(--potrol-surface-soft);
                --background-color: var(--potrol-bg-end);
                --text-color: var(--potrol-text);
            }

            html, body, [class*="st-"], [data-testid="stAppViewContainer"] {
                color: var(--potrol-text);
                font-family: "Google Sans Text", "Google Sans", "Roboto Flex", "Noto Sans", "Segoe UI", sans-serif;
                letter-spacing: 0.005em;
            }

            html, body, [data-testid="stAppViewContainer"] {
                background-color: var(--potrol-bg-end) !important;
            }
            .stApp,
            [data-testid="stAppViewContainer"] {
                background: linear-gradient(180deg, var(--potrol-bg-start) 0%, var(--potrol-bg-end) 100%);
            }
            [data-testid="stDecoration"],
            [data-testid="stToolbar"],
            [data-testid="stHeader"],
            [data-testid="stStatusWidget"],
            [data-testid="stSidebar"],
            [data-testid="stSidebarCollapsedControl"],
            [data-testid="collapsedControl"] {
                display: none !important;
                visibility: hidden !important;
            }
            [data-testid="stAppViewContainer"] > .main {
                padding-top: 0 !important;
            }
            [data-testid="stAppViewContainer"] > .main .block-container,
            [data-testid="stMainBlockContainer"] {
                padding-top: 0.6rem !important;
                padding-bottom: 1.1rem !important;
                padding-left: 0.9rem !important;
                padding-right: 0.9rem !important;
                max-width: 1240px;
                box-sizing: border-box;
            }
            [data-testid="column"] {
                min-width: 0 !important;
            }
            [data-testid="column"] > div {
                min-width: 0 !important;
            }
            .stApp h1,
            .stApp h2,
            .stApp h3 {
                color: var(--potrol-text);
                font-weight: 600;
                letter-spacing: -0.01em;
            }
            [data-testid="stHeaderActionElements"],
            h1 a,
            h2 a,
            h3 a,
            h4 a,
            h5 a,
            h6 a {
                display: none !important;
                visibility: hidden !important;
            }
            .stApp p, .stApp label, .stApp [data-testid="stCaptionContainer"] {
                color: var(--potrol-muted);
            }
            [data-testid="stTabs"] {
                margin-top: 0.15rem;
            }
            [data-testid="stTabs"] [data-baseweb="tab-list"] {
                background: var(--potrol-surface-soft);
                border: 1px solid var(--potrol-border);
                border-radius: 999px;
                padding: 4px;
                gap: 0.3rem;
            }
            [data-testid="stTabs"] [data-baseweb="tab-border"] {
                display: none !important;
            }
            [data-testid="stTabs"] [data-baseweb="tab-highlight"] {
                display: none !important;
            }
            [data-testid="stTabs"] [data-baseweb="tab"] {
                border-radius: 999px;
                height: 2.35rem;
                padding: 0 0.95rem;
                color: var(--potrol-muted);
                font-weight: 600;
                transition: background-color 140ms ease, color 140ms ease;
            }
            [data-testid="stTabs"] [data-baseweb="tab"][aria-selected="true"] {
                background: var(--potrol-surface);
                color: var(--potrol-text) !important;
                box-shadow: 0 1px 2px rgba(17, 24, 39, 0.16);
            }
            [data-testid="stForm"],
            [data-testid="stDataFrame"],
            [data-testid="stAlert"],
            [data-testid="stExpander"] > details {
                background: var(--potrol-surface);
                border: 1px solid var(--potrol-border);
                border-radius: var(--potrol-radius-lg);
                box-shadow: 0 1px 2px rgba(17, 24, 39, 0.09), 0 8px 24px rgba(17, 24, 39, 0.06);
            }
            [data-testid="stForm"] {
                padding: 1rem 1rem 0.45rem 1rem;
            }
            [data-testid="stDataFrame"] {
                overflow: hidden;
            }
            [data-testid="stMetric"] {
                border: 1px solid var(--potrol-outline);
                border-radius: var(--potrol-radius);
                padding: 0.55rem;
                background: var(--potrol-surface-soft);
                box-shadow: 0 1px 2px rgba(17, 24, 39, 0.08);
            }
            [data-testid="stMetricLabel"],
            [data-testid="stMetricValue"],
            [data-testid="stMetricDelta"] {
                color: var(--potrol-text) !important;
            }
            [data-baseweb="input"],
            [data-baseweb="base-input"],
            [data-baseweb="select"],
            [data-baseweb="textarea"] {
                border-radius: var(--potrol-control-radius) !important;
            }
            [data-testid="stTextInput"] [data-baseweb="input"],
            [data-testid="stNumberInput"] [data-baseweb="input"],
            [data-testid="stDateInput"] [data-baseweb="input"],
            [data-testid="stSelectbox"] [data-baseweb="select"],
            [data-testid="stTextArea"] [data-baseweb="textarea"],
            [data-testid="stDialog"] [data-baseweb="input"],
            [data-testid="stDialog"] [data-baseweb="base-input"],
            [data-testid="stDialog"] [data-baseweb="select"],
            [data-testid="stDialog"] [data-baseweb="textarea"] {
                border: 0 !important;
                box-shadow: none !important;
                background: transparent !important;
                border-radius: var(--potrol-control-radius) !important;
            }
            [data-baseweb="input"] > div,
            [data-baseweb="base-input"] > div,
            [data-baseweb="select"] > div,
            [data-baseweb="textarea"] > div {
                background: var(--potrol-surface) !important;
                border: var(--potrol-control-border-width) solid var(--potrol-control-border-color) !important;
                border-radius: var(--potrol-control-radius) !important;
                min-height: var(--potrol-field-height);
                overflow: hidden !important;
                box-sizing: border-box !important;
                background-clip: padding-box !important;
                box-shadow: none !important;
                display: flex !important;
                align-items: center !important;
                transition: border-color 120ms ease, box-shadow 120ms ease, background-color 120ms ease;
            }
            [data-baseweb="input"] > div > div,
            [data-baseweb="base-input"] > div > div {
                border-radius: inherit !important;
                width: 100% !important;
                min-width: 0 !important;
            }
            [data-baseweb="textarea"] > div > div {
                border-radius: inherit !important;
                width: 100% !important;
                min-width: 0 !important;
            }
            [data-baseweb="select"] > div > div:first-child {
                border-radius: inherit !important;
                min-width: 0 !important;
                flex: 1 1 auto !important;
            }
            [data-baseweb="select"] > div > div:last-child {
                width: auto !important;
                min-width: 0 !important;
                flex: 0 0 auto !important;
                border-radius: inherit !important;
            }
            [data-testid="stNumberInput"] [data-baseweb="input"] {
                width: 100% !important;
                min-width: 0 !important;
            }
            [data-testid="stNumberInput"] [data-baseweb="input"] > div {
                width: 100% !important;
                max-width: 100% !important;
                min-width: 0 !important;
                background: var(--potrol-surface) !important;
                border: var(--potrol-control-border-width) solid var(--potrol-control-border-color) !important;
                box-shadow: none !important;
            }
            [data-testid="stNumberInput"] [data-baseweb="input"] > div > div:first-child {
                min-width: 0 !important;
                flex: 1 1 auto !important;
            }
            [data-testid="stNumberInput"] [data-baseweb="input"] > div > div,
            [data-testid="stNumberInput"] [data-baseweb="input"] > div > div > div,
            [data-testid="stNumberInput"] [data-baseweb="input"] > div > div::before,
            [data-testid="stNumberInput"] [data-baseweb="input"] > div > div::after,
            [data-testid="stNumberInput"] [data-baseweb="input"] > div > div > div::before,
            [data-testid="stNumberInput"] [data-baseweb="input"] > div > div > div::after {
                border: 0 !important;
                outline: 0 !important;
                box-shadow: none !important;
                background: transparent !important;
            }
            [data-testid="stNumberInput"] [data-baseweb="input"] > div > div:first-child,
            [data-testid="stNumberInput"] [data-baseweb="input"] > div > div:first-child * {
                border: 0 !important;
                box-shadow: none !important;
                background: transparent !important;
            }
            [data-testid="stNumberInput"] input[type="number"] {
                border: 0 !important;
                outline: none !important;
                box-shadow: none !important;
                background: transparent !important;
                -moz-appearance: textfield !important;
                -webkit-appearance: none !important;
                -webkit-appearance: textfield !important;
                appearance: textfield !important;
            }
            [data-testid="stNumberInput"] input[type="number"]::-webkit-outer-spin-button,
            [data-testid="stNumberInput"] input[type="number"]::-webkit-inner-spin-button {
                -webkit-appearance: none !important;
                margin: 0 !important;
            }
            [data-testid="stNumberInput"] [data-baseweb="input"] input {
                width: 100% !important;
                min-width: 0 !important;
            }
            [data-testid="stNumberInput"] [data-baseweb="input"] + div,
            [data-testid="stNumberInput"] [data-baseweb="input"] > div > div:last-child,
            [data-testid="stNumberInput"] button[data-testid="stNumberInputStepUp"],
            [data-testid="stNumberInput"] button[data-testid="stNumberInputStepDown"],
            [data-testid="stNumberInput"] [data-baseweb="input"] [role="button"],
            [data-testid="stNumberInput"] [data-baseweb="input"] [data-testid*="Step"],
            [data-testid="stNumberInput"] [data-baseweb="input"] [aria-label*="Increase"],
            [data-testid="stNumberInput"] [data-baseweb="input"] [aria-label*="Decrease"] {
                display: none !important;
            }
            [data-testid="stNumberInput"] > div,
            [data-testid="stNumberInput"] [data-baseweb="input"],
            [data-testid="stNumberInput"] [data-baseweb="input"] > div,
            [data-testid="stNumberInput"] [data-baseweb="input"] > div:hover,
            [data-testid="stNumberInput"] [data-baseweb="input"] > div:focus-within {
                border-color: var(--potrol-control-border-color) !important;
                box-shadow: none !important;
                outline: none !important;
                background: var(--potrol-surface) !important;
            }
            [data-testid="stNumberInput"] [data-baseweb="input"]::before,
            [data-testid="stNumberInput"] [data-baseweb="input"]::after,
            [data-testid="stNumberInput"] [data-baseweb="input"] > div::before,
            [data-testid="stNumberInput"] [data-baseweb="input"] > div::after,
            [data-testid="stNumberInput"] fieldset,
            [data-testid="stNumberInput"] legend {
                border: 0 !important;
                outline: 0 !important;
                box-shadow: none !important;
                background: transparent !important;
            }
            [data-testid="stNumberInput"] input,
            [data-testid="stNumberInput"] input:hover,
            [data-testid="stNumberInput"] input:focus,
            [data-testid="stNumberInput"] input:focus-visible {
                border: 0 !important;
                outline: none !important;
                box-shadow: none !important;
                background: transparent !important;
            }
            [data-baseweb="input"] > div:has(input:disabled),
            [data-baseweb="base-input"] > div:has(input:disabled),
            [data-baseweb="textarea"] > div:has(textarea:disabled) {
                background: var(--potrol-disabled-bg) !important;
                border-color: var(--potrol-control-border-color) !important;
                box-shadow: none !important;
            }
            [data-baseweb="input"] > div:focus-within,
            [data-baseweb="base-input"] > div:focus-within,
            [data-baseweb="select"] > div:focus-within,
            [data-baseweb="textarea"] > div:focus-within {
                border-color: var(--potrol-control-border-color) !important;
                box-shadow: none !important;
            }
            [data-testid="stTextInput"],
            [data-testid="stNumberInput"],
            [data-testid="stSelectbox"],
            [data-testid="stDateInput"] {
                overflow: visible !important;
            }
            [data-testid="stTextInput"] > div {
                padding-bottom: 0.12rem !important;
            }
            [data-testid="stTextInput"] [data-baseweb="input"] > div {
                min-height: calc(var(--potrol-field-height) + 0.08rem) !important;
            }
            [data-testid="stNumberInput"] > div,
            [data-testid="stSelectbox"] > div,
            [data-testid="stDateInput"] > div {
                padding-bottom: 0.02rem !important;
            }
            [data-baseweb="input"] input,
            [data-baseweb="base-input"] input,
            [data-baseweb="textarea"] textarea,
            [data-baseweb="select"] * {
                color: var(--potrol-text) !important;
            }
            [data-testid="stSelectbox"] [data-baseweb="select"],
            [data-testid="stSelectbox"] [data-baseweb="select"] *,
            [role="listbox"],
            [role="listbox"] * {
                color: var(--potrol-text) !important;
                -webkit-text-fill-color: var(--potrol-text) !important;
                opacity: 1 !important;
            }
            [data-testid="stSelectbox"] [data-baseweb="select"] [role="combobox"] > div,
            [data-testid="stSelectbox"] [data-baseweb="select"] span,
            [data-testid="stSelectbox"] [data-baseweb="select"] p {
                color: var(--potrol-text) !important;
                -webkit-text-fill-color: var(--potrol-text) !important;
                opacity: 1 !important;
            }
            [data-baseweb="input"] input,
            [data-baseweb="base-input"] input,
            [data-baseweb="select"] input,
            [data-baseweb="select"] [role="combobox"],
            [data-baseweb="textarea"] textarea {
                text-align: left !important;
                line-height: 1.3 !important;
            }
            [data-baseweb="input"] input,
            [data-baseweb="base-input"] input,
            [data-baseweb="textarea"] textarea {
                background: transparent !important;
            }
            [data-baseweb="input"] input:focus,
            [data-baseweb="base-input"] input:focus,
            [data-baseweb="select"] input:focus,
            [data-baseweb="select"] [role="combobox"]:focus,
            [data-baseweb="textarea"] textarea:focus,
            [data-baseweb="input"] button:focus,
            [data-baseweb="select"] button:focus,
            [data-baseweb="base-input"] button:focus {
                outline: none !important;
                box-shadow: none !important;
            }
            [data-baseweb="input"] button,
            [data-baseweb="base-input"] button,
            [data-baseweb="select"] button {
                border: 0 !important;
                border-radius: var(--potrol-control-radius) !important;
                box-shadow: none !important;
                background: transparent !important;
            }
            [data-testid="stDialog"] [data-baseweb="input"] > div,
            [data-testid="stDialog"] [data-baseweb="base-input"] > div,
            [data-testid="stDialog"] [data-baseweb="select"] > div,
            [data-testid="stDialog"] [data-baseweb="textarea"] > div {
                background: var(--potrol-surface) !important;
                border: var(--potrol-control-border-width) solid var(--potrol-control-border-color) !important;
                box-shadow: none !important;
            }
            [data-testid="stDialog"] [data-baseweb="input"] > div:focus-within,
            [data-testid="stDialog"] [data-baseweb="base-input"] > div:focus-within,
            [data-testid="stDialog"] [data-baseweb="select"] > div:focus-within,
            [data-testid="stDialog"] [data-baseweb="textarea"] > div:focus-within {
                border-color: var(--potrol-control-border-color) !important;
                box-shadow: none !important;
            }
            [data-baseweb="input"] input:disabled,
            [data-baseweb="base-input"] input:disabled,
            [data-baseweb="textarea"] textarea:disabled {
                color: var(--potrol-disabled-text) !important;
                -webkit-text-fill-color: var(--potrol-disabled-text) !important;
                background: transparent !important;
                opacity: 1 !important;
            }
            [data-baseweb="input"] input::placeholder,
            [data-baseweb="textarea"] textarea::placeholder {
                color: var(--potrol-placeholder) !important;
            }
            input[type="checkbox"],
            input[type="radio"] {
                accent-color: var(--potrol-accent) !important;
            }
            [data-testid="stCheckbox"] p,
            [data-testid="stSelectbox"] label p,
            [data-testid="stTextInput"] label p,
            [data-testid="stNumberInput"] label p {
                color: var(--potrol-text) !important;
            }
            [role="listbox"] {
                border: var(--potrol-control-border-width) solid var(--potrol-control-border-color) !important;
                border-radius: var(--potrol-control-radius) !important;
                background: var(--potrol-surface) !important;
            }
            [role="option"] {
                color: var(--potrol-text) !important;
            }
            [role="option"][aria-selected="true"],
            [role="option"]:hover {
                background: rgba(var(--potrol-accent-rgb), 0.14) !important;
                color: var(--potrol-text) !important;
            }
            [data-baseweb="popover"],
            [data-baseweb="popover"] > div {
                background: var(--potrol-surface) !important;
                color: var(--potrol-text) !important;
                border-color: var(--potrol-border) !important;
            }
            [data-testid="stDataFrame"] [role="grid"],
            [data-testid="stDataFrame"] [role="row"],
            [data-testid="stDataFrame"] [role="gridcell"] {
                background: var(--potrol-surface) !important;
                color: var(--potrol-text) !important;
                border-color: var(--potrol-border) !important;
            }
            [data-testid="stDataFrame"] [role="columnheader"] {
                background: var(--potrol-surface-soft) !important;
                color: var(--potrol-text) !important;
                border-color: var(--potrol-border) !important;
            }
            [data-testid="stRadio"] [role="radiogroup"] {
                display: flex;
                flex-wrap: wrap;
                gap: 0.45rem;
            }
            [data-testid="stRadio"] [role="radiogroup"] > label {
                border: 1px solid var(--potrol-border);
                border-radius: 999px;
                background: var(--potrol-surface-soft);
                padding: 0.25rem 0.75rem;
                margin: 0 !important;
            }
            [data-testid="stRadio"] [role="radiogroup"] > label:has(input:checked) {
                border-color: var(--potrol-accent);
                background: var(--potrol-surface);
                box-shadow: inset 0 0 0 1px var(--potrol-accent);
            }
            [data-testid="stRadio"][class*="st-key-reports_scope_mode"] [data-baseweb="radio"] input + div,
            [data-testid="stRadio"][class*="st-key-reports-scope-mode"] [data-baseweb="radio"] input + div {
                border-color: var(--potrol-border) !important;
                background: var(--potrol-surface) !important;
            }
            [data-testid="stRadio"][class*="st-key-reports_scope_mode"] [data-baseweb="radio"] input:checked + div,
            [data-testid="stRadio"][class*="st-key-reports-scope-mode"] [data-baseweb="radio"] input:checked + div {
                border-color: var(--potrol-accent) !important;
                background: var(--potrol-surface) !important;
            }
            [data-testid="stRadio"][class*="st-key-reports_scope_mode"] [data-baseweb="radio"] input:checked + div > div,
            [data-testid="stRadio"][class*="st-key-reports-scope-mode"] [data-baseweb="radio"] input:checked + div > div {
                background: var(--potrol-accent) !important;
            }
            [data-testid="stRadio"][class*="st-key-reports_scope_mode"] [data-baseweb="radio"] input:focus + div,
            [data-testid="stRadio"][class*="st-key-reports-scope-mode"] [data-baseweb="radio"] input:focus + div {
                box-shadow: 0 0 0 1px rgba(var(--potrol-accent-rgb), 0.45) !important;
            }
            [data-testid="stRadio"][class*="st-key-reports_scope_mode"] [data-baseweb="radio"] svg,
            [data-testid="stRadio"][class*="st-key-reports_scope_mode"] [data-baseweb="radio"] svg *,
            [data-testid="stRadio"][class*="st-key-reports-scope-mode"] [data-baseweb="radio"] svg,
            [data-testid="stRadio"][class*="st-key-reports-scope-mode"] [data-baseweb="radio"] svg * {
                stroke: var(--potrol-border) !important;
                fill: var(--potrol-border) !important;
            }
            [data-testid="stRadio"][class*="st-key-reports_scope_mode"] [data-baseweb="radio"]:has(input:checked) svg,
            [data-testid="stRadio"][class*="st-key-reports_scope_mode"] [data-baseweb="radio"]:has(input:checked) svg *,
            [data-testid="stRadio"][class*="st-key-reports-scope-mode"] [data-baseweb="radio"]:has(input:checked) svg,
            [data-testid="stRadio"][class*="st-key-reports-scope-mode"] [data-baseweb="radio"]:has(input:checked) svg * {
                stroke: var(--potrol-accent) !important;
                fill: var(--potrol-accent) !important;
            }
            [data-testid="stPopover"] > div > button,
            [data-testid="stPopoverButton"] > button {
                background: var(--potrol-accent) !important;
                color: #ffffff !important;
                border: 1px solid var(--potrol-accent) !important;
                border-radius: 999px !important;
                min-height: 2.6rem !important;
                padding: 0.45rem 1rem !important;
                font-weight: 600 !important;
                line-height: 1.2 !important;
                white-space: nowrap !important;
                box-shadow: 0 1px 2px rgba(0, 0, 0, 0.14);
            }
            [data-testid="stPopover"] > div > button > div,
            [data-testid="stPopoverButton"] > button > div {
                background: transparent !important;
            }
            [data-testid="stPopover"] > div > button > div > span,
            [data-testid="stPopoverButton"] > button > div > span {
                color: #ffffff !important;
            }
            [data-testid="stPopover"] > div > button [data-testid="stButtonIcon"],
            [data-testid="stPopoverButton"] > button [data-testid="stButtonIcon"],
            [data-testid="stPopoverButton"] > button [data-testid="stButtonIcon"] *,
            [data-testid="stPopover"] > div > button svg,
            [data-testid="stPopoverButton"] > button svg,
            [data-testid="stPopover"] > div > button [aria-hidden="true"],
            [data-testid="stPopoverButton"] > button [aria-hidden="true"] {
                display: none !important;
                visibility: hidden !important;
            }
            [data-testid="stPopover"] > div > button:hover,
            [data-testid="stPopoverButton"] > button:hover {
                background: var(--potrol-accent-strong) !important;
                border-color: var(--potrol-accent-strong) !important;
                box-shadow: 0 4px 10px rgba(var(--potrol-accent-strong-rgb), 0.25);
            }
            .stButton > button,
            [data-testid="stDownloadButton"] > button,
            [data-testid="stFormSubmitButton"] > button {
                background: var(--potrol-accent) !important;
                color: #ffffff !important;
                border: 1px solid var(--potrol-accent) !important;
                border-radius: 999px !important;
                min-height: 2.6rem !important;
                padding: 0.45rem 1rem !important;
                font-weight: 600 !important;
                line-height: 1.2 !important;
                white-space: nowrap;
                box-shadow: 0 1px 2px rgba(0, 0, 0, 0.14);
                transition: transform 120ms ease, background-color 120ms ease, box-shadow 120ms ease;
            }
            .stButton > button *,
            [data-testid="stDownloadButton"] > button *,
            [data-testid="stFormSubmitButton"] > button * {
                color: #ffffff !important;
                fill: #ffffff !important;
            }
            .stButton > button:hover,
            [data-testid="stDownloadButton"] > button:hover,
            [data-testid="stFormSubmitButton"] > button:hover {
                background: var(--potrol-accent-strong) !important;
                border-color: var(--potrol-accent-strong) !important;
                color: #ffffff !important;
                box-shadow: 0 4px 10px rgba(var(--potrol-accent-strong-rgb), 0.25);
                transform: translateY(-1px);
            }
            .stButton > button:focus-visible,
            [data-testid="stDownloadButton"] > button:focus-visible,
            [data-testid="stFormSubmitButton"] > button:focus-visible {
                outline: none !important;
                box-shadow: 0 0 0 3px rgba(var(--potrol-accent-rgb), 0.3) !important;
            }
            .stButton > button:disabled,
            [data-testid="stDownloadButton"] > button:disabled,
            [data-testid="stFormSubmitButton"] > button:disabled {
                background: var(--potrol-surface-soft) !important;
                color: var(--potrol-muted) !important;
                border-color: var(--potrol-border) !important;
                box-shadow: none !important;
                transform: none !important;
                opacity: 1 !important;
            }
            .stButton > button:disabled *,
            [data-testid="stDownloadButton"] > button:disabled *,
            [data-testid="stFormSubmitButton"] > button:disabled * {
                color: var(--potrol-muted) !important;
                fill: var(--potrol-muted) !important;
            }
            [data-testid="stPopoverContent"] {
                border: 1px solid var(--potrol-border) !important;
                border-radius: var(--potrol-radius) !important;
                background: var(--potrol-surface) !important;
            }
            [data-testid="stDialog"] [role="dialog"] {
                border-radius: var(--potrol-radius-lg) !important;
                border: 1px solid var(--potrol-border) !important;
                background: linear-gradient(180deg, var(--potrol-surface) 0%, var(--potrol-surface-soft) 100%)
                    !important;
                box-shadow: 0 18px 44px rgba(17, 24, 39, 0.18) !important;
            }
            [data-testid="stDialog"] [data-testid="stVerticalBlock"],
            [data-testid="stDialog"] [data-testid="stHorizontalBlock"],
            [data-testid="stDialog"] [data-testid="stVerticalBlockBorderWrapper"] {
                background: transparent !important;
            }
            [data-testid="stDataFrame"] * {
                color: var(--potrol-text) !important;
            }
            [data-testid="StyledFullScreenButton"],
            button[title="View fullscreen"] {
                display: none !important;
                visibility: hidden !important;
            }
            [data-testid="stAppDeployButton"],
            [data-testid="stDeployButton"],
            button[title="Deploy"],
            a[title="Deploy"] {
                display: none !important;
                visibility: hidden !important;
            }
            .potrol-logo-wrap {
                width: fit-content;
                max-width: 100%;
            }
            .potrol-theme-card {
                margin-top: 0.08rem;
                border: 1px solid var(--potrol-border);
                border-radius: 10px;
                background: var(--potrol-surface);
                padding: 0.3rem;
                min-height: 5.2rem;
                box-sizing: border-box;
                overflow: hidden;
                display: block;
            }
            .potrol-theme-card-link {
                display: block;
                text-decoration: none !important;
                cursor: pointer;
            }
            .potrol-theme-card-link:focus-visible .potrol-theme-card {
                outline: 2px solid rgba(var(--potrol-accent-rgb), 0.48);
                outline-offset: 2px;
            }
            .potrol-theme-card-active {
                border-color: var(--potrol-accent);
                box-shadow: 0 0 0 1px rgba(var(--potrol-accent-rgb), 0.25);
            }
            .potrol-theme-bar {
                height: 0.58rem;
                border-radius: 999px;
                border: 1px solid var(--preview-border, var(--potrol-border));
            }
            .potrol-theme-head {
                margin-top: 0.26rem;
                display: flex;
                align-items: center;
                justify-content: space-between;
                gap: 0.24rem;
            }
            .potrol-theme-name {
                color: var(--preview-text, var(--potrol-text)) !important;
                font-size: 0.71rem;
                font-weight: 600;
                line-height: 1.2;
                white-space: nowrap;
                overflow: hidden;
                text-overflow: ellipsis;
            }
            .potrol-theme-section-title {
                margin: 0.55rem 0 0.14rem 0;
                color: var(--potrol-text);
                font-size: 0.84rem;
                font-weight: 700;
                line-height: 1.2;
            }
            .potrol-theme-meta {
                margin-top: 0.2rem;
                display: flex;
                align-items: center;
                gap: 0.24rem;
            }
            .potrol-theme-pill {
                border: 1px solid var(--preview-border, var(--potrol-border)) !important;
                color: var(--preview-text, var(--potrol-text)) !important;
                background: var(--preview-pill-bg, rgba(17, 24, 39, 0.08)) !important;
                border-radius: 999px;
                padding: 0.06rem 0.45rem;
                font-size: 0.62rem;
                line-height: 1.2;
                font-weight: 700;
                text-transform: uppercase;
                letter-spacing: 0.03em;
            }
            .potrol-theme-apply {
                display: inline-flex;
                align-items: center;
                justify-content: center;
                text-decoration: none !important;
                border-radius: 999px;
                min-height: 1.22rem;
                padding: 0.03rem 0.42rem;
                font-size: 0.63rem;
                font-weight: 700;
                line-height: 1.1;
                border: 1px solid transparent;
                white-space: nowrap;
            }
            .potrol-theme-apply:hover {
                filter: brightness(0.95);
            }
            .potrol-theme-swatches {
                display: flex;
                gap: 0.22rem;
                flex-wrap: wrap;
                margin-top: 0.24rem;
            }
            .potrol-theme-swatch {
                width: 0.68rem;
                height: 0.68rem;
                border-radius: 999px;
                border: 1px solid var(--preview-border, var(--potrol-border));
                box-shadow: 0 1px 2px rgba(var(--potrol-accent-rgb), 0.16);
            }
            .potrol-stat-badge {
                display: inline-block;
                padding: 0.28rem 0.7rem;
                border: 1px solid var(--potrol-border);
                border-radius: 999px;
                background: var(--potrol-surface-soft);
                color: var(--potrol-text);
                font-size: 0.8rem;
                font-weight: 600;
                line-height: 1.3;
            }
            [data-testid="stSlider"] > div {
                padding-top: 0.2rem;
            }
            [data-testid="stSlider"] [data-baseweb="slider"] > div > div {
                border-radius: 999px !important;
            }
            [data-testid="stSlider"] [data-baseweb="slider"] > div > div:first-child {
                background: rgba(var(--potrol-accent-rgb), 0.24) !important;
            }
            [data-testid="stSlider"] [data-baseweb="slider"] > div > div:first-child > div {
                background: var(--potrol-accent) !important;
            }
            [data-testid="stSlider"] [data-baseweb="slider"] div[aria-hidden="true"] {
                background: rgba(var(--potrol-accent-rgb), 0.24) !important;
            }
            [data-testid="stSlider"] [data-baseweb="slider"] div[aria-hidden="true"] > div {
                background: var(--potrol-accent) !important;
            }
            [data-testid="stSlider"] [data-baseweb="slider"] [role="slider"] {
                width: 1.02rem !important;
                height: 1.02rem !important;
                border-radius: 999px !important;
                border: 2px solid var(--potrol-accent-strong) !important;
                background: var(--potrol-surface) !important;
                box-shadow: 0 0 0 3px rgba(var(--potrol-accent-rgb), 0.22) !important;
            }
            [data-testid="stSlider"] [data-baseweb="slider"] p {
                color: var(--potrol-muted) !important;
                font-weight: 600 !important;
            }
            .potrol-report-metric {
                border: 1px solid var(--potrol-border);
                border-radius: 18px;
                background: linear-gradient(180deg, var(--potrol-surface) 0%, var(--potrol-surface-soft) 100%);
                padding: 0.72rem 0.82rem;
                box-shadow: 0 1px 2px rgba(17, 24, 39, 0.08), 0 10px 26px rgba(17, 24, 39, 0.05);
                min-height: 5.25rem;
            }
            .potrol-report-metric-label {
                color: var(--potrol-muted);
                font-size: 0.74rem;
                font-weight: 600;
                text-transform: uppercase;
                letter-spacing: 0.04em;
                line-height: 1.2;
            }
            .potrol-report-metric-value {
                margin-top: 0.32rem;
                color: var(--potrol-text);
                font-size: 1.34rem;
                font-weight: 700;
                line-height: 1.1;
                letter-spacing: -0.01em;
            }
            .potrol-report-metric-note {
                margin-top: 0.28rem;
                color: var(--potrol-muted);
                font-size: 0.74rem;
                line-height: 1.3;
            }
            .potrol-report-card-title {
                color: var(--potrol-text);
                font-size: 0.96rem;
                font-weight: 700;
                line-height: 1.2;
                margin-bottom: 0.12rem;
            }
            .potrol-report-card-sub {
                color: var(--potrol-muted);
                font-size: 0.78rem;
                line-height: 1.3;
                margin-bottom: 0.42rem;
            }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        f"""
        <style>
            :root,
            html,
            body,
            .stApp,
            [data-testid="stApp"],
            [data-testid="stAppViewContainer"],
            [data-testid="stDialog"] [role="dialog"] {{
                color-scheme: {theme_color_scheme};
                --potrol-bg-start: {theme_palette["bg_start"]};
                --potrol-bg-end: {theme_palette["bg_end"]};
                --potrol-surface: {theme_palette["surface"]};
                --potrol-surface-soft: {theme_palette["surface_soft"]};
                --potrol-border: {theme_palette["border"]};
                --potrol-outline: {theme_outline};
                --potrol-text: {theme_palette["text"]};
                --potrol-muted: {theme_palette["muted"]};
                --potrol-accent: {theme_palette["accent"]};
                --potrol-accent-strong: {theme_palette["accent_strong"]};
                --potrol-accent-rgb: {accent_rgb};
                --potrol-accent-strong-rgb: {accent_strong_rgb};
                --potrol-disabled-bg: {theme_palette["surface_soft"]};
                --potrol-placeholder: {theme_placeholder};
                --potrol-disabled-text: {theme_disabled_text};
                --primary-color: {theme_palette["accent"]};
                --secondary-background-color: {theme_palette["surface_soft"]};
                --background-color: {theme_palette["bg_end"]};
                --text-color: {theme_palette["text"]};
            }}
            html,
            body,
            .stApp,
            [data-testid="stApp"],
            [data-testid="stAppViewContainer"] {{
                background-color: var(--potrol-bg-end) !important;
            }}
            .stApp,
            [data-testid="stAppViewContainer"] {{
                background: linear-gradient(180deg, var(--potrol-bg-start) 0%, var(--potrol-bg-end) 100%)
                    !important;
            }}
        </style>
        """,
        unsafe_allow_html=True,
    )

    if "location_options" not in st.session_state:
        st.session_state["location_options"] = load_location_options()
    if WORKBOOK_PATH_STATE_KEY not in st.session_state:
        saved_workbook_path = str(app_settings.get("workbook_path", "")).strip()
        st.session_state[WORKBOOK_PATH_STATE_KEY] = saved_workbook_path or str(DEFAULT_WORKBOOK_PATH)
    if BACKUP_DIR_STATE_KEY not in st.session_state:
        saved_backup_dir = str(app_settings.get("backup_dir", "")).strip()
        if saved_backup_dir:
            st.session_state[BACKUP_DIR_STATE_KEY] = saved_backup_dir
        else:
            initial_workbook_path = Path(st.session_state[WORKBOOK_PATH_STATE_KEY]).expanduser()
            st.session_state[BACKUP_DIR_STATE_KEY] = str(initial_workbook_path.parent / "PO_Backups")
    if BACKUP_KEEP_LATEST_STATE_KEY not in st.session_state:
        st.session_state[BACKUP_KEEP_LATEST_STATE_KEY] = normalize_backup_keep_latest(
            app_settings.get("backup_keep_latest", DEFAULT_BACKUP_KEEP_LATEST)
        )
    else:
        st.session_state[BACKUP_KEEP_LATEST_STATE_KEY] = normalize_backup_keep_latest(
            st.session_state.get(BACKUP_KEEP_LATEST_STATE_KEY, DEFAULT_BACKUP_KEEP_LATEST)
        )
    if SETTINGS_TAB_STATE_KEY not in st.session_state:
        st.session_state[SETTINGS_TAB_STATE_KEY] = "workbook"
    if UPDATE_MANIFEST_URL_STATE_KEY not in st.session_state:
        st.session_state[UPDATE_MANIFEST_URL_STATE_KEY] = str(
            app_settings.get("update_manifest_url", DEFAULT_UPDATE_MANIFEST_URL)
        ).strip()
    if OPEN_SETTINGS_ONCE_STATE_KEY not in st.session_state:
        st.session_state[OPEN_SETTINGS_ONCE_STATE_KEY] = False

    def apply_theme_selection(theme_name: str, reopen_settings: bool = False) -> None:
        requested_theme_name = canonical_theme_name(str(theme_name).strip())
        if requested_theme_name not in THEME_PRESETS:
            return
        st.session_state[THEME_STATE_KEY] = requested_theme_name
        st.session_state[SETTINGS_TAB_PENDING_STATE_KEY] = "theme"
        if reopen_settings:
            st.session_state[OPEN_SETTINGS_ONCE_STATE_KEY] = True
        save_app_settings(
            workbook_path=st.session_state[WORKBOOK_PATH_STATE_KEY],
            backup_dir=st.session_state[BACKUP_DIR_STATE_KEY],
            theme=requested_theme_name,
            update_manifest_url=str(
                st.session_state.get(UPDATE_MANIFEST_URL_STATE_KEY, DEFAULT_UPDATE_MANIFEST_URL)
            ).strip(),
        )

    requested_theme = get_query_param_text("apply_theme")
    if requested_theme:
        apply_theme_selection(requested_theme, reopen_settings=True)
        clear_query_param("apply_theme")
        st.rerun()

    @st.dialog("Settings", width="large")
    def show_settings_dialog() -> None:
        tab_items: list[tuple[str, str]] = [
            ("workbook", ":material/folder_open: Workbook"),
            ("locations", ":material/location_on: Locations"),
            ("theme", ":material/palette: Theme"),
            ("diagnostics", ":material/health_and_safety: Diagnostics"),
            ("about", ":material/info: About"),
        ]
        tab_keys = [key for key, _ in tab_items]
        tab_labels_by_key = {key: label for key, label in tab_items}

        pending_tab_key = str(st.session_state.pop(SETTINGS_TAB_PENDING_STATE_KEY, "")).strip().lower()
        if pending_tab_key in tab_keys:
            st.session_state[SETTINGS_TAB_STATE_KEY] = pending_tab_key

        current_tab_key = str(st.session_state.get(SETTINGS_TAB_STATE_KEY, "workbook")).strip().lower()
        if current_tab_key not in tab_keys:
            current_tab_key = "workbook"
            st.session_state[SETTINGS_TAB_STATE_KEY] = current_tab_key

        selected_tab = st.radio(
            "Settings section",
            options=tab_keys,
            key=SETTINGS_TAB_STATE_KEY,
            format_func=lambda option_key: tab_labels_by_key.get(option_key, str(option_key)),
            horizontal=True,
            label_visibility="collapsed",
        )

        if selected_tab == "workbook":
            st.subheader("Workbook Settings")
            current_workbook_value = str(st.session_state.get(WORKBOOK_PATH_STATE_KEY, "")).strip()
            if not current_workbook_value:
                current_workbook_value = str(DEFAULT_WORKBOOK_PATH)
                st.session_state[WORKBOOK_PATH_STATE_KEY] = current_workbook_value

            current_backup_value = str(st.session_state.get(BACKUP_DIR_STATE_KEY, "")).strip()
            if not current_backup_value:
                current_backup_value = str(Path(current_workbook_value).expanduser().parent / "PO_Backups")
                st.session_state[BACKUP_DIR_STATE_KEY] = current_backup_value
            workbook_input_state_key = "settings_workbook_path_input"
            backup_input_state_key = "settings_backup_dir_input"
            workbook_input_pending_key = "settings_workbook_path_input_pending"
            backup_input_pending_key = "settings_backup_dir_input_pending"
            backup_keep_input_state_key = "settings_backup_keep_latest_input"
            backup_keep_input_pending_key = "settings_backup_keep_latest_input_pending"

            pending_workbook_input = st.session_state.pop(workbook_input_pending_key, None)
            if pending_workbook_input is not None:
                st.session_state[workbook_input_state_key] = str(pending_workbook_input)

            pending_backup_input = st.session_state.pop(backup_input_pending_key, None)
            if pending_backup_input is not None:
                st.session_state[backup_input_state_key] = str(pending_backup_input)

            pending_backup_keep = st.session_state.pop(backup_keep_input_pending_key, None)
            if pending_backup_keep is not None:
                st.session_state[backup_keep_input_state_key] = normalize_backup_keep_latest(pending_backup_keep)

            if workbook_input_state_key not in st.session_state:
                st.session_state[workbook_input_state_key] = current_workbook_value
            if backup_input_state_key not in st.session_state:
                st.session_state[backup_input_state_key] = current_backup_value
            if backup_keep_input_state_key not in st.session_state:
                st.session_state[backup_keep_input_state_key] = normalize_backup_keep_latest(
                    st.session_state.get(BACKUP_KEEP_LATEST_STATE_KEY, DEFAULT_BACKUP_KEEP_LATEST)
                )

            def persist_workbook_settings(workbook_text: str, backup_text: str, backup_keep_latest: int) -> None:
                normalized_keep_latest = normalize_backup_keep_latest(backup_keep_latest)
                st.session_state[WORKBOOK_PATH_STATE_KEY] = workbook_text
                st.session_state[BACKUP_DIR_STATE_KEY] = backup_text
                st.session_state[BACKUP_KEEP_LATEST_STATE_KEY] = normalized_keep_latest
                st.session_state[workbook_input_pending_key] = workbook_text
                st.session_state[backup_input_pending_key] = backup_text
                st.session_state[backup_keep_input_pending_key] = normalized_keep_latest
                save_app_settings(
                    workbook_path=st.session_state[WORKBOOK_PATH_STATE_KEY],
                    backup_dir=st.session_state[BACKUP_DIR_STATE_KEY],
                    backup_keep_latest=st.session_state[BACKUP_KEEP_LATEST_STATE_KEY],
                )

            st.caption("Manage workbook and backup paths used by POtrol.")

            with st.container(border=True):
                st.markdown("**Paths**")
                st.caption("Use local or network paths. Workbook path must target an Excel file.")

                workbook_col, workbook_browse_col = st.columns([5.4, 1.15], gap="small")
                with workbook_col:
                    st.text_input(
                        "Workbook path",
                        key=workbook_input_state_key,
                        placeholder=str(DEFAULT_WORKBOOK_PATH),
                    )
                with workbook_browse_col:
                    st.markdown("<div style='height: 1.88rem;'></div>", unsafe_allow_html=True)
                    browse_workbook_clicked = st.button(
                        "Browse",
                        key="settings_browse_workbook_button",
                        use_container_width=True,
                    )

                backup_col, backup_browse_col = st.columns([5.4, 1.15], gap="small")
                with backup_col:
                    st.text_input(
                        "Backup folder",
                        key=backup_input_state_key,
                        placeholder=str(Path(current_workbook_value).expanduser().parent / "PO_Backups"),
                    )
                with backup_browse_col:
                    st.markdown("<div style='height: 1.88rem;'></div>", unsafe_allow_html=True)
                    browse_backup_clicked = st.button(
                        "Browse",
                        key="settings_browse_backup_folder_button",
                        use_container_width=True,
                    )

                keep_col, keep_hint_col = st.columns([2.3, 4.25], gap="small")
                with keep_col:
                    st.number_input(
                        "Backups to keep",
                        min_value=MIN_BACKUP_KEEP_LATEST,
                        max_value=MAX_BACKUP_KEEP_LATEST,
                        step=1,
                        key=backup_keep_input_state_key,
                        help="How many recent workbook backups to retain.",
                    )
                with keep_hint_col:
                    st.markdown("<div style='height: 1.95rem;'></div>", unsafe_allow_html=True)
                    st.caption("Older backups are deleted automatically after each save.")

                action_col_1, action_col_2 = st.columns(2, gap="small")
                with action_col_1:
                    apply_paths_clicked = st.button(
                        "Save Path Changes",
                        key="settings_apply_path_changes_button",
                        type="primary",
                        use_container_width=True,
                    )
                with action_col_2:
                    use_default_backup_clicked = st.button(
                        "Use Default Backup Folder",
                        key="settings_use_default_backup_button",
                        use_container_width=True,
                    )

            if use_default_backup_clicked:
                workbook_seed_text = str(st.session_state.get(workbook_input_state_key, "")).strip()
                workbook_seed_path = Path(workbook_seed_text or current_workbook_value).expanduser()
                st.session_state[backup_input_pending_key] = str(workbook_seed_path.parent / "PO_Backups")
                st.rerun()

            if browse_workbook_clicked:
                workbook_seed_text = str(st.session_state.get(workbook_input_state_key, "")).strip()
                workbook_seed_path = Path(workbook_seed_text or current_workbook_value).expanduser()
                current_backup_text = str(st.session_state.get(backup_input_state_key, "")).strip()
                current_default_backup = str(workbook_seed_path.parent / "PO_Backups")
                selected_workbook = browse_workbook_file(workbook_seed_path)
                if selected_workbook is not None:
                    selected_workbook_text = str(selected_workbook)
                    validation_error = validate_workbook_input(selected_workbook_text, selected_workbook)
                    if validation_error:
                        st.warning(validation_error)
                    else:
                        if not current_backup_text or path_key(Path(current_backup_text)) == path_key(
                            Path(current_default_backup)
                        ):
                            selected_backup_text = str(selected_workbook.parent / "PO_Backups")
                        else:
                            selected_backup_text = current_backup_text
                        persist_workbook_settings(
                            selected_workbook_text,
                            selected_backup_text,
                            int(st.session_state.get(backup_keep_input_state_key, DEFAULT_BACKUP_KEEP_LATEST)),
                        )
                        load_sheet_data.clear()
                        st.success("Workbook path updated.")
                        st.rerun()

            if browse_backup_clicked:
                backup_seed_text = str(st.session_state.get(backup_input_state_key, "")).strip()
                backup_seed_path = Path(backup_seed_text or current_backup_value).expanduser()
                selected_backup_folder = browse_folder(backup_seed_path)
                if selected_backup_folder is not None:
                    selected_backup_text = str(selected_backup_folder)
                    workbook_for_backup = str(st.session_state.get(WORKBOOK_PATH_STATE_KEY, current_workbook_value))
                    persist_workbook_settings(
                        workbook_for_backup,
                        selected_backup_text,
                        int(st.session_state.get(backup_keep_input_state_key, DEFAULT_BACKUP_KEEP_LATEST)),
                    )
                    st.success("Backup folder updated.")
                    st.rerun()

            if apply_paths_clicked:
                new_workbook_text = str(st.session_state.get(workbook_input_state_key, "")).strip()
                new_backup_text = str(st.session_state.get(backup_input_state_key, "")).strip()
                new_backup_keep = normalize_backup_keep_latest(
                    st.session_state.get(backup_keep_input_state_key, DEFAULT_BACKUP_KEEP_LATEST)
                )

                if not new_workbook_text:
                    st.warning("Workbook path cannot be blank.")
                else:
                    new_workbook_path = Path(new_workbook_text).expanduser()
                    validation_error = validate_workbook_input(new_workbook_text, new_workbook_path)
                    if validation_error:
                        st.warning(validation_error)
                    else:
                        normalized_workbook_text = str(new_workbook_path)
                        if new_backup_text:
                            normalized_backup_text = str(Path(new_backup_text).expanduser())
                        else:
                            normalized_backup_text = str(new_workbook_path.parent / "PO_Backups")
                        persist_workbook_settings(
                            normalized_workbook_text,
                            normalized_backup_text,
                            new_backup_keep,
                        )
                        load_sheet_data.clear()
                        st.success("Workbook settings updated.")
                        st.rerun()

            workbook_path_for_backup_text = str(
                st.session_state.get(workbook_input_state_key, current_workbook_value)
            ).strip() or current_workbook_value
            backup_dir_for_backup_text = str(
                st.session_state.get(backup_input_state_key, current_backup_value)
            ).strip()
            workbook_in_settings = Path(workbook_path_for_backup_text).expanduser()
            if not backup_dir_for_backup_text:
                backup_dir_for_backup_text = str(workbook_in_settings.parent / "PO_Backups")
            backup_dir_in_settings = Path(backup_dir_for_backup_text).expanduser()

            with st.container(border=True):
                st.markdown("**Backup Management**")
                st.caption(f"Workbook: `{workbook_in_settings}`")
                st.caption(f"Backup folder: `{backup_dir_in_settings}`")
                configured_keep_latest = normalize_backup_keep_latest(
                    st.session_state.get(backup_keep_input_state_key, DEFAULT_BACKUP_KEEP_LATEST)
                )
                backup_files = list_backups(workbook_in_settings, backup_dir_in_settings)
                if backup_files:
                    st.caption(f"Found {len(backup_files)} backup(s). Latest: `{backup_files[0].name}`")
                    restore_backup_name_key = "settings_restore_backup_name"
                    backup_file_names = [backup_file.name for backup_file in backup_files]
                    if (
                        restore_backup_name_key not in st.session_state
                        or st.session_state.get(restore_backup_name_key) not in backup_file_names
                    ):
                        st.session_state[restore_backup_name_key] = backup_file_names[0]

                    selected_backup_name = st.selectbox(
                        "Select backup to restore",
                        options=backup_file_names,
                        key=restore_backup_name_key,
                    )

                    restore_col_selected, restore_col_latest = st.columns(2, gap="small")
                    with restore_col_selected:
                        restore_selected_clicked = st.button(
                            "Restore Selected Backup",
                            key="settings_restore_selected_backup_button",
                            use_container_width=True,
                        )
                    with restore_col_latest:
                        restore_latest_clicked = st.button(
                            "Restore Latest Backup",
                            key="settings_restore_latest_backup_button",
                            use_container_width=True,
                        )

                    if restore_selected_clicked or restore_latest_clicked:
                        if restore_latest_clicked:
                            target_backup = backup_files[0]
                        else:
                            target_backup = next(
                                (
                                    backup_file
                                    for backup_file in backup_files
                                    if backup_file.name == selected_backup_name
                                ),
                                None,
                            )
                        try:
                            if target_backup is None:
                                st.info("No backup found to restore.")
                            else:
                                restored_from = restore_backup(
                                    workbook_in_settings,
                                    backup_dir_in_settings,
                                    target_backup,
                                )
                                if restored_from is None:
                                    st.info("No backup found to restore.")
                                else:
                                    load_sheet_data.clear()
                                    st.success(f"Restored workbook from `{restored_from.name}`.")
                                    st.rerun()
                        except Exception as exc:
                            log_runtime_error("settings.restore_backup", exc)
                            st.error(f"Restore failed: {exc}")
                else:
                    st.caption("No backups found yet.")
                st.caption(
                    f"Backup retention: keeping the most recent {configured_keep_latest} backup(s)."
                )
        elif selected_tab == "locations":
            st.subheader("Location Settings")
            new_location = st.text_input(
                "Add location code",
                placeholder="e.g. DAL",
                key="location_add_code",
            )
            if st.button(
                "Add Location",
                key="settings_add_location_button",
                use_container_width=True,
            ):
                normalized_location = normalize_location_code(new_location)
                existing_locations = st.session_state["location_options"]
                if not normalized_location:
                    st.warning("Enter a location code first.")
                elif normalized_location in existing_locations:
                    st.info(f"`{normalized_location}` already exists.")
                else:
                    updated_locations = sorted(existing_locations + [normalized_location])
                    st.session_state["location_options"] = updated_locations
                    save_location_options(updated_locations)
                    st.success(f"Added `{normalized_location}`.")
                    st.rerun()

            current_locations = st.session_state["location_options"]
            st.caption("Current locations")
            st.write(", ".join(current_locations))
        elif selected_tab == "theme":
            st.subheader("Theme")
            theme_names = sorted(THEME_PRESETS.keys(), key=lambda value: value.casefold())
            theme_preview_palettes = {
                theme_name: resolve_theme_palette(theme_name) for theme_name in theme_names
            }
            current_theme_name = str(st.session_state.get(THEME_STATE_KEY, DEFAULT_THEME_NAME)).strip()
            if current_theme_name not in theme_names:
                current_theme_name = DEFAULT_THEME_NAME
            st.caption("Click any theme to apply it instantly.")
            controls_col_1, controls_col_2 = st.columns([2.4, 1.7], gap="small")
            with controls_col_1:
                theme_search_text = st.text_input(
                    "Find Theme",
                    key="settings_theme_search_text",
                    placeholder="Search by theme name...",
                )
            with controls_col_2:
                theme_filter = st.radio(
                    "Theme filter",
                    options=["All", "Light", "Dark"],
                    horizontal=True,
                    key="settings_theme_filter_mode",
                    label_visibility="collapsed",
                )

            def get_theme_scheme(theme_name: str) -> str:
                raw_scheme = str(theme_preview_palettes[theme_name].get("color_scheme", "light")).strip().lower()
                return "dark" if raw_scheme == "dark" else "light"

            def prioritize_active_theme(theme_list: list[str]) -> list[str]:
                ordered = sorted(theme_list, key=lambda value: value.casefold())
                if current_theme_name in ordered:
                    active_index = ordered.index(current_theme_name)
                    ordered.insert(0, ordered.pop(active_index))
                return ordered

            search_token = str(theme_search_text).strip().casefold()
            filtered_theme_names: list[str] = []
            for theme_name in theme_names:
                if search_token and search_token not in theme_name.casefold():
                    continue
                scheme_name = get_theme_scheme(theme_name)
                if theme_filter != "All" and scheme_name != theme_filter.casefold():
                    continue
                filtered_theme_names.append(theme_name)

            light_theme_names = prioritize_active_theme(
                [theme_name for theme_name in filtered_theme_names if get_theme_scheme(theme_name) == "light"]
            )
            dark_theme_names = prioritize_active_theme(
                [theme_name for theme_name in filtered_theme_names if get_theme_scheme(theme_name) == "dark"]
            )
            st.caption(f"Showing {len(filtered_theme_names)} of {len(theme_names)} themes.")

            def render_theme_group(section_label: str, grouped_theme_names: list[str], scheme_label: str) -> None:
                if not grouped_theme_names:
                    return

                st.markdown(
                    f"<div class='potrol-theme-section-title'>{section_label} ({len(grouped_theme_names)})</div>",
                    unsafe_allow_html=True,
                )

                cards_per_row = 3
                for start_index in range(0, len(grouped_theme_names), cards_per_row):
                    row_theme_names = grouped_theme_names[start_index : start_index + cards_per_row]
                    row_columns = st.columns(cards_per_row, gap="small")

                    for column_index, column in enumerate(row_columns):
                        if column_index >= len(row_theme_names):
                            with column:
                                st.empty()
                            continue

                        theme_name = row_theme_names[column_index]
                        preview_theme = theme_preview_palettes[theme_name]
                        preview_scheme = get_theme_scheme(theme_name)
                        is_active_theme = theme_name == current_theme_name
                        active_class = " potrol-theme-card-active" if is_active_theme else ""
                        preview_text_color = "#f5f8ff" if preview_scheme == "dark" else "#151922"
                        preview_pill_background = (
                            "rgba(255, 255, 255, 0.18)"
                            if preview_scheme == "dark"
                            else "rgba(17, 24, 39, 0.08)"
                        )
                        preview_pill_border = (
                            "rgba(255, 255, 255, 0.46)"
                            if preview_scheme == "dark"
                            else "rgba(17, 24, 39, 0.24)"
                        )
                        scheme_pill_html = (
                            f"<span class='potrol-theme-pill' style='color:{preview_text_color}; "
                            f"border-color:{preview_pill_border}; background:{preview_pill_background};'>{scheme_label}</span>"
                        )
                        active_pill_html = (
                            (
                                f"<span class='potrol-theme-pill' style='color:{preview_text_color}; "
                                f"border-color:{preview_pill_border}; background:{preview_pill_background};'>Active</span>"
                            )
                            if is_active_theme
                            else ""
                        )

                        with column:
                            theme_card_html = (
                                f'<div class="potrol-theme-card{active_class}" '
                                f'style="background:linear-gradient(135deg, {preview_theme["bg_start"]} 0%, {preview_theme["bg_end"]} 100%); '
                                f'border-color:{preview_theme["border"]}; --preview-text:{preview_text_color}; --preview-border:{preview_theme["border"]};">'
                                f'<div class="potrol-theme-bar" style="background:linear-gradient(120deg, {preview_theme["accent"]} 0%, {preview_theme["accent_strong"]} 100%);"></div>'
                                '<div class="potrol-theme-head">'
                                f'<span class="potrol-theme-name">{theme_name}</span>'
                                "</div>"
                                '<div class="potrol-theme-meta">'
                                f"{scheme_pill_html}"
                                f"{active_pill_html}"
                                "</div>"
                                '<div class="potrol-theme-swatches">'
                                f'<span class="potrol-theme-swatch" style="background:{preview_theme["accent"]};" title="Accent"></span>'
                                f'<span class="potrol-theme-swatch" style="background:{preview_theme["accent_strong"]};" title="Accent Strong"></span>'
                                f'<span class="potrol-theme-swatch" style="background:{preview_theme["surface"]};" title="Surface"></span>'
                                f'<span class="potrol-theme-swatch" style="background:{preview_theme["surface_soft"]};" title="Surface Soft"></span>'
                                f'<span class="potrol-theme-swatch" style="background:{preview_theme["border"]};" title="Border"></span>'
                                "</div>"
                                "</div>"
                            )
                            if is_active_theme:
                                card_markup = theme_card_html
                            else:
                                apply_theme_href = f"?apply_theme={quote(theme_name, safe='')}"
                                card_markup = (
                                    f'<a class="potrol-theme-card-link" href="{apply_theme_href}" '
                                    f'title="Apply {theme_name}">{theme_card_html}</a>'
                                )
                            st.markdown(card_markup, unsafe_allow_html=True)

            if not filtered_theme_names:
                st.info("No themes match that search/filter.")
            else:
                if theme_filter == "All":
                    render_theme_group("Light Themes", light_theme_names, "Light")
                    render_theme_group("Dark Themes", dark_theme_names, "Dark")
                elif theme_filter == "Light":
                    render_theme_group("Light Themes", light_theme_names, "Light")
                else:
                    render_theme_group("Dark Themes", dark_theme_names, "Dark")
        elif selected_tab == "diagnostics":
            st.subheader("Diagnostics")
            st.text_input(
                "Update manifest URL (optional)",
                key=UPDATE_MANIFEST_URL_STATE_KEY,
                placeholder="https://example.com/potrol-update.json",
            )
            if st.button(
                "Save Update URL",
                key="settings_save_update_url_button",
                use_container_width=True,
            ):
                update_manifest_url_value = str(
                    st.session_state.get(UPDATE_MANIFEST_URL_STATE_KEY, "")
                ).strip()
                save_app_settings(
                    workbook_path=st.session_state[WORKBOOK_PATH_STATE_KEY],
                    backup_dir=st.session_state[BACKUP_DIR_STATE_KEY],
                    update_manifest_url=update_manifest_url_value,
                )
                st.success("Update URL saved.")

            if st.button(
                "Check for Updates",
                key="settings_check_updates_button",
                use_container_width=True,
            ):
                update_url = str(st.session_state.get(UPDATE_MANIFEST_URL_STATE_KEY, "")).strip()
                if not update_url:
                    st.info("Enter an update manifest URL to check for updates.")
                else:
                    try:
                        manifest = fetch_update_manifest(update_url)
                        available_version = manifest["version"]
                        if is_version_newer(available_version, APP_VERSION):
                            st.success(f"Update available: {available_version} (current: {APP_VERSION})")
                            if manifest["download_url"]:
                                st.link_button(
                                    "Download Update",
                                    manifest["download_url"],
                                    use_container_width=True,
                                )
                            if manifest["notes"]:
                                st.caption(manifest["notes"])
                        else:
                            st.info(f"You are up to date ({APP_VERSION}).")
                    except URLError as exc:
                        log_runtime_error("diagnostics.check_updates.network", exc)
                        st.error(f"Update check failed: {exc.reason}")
                    except Exception as exc:
                        log_runtime_error("diagnostics.check_updates", exc)
                        st.error(f"Update check failed: {exc}")

            diagnostics_payload = build_diagnostics_payload(
                workbook_path=Path(st.session_state[WORKBOOK_PATH_STATE_KEY]).expanduser(),
                sheet_name=str(st.session_state.get(SHEET_SELECT_STATE_KEY, "")),
                theme_name=str(st.session_state.get(THEME_STATE_KEY, DEFAULT_THEME_NAME)),
                update_manifest_url=str(st.session_state.get(UPDATE_MANIFEST_URL_STATE_KEY, "")).strip(),
                backup_keep_latest=normalize_backup_keep_latest(
                    st.session_state.get(BACKUP_KEEP_LATEST_STATE_KEY, DEFAULT_BACKUP_KEEP_LATEST)
                ),
            )
            diagnostics_text = json.dumps(diagnostics_payload, indent=2)
            st.code(diagnostics_text, language="json")
            st.download_button(
                "Download Diagnostics JSON",
                data=diagnostics_text.encode("utf-8"),
                file_name=f"potrol_diagnostics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json",
                use_container_width=True,
            )

            runtime_log_lines = read_runtime_log_tail(max_lines=200)
            with st.container(border=True):
                st.markdown("**Runtime Log**")
                st.caption(f"Log file: `{APP_RUNTIME_LOG_PATH}`")
                if runtime_log_lines:
                    st.code("\n".join(runtime_log_lines), language="text")
                else:
                    st.caption("No runtime errors logged yet.")

                log_action_col_1, log_action_col_2 = st.columns(2, gap="small")
                with log_action_col_1:
                    runtime_log_text = "\n".join(runtime_log_lines)
                    st.download_button(
                        "Download Runtime Log",
                        data=runtime_log_text.encode("utf-8"),
                        file_name=f"potrol_runtime_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                        mime="text/plain",
                        use_container_width=True,
                    )
                with log_action_col_2:
                    if st.button(
                        "Clear Runtime Log",
                        key="settings_clear_runtime_log_button",
                        use_container_width=True,
                    ):
                        clear_runtime_log()
                        st.success("Runtime log cleared.")
                        st.rerun()
        else:
            st.subheader("About")
            st.write("Developer: Corry Holt")
            st.write("Email: corryrholt@gmail.com")
            st.caption("POtrol desktop purchase order tracker.")

    header_col, menu_col = st.columns([10, 2], gap="small")
    with header_col:
        render_logo_image(APP_LOGO_PATH, width=360, palette=theme_palette)
    with menu_col:
        with st.popover("Menu", use_container_width=False):
            open_settings = st.button(
                "Settings",
                key="open_settings_menu_button",
                use_container_width=True,
            )
            open_about = st.button(
                "About",
                key="open_about_menu_button",
                use_container_width=True,
            )

    if open_settings:
        st.session_state[SETTINGS_TAB_STATE_KEY] = "workbook"
        show_settings_dialog()
    elif open_about:
        st.session_state[SETTINGS_TAB_STATE_KEY] = "about"
        show_settings_dialog()
    elif st.session_state.pop(OPEN_SETTINGS_ONCE_STATE_KEY, False):
        show_settings_dialog()

    workbook_path_text = str(st.session_state.get(WORKBOOK_PATH_STATE_KEY, "")).strip()
    workbook_path = Path(workbook_path_text).expanduser()

    workbook_input_error = validate_workbook_input(workbook_path_text, workbook_path)
    if workbook_input_error:
        st.error(workbook_input_error)
        st.stop()

    if not str(st.session_state.get(BACKUP_DIR_STATE_KEY, "")).strip():
        st.session_state[BACKUP_DIR_STATE_KEY] = str(workbook_path.parent / "PO_Backups")
        save_app_settings(
            workbook_path=st.session_state[WORKBOOK_PATH_STATE_KEY],
            backup_dir=st.session_state[BACKUP_DIR_STATE_KEY],
        )
    backup_dir = Path(st.session_state[BACKUP_DIR_STATE_KEY]).expanduser()
    keep_backups = normalize_backup_keep_latest(
        st.session_state.get(BACKUP_KEEP_LATEST_STATE_KEY, DEFAULT_BACKUP_KEEP_LATEST)
    )
    st.session_state[BACKUP_KEEP_LATEST_STATE_KEY] = keep_backups

    location_options = st.session_state["location_options"]
    if not location_options:
        location_options = DEFAULT_LOCATION_OPTIONS.copy()
        st.session_state["location_options"] = location_options
        save_location_options(location_options)

    if not workbook_path.exists():
        if is_network_path(workbook_path):
            st.error(
                "Workbook path is on a network share and is currently unavailable. "
                "Verify network/VPN access and that the share is online."
            )
            if st.button("Retry Workbook Connection", key="retry_network_workbook_path"):
                st.rerun()
        else:
            st.warning(f"Workbook not found at `{workbook_path}`.")
            st.markdown("Create it now to start using the tracker.")

            new_sheet_name = st.text_input("Sheet name", value=DEFAULT_SHEET_NAME)
            st.caption("POtrol will create the workbook using the required PO columns.")

            if st.button("Create Workbook", type="primary"):
                create_workbook(
                    workbook_path,
                    new_sheet_name.strip() or DEFAULT_SHEET_NAME,
                    DEFAULT_HEADERS.copy(),
                )
                load_sheet_data.clear()
                st.success(f"Workbook created: `{workbook_path}`")
                st.rerun()
        st.stop()

    try:
        sheet_names = get_sheet_names(workbook_path)
    except InvalidFileException:
        st.error(
            "The selected workbook path is not a supported Excel workbook. "
            "Use .xlsx, .xlsm, .xltx, or .xltm."
        )
        st.stop()
    except Exception as exc:
        log_runtime_error("workbook.open_sheet_names", exc)
        if is_network_path(workbook_path):
            st.error(
                "Could not open workbook on the network path. "
                "Check network access, file locks, and permissions, then retry."
            )
        else:
            st.error(f"Could not open workbook: {exc}")
        if st.button("Retry Workbook Access", key="retry_open_workbook_button"):
            st.rerun()
        st.stop()
    if not sheet_names:
        st.error("Workbook has no sheets. Add a sheet in Excel and reload.")
        st.stop()

    default_sheet_name = choose_default_sheet_name(sheet_names, year=date.today().year)
    selected_workbook = st.session_state.get(SHEET_SELECT_WORKBOOK_STATE_KEY)
    if selected_workbook != str(workbook_path):
        st.session_state[SHEET_SELECT_WORKBOOK_STATE_KEY] = str(workbook_path)
        st.session_state[SHEET_SELECT_STATE_KEY] = default_sheet_name
    elif st.session_state.get(SHEET_SELECT_STATE_KEY) not in sheet_names:
        st.session_state[SHEET_SELECT_STATE_KEY] = default_sheet_name

    sheet_name = st.selectbox("Worksheet", options=sheet_names, key=SHEET_SELECT_STATE_KEY)
    try:
        headers, rows, row_numbers = load_sheet_data(str(workbook_path), sheet_name)
    except InvalidFileException:
        st.error(
            "The selected workbook path is not a supported Excel workbook. "
            "Use .xlsx, .xlsm, .xltx, or .xltm."
        )
        st.stop()
    except Exception as exc:
        log_runtime_error("workbook.load_sheet_data", exc)
        if is_network_path(workbook_path):
            st.error(
                "Could not read worksheet data from the network path. "
                "Check connectivity and permissions, then retry."
            )
        else:
            st.error(f"Could not read worksheet data: {exc}")
        if st.button("Retry Worksheet Read", key="retry_read_sheet_button"):
            st.rerun()
        st.stop()

    if not headers:
        headers = DEFAULT_HEADERS.copy()
    if len(row_numbers) != len(rows):
        row_numbers = [row_index + 2 for row_index in range(len(rows))]
    entry_mode, header_map, write_headers = build_entry_schema(headers)

    entry_scope_token = hashlib.sha1(
        f"{str(workbook_path)}::{sheet_name}".encode("utf-8")
    ).hexdigest()[:12]
    vendor_key = field_key(sheet_name, "Vendor/Store", scope=entry_scope_token)
    department_key = field_key(sheet_name, "Department", scope=entry_scope_token)
    location_key = field_key(sheet_name, "Location", scope=entry_scope_token)
    line_items_state_key = f"line_items::{entry_scope_token}"
    shipping_cost_key = field_key(sheet_name, "Shipping Cost", scope=entry_scope_token)
    sales_tax_key = field_key(sheet_name, "Sales Tax", scope=entry_scope_token)
    purchase_reason_key = field_key(sheet_name, "Purchase Reason", scope=entry_scope_token)
    entry_reset_flag_key = f"{ENTRY_FORM_RESET_KEY_PREFIX}::{entry_scope_token}"
    reservation_po_state_key = f"reserved_po::{entry_scope_token}"
    reservation_sync_at_state_key = f"reserved_po_sync_at::{entry_scope_token}"
    draft_restored_state_key = f"draft_restored::{entry_scope_token}"
    draft_hash_state_key = f"draft_hash::{entry_scope_token}"
    draft_saved_at_state_key = f"draft_saved_at::{entry_scope_token}"
    draft_error_state_key = f"draft_error::{entry_scope_token}"
    workbook_signature_state_key = f"workbook_signature::{str(workbook_path).casefold()}"
    workbook_last_sync_state_key = f"workbook_last_sync::{str(workbook_path).casefold()}"

    entry_was_reset = bool(st.session_state.pop(entry_reset_flag_key, False))
    if entry_was_reset:
        line_item_widget_prefix = f"{line_items_state_key}::"
        keys_to_clear = [
            state_key
            for state_key in list(st.session_state.keys())
            if isinstance(state_key, str) and state_key.startswith(line_item_widget_prefix)
        ]
        for state_key in keys_to_clear:
            st.session_state.pop(state_key, None)

        st.session_state[vendor_key] = ""
        st.session_state[department_key] = "IT"
        st.session_state[line_items_state_key] = default_line_items()
        st.session_state[shipping_cost_key] = 0.0
        st.session_state[sales_tax_key] = 0.0
        st.session_state[purchase_reason_key] = ""
        clear_entry_draft(workbook_path, sheet_name)
        st.session_state.pop(draft_hash_state_key, None)
        st.session_state.pop(draft_saved_at_state_key, None)
        st.session_state.pop(draft_error_state_key, None)
        st.session_state[draft_restored_state_key] = False
        if location_options:
            st.session_state[location_key] = location_options[0]

    if vendor_key not in st.session_state:
        st.session_state[vendor_key] = ""
    if department_key not in st.session_state:
        st.session_state[department_key] = "IT"
    elif st.session_state[department_key] not in DEFAULT_DEPARTMENT_OPTIONS:
        st.session_state[department_key] = "Other"
    if line_items_state_key not in st.session_state or not isinstance(
        st.session_state[line_items_state_key], list
    ):
        st.session_state[line_items_state_key] = default_line_items()
    else:
        st.session_state[line_items_state_key] = ensure_line_item_rows(
            st.session_state[line_items_state_key]
        )
    if shipping_cost_key not in st.session_state:
        st.session_state[shipping_cost_key] = 0.0
    if sales_tax_key not in st.session_state:
        st.session_state[sales_tax_key] = 0.0
    if purchase_reason_key not in st.session_state:
        st.session_state[purchase_reason_key] = ""
    if location_key not in st.session_state or st.session_state[location_key] not in location_options:
        st.session_state[location_key] = location_options[0]

    if draft_restored_state_key not in st.session_state:
        st.session_state[draft_restored_state_key] = False
    if not st.session_state[draft_restored_state_key]:
        saved_draft = load_entry_draft(workbook_path, sheet_name)
        if isinstance(saved_draft, dict):
            st.session_state[vendor_key] = str(saved_draft.get("vendor", "")).strip()
            saved_department = str(saved_draft.get("department", "IT")).strip()
            st.session_state[department_key] = (
                saved_department if saved_department in DEFAULT_DEPARTMENT_OPTIONS else "Other"
            )
            saved_location = str(saved_draft.get("location", "")).strip()
            st.session_state[location_key] = (
                saved_location if saved_location in location_options else location_options[0]
            )
            saved_items = saved_draft.get("line_items", default_line_items())
            if not isinstance(saved_items, list):
                saved_items = default_line_items()
            st.session_state[line_items_state_key] = ensure_line_item_rows(saved_items)
            st.session_state[shipping_cost_key] = round(
                parse_float(saved_draft.get("shipping_cost", 0.0), 0.0),
                2,
            )
            st.session_state[sales_tax_key] = round(
                parse_float(saved_draft.get("sales_tax", 0.0), 0.0),
                2,
            )
            st.session_state[purchase_reason_key] = str(saved_draft.get("purchase_reason", "")).strip()
            st.session_state[draft_hash_state_key] = draft_payload_hash(
                {
                    "vendor": st.session_state[vendor_key],
                    "department": st.session_state[department_key],
                    "location": st.session_state[location_key],
                    "line_items": st.session_state[line_items_state_key],
                    "shipping_cost": st.session_state[shipping_cost_key],
                    "sales_tax": st.session_state[sales_tax_key],
                    "purchase_reason": st.session_state[purchase_reason_key],
                }
            )
            st.session_state[draft_saved_at_state_key] = float(saved_draft.get("saved_at_ts", time.time()))
        st.session_state[draft_restored_state_key] = True

    if workbook_signature_state_key not in st.session_state:
        st.session_state[workbook_signature_state_key] = get_workbook_signature(workbook_path)
    if workbook_last_sync_state_key not in st.session_state:
        st.session_state[workbook_last_sync_state_key] = datetime.now().strftime("%H:%M:%S")

    def sync_reserved_po_number(force_refresh: bool = False) -> str:
        cached_po = str(st.session_state.get(reservation_po_state_key, "")).strip()
        if not cached_po:
            try:
                cached_po = get_next_po_number(
                    workbook_path,
                    sheet_name=sheet_name,
                    prefix=PO_PREFIX,
                    starting_number=PO_START_NUMBER,
                )
            except Exception:
                cached_po = f"{PO_PREFIX}{PO_START_NUMBER}"

        now_ts = time.time()
        last_sync = float(st.session_state.get(reservation_sync_at_state_key, 0.0) or 0.0)
        if not force_refresh and cached_po and (now_ts - last_sync) < PO_RESERVATION_SYNC_SECONDS:
            return cached_po

        try:
            reserved_po = reserve_session_po_number(
                path=workbook_path,
                session_id=SESSION_ID,
                owner_label=SESSION_OWNER,
                sheet_name=sheet_name,
                prefix=PO_PREFIX,
                starting_number=PO_START_NUMBER,
                stale_seconds=PO_RESERVATION_STALE_SECONDS,
                lock_timeout_seconds=5.0 if force_refresh else 2.2,
            )
            st.session_state[reservation_po_state_key] = reserved_po
        except Exception:
            st.session_state[reservation_po_state_key] = cached_po

        st.session_state[reservation_sync_at_state_key] = now_ts
        return str(st.session_state.get(reservation_po_state_key, cached_po)).strip() or cached_po

    next_po_number = sync_reserved_po_number(force_refresh=entry_was_reset)

    entry_tab, view_tab, reports_tab = st.tabs(["PO Entry", "Search, View, and Edit", "Reports"])

    with entry_tab:
        st.subheader("PO Entry")

        @st.fragment(run_every=f"{LIVE_PO_REFRESH_INTERVAL_SECONDS}s")
        def render_live_po_number() -> None:
            latest_signature = get_workbook_signature(workbook_path)
            previous_signature = str(
                st.session_state.get(workbook_signature_state_key, latest_signature)
            )
            workbook_changed = latest_signature != previous_signature
            st.session_state[workbook_signature_state_key] = latest_signature
            if workbook_changed:
                load_sheet_data.clear()
                st.session_state[workbook_last_sync_state_key] = datetime.now().strftime("%H:%M:%S")
                sync_reserved_po_number(force_refresh=True)
                st.rerun()

            refreshed_po_number = sync_reserved_po_number(force_refresh=False)
            st.text_input("PO Number", value=refreshed_po_number, disabled=True)

        render_live_po_number()
        st.caption(
            f"PO auto-sync every {LIVE_PO_REFRESH_INTERVAL_SECONDS} seconds | "
            f"Workbook sync: {st.session_state.get(workbook_last_sync_state_key, '--:--:--')}"
        )
        entry_date_value = date.today()
        st.date_input("Date", value=entry_date_value, disabled=True)
        st.text_input("Vendor/Store", key=vendor_key, placeholder="Amazon")
        st.selectbox(
            "Department",
            options=DEFAULT_DEPARTMENT_OPTIONS,
            key=department_key,
        )

        st.selectbox("Location", options=location_options, key=location_key)
        st.caption("Line items (one row per item)")
        line_items = ensure_line_item_rows(st.session_state[line_items_state_key])
        st.session_state[line_items_state_key] = line_items

        header_cols = st.columns([5, 2, 2, 1], gap="small")
        with header_cols[0]:
            st.caption("Item")
        with header_cols[1]:
            st.caption("Price Per Item")
        with header_cols[2]:
            st.caption("Quantity")
        with header_cols[3]:
            st.caption(" ")

        edited_line_items: list[dict[str, Any]] = []
        for row in line_items:
            row_id = str(row.get("Row ID", "")).strip() or uuid4().hex
            row["Row ID"] = row_id
            item_key = f"{line_items_state_key}::item::{row_id}"
            price_key = f"{line_items_state_key}::price::{row_id}"
            quantity_key = f"{line_items_state_key}::quantity::{row_id}"
            remove_key = f"{line_items_state_key}::remove::{row_id}"

            if item_key not in st.session_state:
                st.session_state[item_key] = str(row.get("Item", ""))
            if price_key not in st.session_state:
                st.session_state[price_key] = round(parse_float(row.get("Price Per Item", 0.0), 0.0), 2)
            if quantity_key not in st.session_state:
                st.session_state[quantity_key] = parse_int(row.get("Quantity", 1), 1)

            row_cols = st.columns([5, 2, 2, 1], gap="small")
            with row_cols[0]:
                st.text_input(
                    "Item",
                    key=item_key,
                    label_visibility="collapsed",
                    placeholder="e.g. USB-C to HDMI Adapter",
                )
            with row_cols[1]:
                st.number_input(
                    "Price Per Item",
                    min_value=0.0,
                    step=0.01,
                    format="%.2f",
                    key=price_key,
                    label_visibility="collapsed",
                )
            with row_cols[2]:
                st.number_input(
                    "Quantity",
                    min_value=1,
                    step=1,
                    format="%d",
                    key=quantity_key,
                    label_visibility="collapsed",
                )
            with row_cols[3]:
                remove_clicked = st.button("Remove", key=remove_key, use_container_width=True)

            if remove_clicked:
                continue

            edited_line_items.append(
                {
                    "Row ID": row_id,
                    "Item": str(st.session_state[item_key]).strip(),
                    "Price Per Item": round(parse_float(st.session_state[price_key], 0.0), 2),
                    "Quantity": parse_int(st.session_state[quantity_key], 1),
                }
            )

        if st.button("Add Item Row", key=f"{line_items_state_key}::add", use_container_width=False):
            edited_line_items.append(create_line_item_row())

        if not edited_line_items:
            edited_line_items = default_line_items()

        st.session_state[line_items_state_key] = edited_line_items
        normalized_line_items, line_item_errors = normalize_line_items(edited_line_items)

        st.number_input(
            "Shipping Cost",
            min_value=0.0,
            step=0.01,
            format="%.2f",
            key=shipping_cost_key,
        )
        st.number_input(
            "Sales Tax",
            min_value=0.0,
            step=0.01,
            format="%.2f",
            key=sales_tax_key,
        )
        st.text_input(
            "Purchase Reason (Optional)",
            key=purchase_reason_key,
            placeholder="Reason for this purchase",
        )

        shipping_cost = round(parse_float(st.session_state[shipping_cost_key], 0.0), 2)
        sales_tax = round(parse_float(st.session_state[sales_tax_key], 0.0), 2)
        sub_total = round(sum(item["Sub Total"] for item in normalized_line_items), 2)
        grand_total = round(sub_total + shipping_cost + sales_tax, 2)

        st.text_input("Sub Total", value=f"{sub_total:,.2f}", disabled=True)
        st.text_input("Grand Total", value=f"{grand_total:,.2f}", disabled=True)

        draft_line_items = ensure_line_item_rows(st.session_state[line_items_state_key])
        draft_snapshot = {
            "vendor": str(st.session_state[vendor_key]).strip(),
            "department": str(st.session_state[department_key]).strip(),
            "location": str(st.session_state[location_key]).strip(),
            "line_items": draft_line_items,
            "shipping_cost": shipping_cost,
            "sales_tax": sales_tax,
            "purchase_reason": str(st.session_state[purchase_reason_key]).strip(),
        }
        has_line_item_content = any(
            str(item.get("Item", "")).strip()
            or parse_float(item.get("Price Per Item", 0.0), 0.0) > 0
            for item in draft_line_items
        )
        has_draft_content = (
            bool(draft_snapshot["vendor"])
            or bool(draft_snapshot["purchase_reason"])
            or float(draft_snapshot["shipping_cost"]) > 0
            or float(draft_snapshot["sales_tax"]) > 0
            or has_line_item_content
        )
        current_draft_hash = draft_payload_hash(draft_snapshot)
        previous_draft_hash = str(st.session_state.get(draft_hash_state_key, "")).strip()
        last_draft_saved_at = float(st.session_state.get(draft_saved_at_state_key, 0.0) or 0.0)
        now_ts = time.time()
        if has_draft_content:
            if current_draft_hash != previous_draft_hash and (
                (now_ts - last_draft_saved_at) >= DRAFT_AUTOSAVE_MIN_SECONDS
            ):
                payload_to_save = dict(draft_snapshot)
                payload_to_save["saved_at_ts"] = now_ts
                try:
                    save_entry_draft(workbook_path, sheet_name, payload_to_save)
                    st.session_state[draft_hash_state_key] = current_draft_hash
                    st.session_state[draft_saved_at_state_key] = now_ts
                    st.session_state.pop(draft_error_state_key, None)
                except Exception as exc:
                    st.session_state[draft_error_state_key] = (
                        f"Draft autosave failed: {str(exc).strip() or 'unknown error'}"
                    )
        elif previous_draft_hash:
            try:
                clear_entry_draft(workbook_path, sheet_name)
                st.session_state.pop(draft_hash_state_key, None)
                st.session_state.pop(draft_saved_at_state_key, None)
                st.session_state.pop(draft_error_state_key, None)
            except Exception as exc:
                st.session_state[draft_error_state_key] = (
                    f"Draft cleanup failed: {str(exc).strip() or 'unknown error'}"
                )

        if st.session_state.get(draft_saved_at_state_key):
            saved_time_text = datetime.fromtimestamp(
                float(st.session_state[draft_saved_at_state_key])
            ).strftime("%I:%M:%S %p").lstrip("0")
            st.caption(f"Draft autosaved: {saved_time_text}")
        if st.session_state.get(draft_error_state_key):
            st.caption(str(st.session_state[draft_error_state_key]))

        save_clicked = st.button("Save PO", type="primary", use_container_width=True)
        if save_clicked:
            vendor_value = str(st.session_state[vendor_key]).strip()
            department_value = str(st.session_state[department_key]).strip()
            location_value = str(st.session_state[location_key]).strip()
            purchase_reason_value = str(st.session_state[purchase_reason_key]).strip()

            if not vendor_value:
                st.error("Vendor/Store is required.")
            elif line_item_errors:
                for error_text in line_item_errors:
                    st.error(error_text)
            elif not normalized_line_items:
                st.error("Add at least one item line before saving.")
            else:
                try:
                    po_number = str(st.session_state.get(reservation_po_state_key, next_po_number)).strip()
                    if not po_number:
                        po_number = next_po_number
                    if isinstance(entry_date_value, date):
                        entry_date_text = entry_date_value.strftime("%Y-%m-%d")
                    else:
                        entry_date_text = str(entry_date_value)

                    entry_values: list[dict[str, Any]] = []
                    if entry_mode == "legacy":
                        po_header = header_map["PO Number"]
                        date_header = header_map["Date"]
                        vendor_header = header_map["Vendor/Store"]
                        department_header = header_map["Department"]
                        location_header = header_map["Location"]
                        item_header = header_map["Items Being Purchased"]
                        price_header = header_map["Price Per Item"]
                        quantity_header = header_map["Quantity"]
                        sub_total_header = header_map["Sub Total"]
                        grand_total_header = header_map["Grand Total"]

                        if department_value and location_value:
                            if department_value.casefold() == location_value.casefold():
                                department_location_value = location_value
                            else:
                                department_location_value = f"{location_value}/{department_value}"
                        else:
                            department_location_value = location_value or department_value

                        for item_index, line_item in enumerate(normalized_line_items):
                            row_values: dict[str, Any] = {header: "" for header in write_headers}
                            if item_index == 0:
                                row_values[po_header] = po_number
                                row_values[date_header] = entry_date_text
                                row_values[vendor_header] = vendor_value
                                if department_header == location_header:
                                    row_values[department_header] = department_location_value
                                else:
                                    row_values[department_header] = department_value
                                    row_values[location_header] = location_value
                            row_values[item_header] = line_item["Items Being Purchased"]
                            row_values[price_header] = line_item["Price Per Item"]
                            row_values[quantity_header] = line_item["Quantity"]
                            row_values[sub_total_header] = line_item["Sub Total"]
                            row_values[grand_total_header] = ""
                            entry_values.append(row_values)

                        if shipping_cost > 0:
                            shipping_row: dict[str, Any] = {header: "" for header in write_headers}
                            shipping_row[item_header] = "Shipping Cost"
                            shipping_row[price_header] = shipping_cost
                            shipping_row[quantity_header] = 1
                            shipping_row[sub_total_header] = shipping_cost
                            shipping_row[grand_total_header] = ""
                            entry_values.append(shipping_row)

                        if sales_tax > 0:
                            tax_row: dict[str, Any] = {header: "" for header in write_headers}
                            tax_row[item_header] = "Tax"
                            tax_row[price_header] = sales_tax
                            tax_row[quantity_header] = 1
                            tax_row[sub_total_header] = sales_tax
                            tax_row[grand_total_header] = ""
                            entry_values.append(tax_row)

                        entry_values[-1][grand_total_header] = grand_total
                    else:
                        for item_index, line_item in enumerate(normalized_line_items):
                            entry_values.append(
                                {
                                    "PO Number": po_number if item_index == 0 else "",
                                    "Date": entry_date_text if item_index == 0 else "",
                                    "Vendor/Store": vendor_value if item_index == 0 else "",
                                    "Department": department_value if item_index == 0 else "",
                                    "Location": location_value if item_index == 0 else "",
                                    "Items Being Purchased": line_item["Items Being Purchased"],
                                    "Price Per Item": line_item["Price Per Item"],
                                    "Quantity": line_item["Quantity"],
                                    "Sub Total": line_item["Sub Total"],
                                    "Shipping Cost": "",
                                    "Sales Tax": "",
                                    "Grand Total": "",
                                }
                            )

                        if shipping_cost > 0:
                            entry_values.append(
                                {
                                    "PO Number": "",
                                    "Date": "",
                                    "Vendor/Store": "",
                                    "Department": "",
                                    "Location": "",
                                    "Items Being Purchased": "Shipping Cost",
                                    "Price Per Item": shipping_cost,
                                    "Quantity": 1,
                                    "Sub Total": shipping_cost,
                                    "Shipping Cost": shipping_cost,
                                    "Sales Tax": "",
                                    "Grand Total": "",
                                }
                            )

                        if sales_tax > 0:
                            entry_values.append(
                                {
                                    "PO Number": "",
                                    "Date": "",
                                    "Vendor/Store": "",
                                    "Department": "",
                                    "Location": "",
                                    "Items Being Purchased": "Tax",
                                    "Price Per Item": sales_tax,
                                    "Quantity": 1,
                                    "Sub Total": sales_tax,
                                    "Shipping Cost": "",
                                    "Sales Tax": sales_tax,
                                    "Grand Total": "",
                                }
                            )

                        entry_values[-1]["Grand Total"] = grand_total

                    with workbook_write_lock(workbook_path):
                        candidate_po_number = str(
                            st.session_state.get(reservation_po_state_key, po_number)
                        ).strip()
                        if (
                            not candidate_po_number
                            or po_number_exists(
                                workbook_path,
                                candidate_po_number,
                                prefix=PO_PREFIX,
                                sheet_name=sheet_name,
                            )
                        ):
                            candidate_po_number = get_next_po_number(
                                workbook_path,
                                sheet_name=sheet_name,
                                prefix=PO_PREFIX,
                                starting_number=PO_START_NUMBER,
                            )
                        po_number = candidate_po_number
                        if entry_values:
                            if entry_mode == "legacy":
                                po_header = header_map["PO Number"]
                                entry_values[0][po_header] = po_number
                            else:
                                entry_values[0]["PO Number"] = po_number
                        backup_path = append_record(
                            path=workbook_path,
                            sheet_name=sheet_name,
                            headers=write_headers,
                            values=entry_values,
                            backup_dir=backup_dir,
                            keep_backups=int(keep_backups),
                            purchase_reason=purchase_reason_value if entry_mode == "legacy" else "",
                            purchase_reason_column_index=(
                                PURCHASE_REASON_COLUMN_INDEX if entry_mode == "legacy" else None
                            ),
                        )
                    load_sheet_data.clear()
                    st.session_state[workbook_signature_state_key] = get_workbook_signature(workbook_path)
                    st.session_state[workbook_last_sync_state_key] = datetime.now().strftime("%H:%M:%S")
                    try:
                        release_session_po_reservation(workbook_path, SESSION_ID)
                    except Exception:
                        pass
                    try:
                        st.session_state[reservation_po_state_key] = reserve_session_po_number(
                            path=workbook_path,
                            session_id=SESSION_ID,
                            owner_label=SESSION_OWNER,
                            sheet_name=sheet_name,
                            prefix=PO_PREFIX,
                            starting_number=PO_START_NUMBER,
                            stale_seconds=PO_RESERVATION_STALE_SECONDS,
                            lock_timeout_seconds=4.0,
                        )
                    except Exception:
                        try:
                            st.session_state[reservation_po_state_key] = get_next_po_number(
                                workbook_path,
                                sheet_name=sheet_name,
                                prefix=PO_PREFIX,
                                starting_number=PO_START_NUMBER,
                            )
                        except Exception:
                            st.session_state[reservation_po_state_key] = po_number
                    st.session_state[reservation_sync_at_state_key] = time.time()
                    clear_entry_draft(workbook_path, sheet_name)
                    st.session_state.pop(draft_hash_state_key, None)
                    st.session_state.pop(draft_saved_at_state_key, None)
                    st.session_state.pop(draft_error_state_key, None)

                    st.session_state[entry_reset_flag_key] = True

                    if backup_path is None:
                        st.success(f"PO `{po_number}` saved.")
                    else:
                        st.success(f"PO `{po_number}` saved. Backup created: `{backup_path.name}`")
                    st.rerun()
                except TimeoutError as exc:
                    log_runtime_error("entry_tab.save.timeout", exc)
                    st.error(str(exc))
                except PermissionError:
                    append_runtime_log(
                        "ERROR",
                        "entry_tab.save.permission",
                        "Could not write to workbook because the file is locked or read-only.",
                    )
                    st.error("Could not write to the workbook. Close Excel and try again.")
                except Exception as exc:
                    log_runtime_error("entry_tab.save", exc)
                    st.error(f"Save failed: {exc}")

    with view_tab:
        st.subheader("Search, View, and Edit")
        st.caption(f"Workbook sync: {st.session_state.get(workbook_last_sync_state_key, '--:--:--')}")
        editor_search_key = f"manual_editor_search::{entry_scope_token}"
        editor_scan_limit_key = f"manual_editor_scan_limit::{entry_scope_token}"
        editor_page_size_key = f"manual_editor_page_size::{entry_scope_token}"
        editor_page_key = f"manual_editor_page::{entry_scope_token}"
        editor_signature_key = f"manual_editor_signature::{entry_scope_token}"
        editor_signature_sync_key = f"manual_editor_signature_sync::{entry_scope_token}"

        if editor_scan_limit_key not in st.session_state:
            st.session_state[editor_scan_limit_key] = DEFAULT_EDITOR_SEARCH_SCAN_LIMIT
        if editor_page_size_key not in st.session_state:
            st.session_state[editor_page_size_key] = DEFAULT_EDITOR_PAGE_SIZE
        if editor_page_key not in st.session_state:
            st.session_state[editor_page_key] = 1

        signature_sync_marker = (
            f"{st.session_state.get(workbook_last_sync_state_key, '--:--:--')}::"
            f"{len(rows)}::{sheet_name.casefold()}"
        )
        if (
            editor_signature_key not in st.session_state
            or st.session_state.get(editor_signature_sync_key) != signature_sync_marker
        ):
            st.session_state[editor_signature_key] = get_workbook_signature(workbook_path)
            st.session_state[editor_signature_sync_key] = signature_sync_marker

        search_col, filter_col, scan_col = st.columns([3.55, 1.0, 1.35], gap="small")
        with search_col:
            search_query = st.text_input(
                "Search",
                placeholder="Type PO number, vendor, department, location, item...",
                key=editor_search_key,
            )
        with filter_col:
            st.markdown("<div style='height: 1.85rem;'></div>", unsafe_allow_html=True)
            newest_first = st.checkbox("Newest first", value=True)
        with scan_col:
            scan_limit_options = [1000, 2500, 5000, 10000, 25000, 0]
            st.selectbox(
                "Search scan",
                options=scan_limit_options,
                key=editor_scan_limit_key,
                format_func=lambda option: "All rows" if int(option) == 0 else f"Last {int(option):,}",
            )

        frame = pd.DataFrame(rows, columns=headers)
        if frame.empty:
            st.info("No entries yet on this worksheet.")
        else:
            configured_scan_limit = int(st.session_state.get(editor_scan_limit_key, DEFAULT_EDITOR_SEARCH_SCAN_LIMIT))
            max_scan_rows = len(frame) if configured_scan_limit == 0 else configured_scan_limit
            filtered_frame, search_truncated, scanned_rows = filter_records_lazy(
                frame,
                search_query,
                max_scan_rows=max_scan_rows,
            )
            if newest_first:
                filtered_frame = filtered_frame.iloc[::-1]

            top_info_col, top_action_col, top_refresh_col = st.columns([2.4, 1, 1], gap="small")
            with top_info_col:
                st.markdown(
                    f"<div class='potrol-stat-badge'>Total Entries: {len(frame):,}</div>",
                    unsafe_allow_html=True,
                )
            with top_action_col:
                csv_data = filtered_frame.to_csv(index=False).encode("utf-8")
                st.download_button(
                    "Download CSV",
                    data=csv_data,
                    file_name=f"{sheet_name.replace(' ', '_').lower()}_filtered.csv",
                    mime="text/csv",
                    use_container_width=True,
                )
            with top_refresh_col:
                st.markdown("<div style='height: 1.85rem;'></div>", unsafe_allow_html=True)
                reload_latest_clicked = st.button(
                    "Reload Latest",
                    key=f"manual_editor_reload::{entry_scope_token}",
                    use_container_width=True,
                )

            if reload_latest_clicked:
                latest_signature = get_workbook_signature(workbook_path)
                load_sheet_data.clear()
                st.session_state[editor_signature_key] = latest_signature
                st.session_state[workbook_signature_state_key] = latest_signature
                st.session_state[workbook_last_sync_state_key] = datetime.now().strftime("%H:%M:%S")
                st.session_state[editor_page_key] = 1
                st.rerun()

            if filtered_frame.empty:
                st.info("No matching entries for the current search.")
            else:
                if search_query.strip() and search_truncated:
                    st.caption(
                        f"Search scanned the most recent {scanned_rows:,} rows for responsiveness."
                    )

                paging_col_1, paging_col_2, paging_col_3 = st.columns([2.4, 1.4, 1.4], gap="small")
                with paging_col_1:
                    filtered_count = len(filtered_frame)
                    st.caption(f"Matching rows: {filtered_count:,}")
                with paging_col_2:
                    page_size_options = [25, 50, 100, 250, 500]
                    st.selectbox(
                        "Rows per page",
                        options=page_size_options,
                        key=editor_page_size_key,
                        format_func=lambda option: f"{int(option):,}",
                    )
                with paging_col_3:
                    page_size = max(1, int(st.session_state.get(editor_page_size_key, DEFAULT_EDITOR_PAGE_SIZE)))
                    total_pages = max(1, (len(filtered_frame) + page_size - 1) // page_size)
                    stored_page_number = max(1, int(st.session_state.get(editor_page_key, 1)))
                    if stored_page_number > total_pages:
                        stored_page_number = total_pages
                        st.session_state[editor_page_key] = stored_page_number
                    st.number_input(
                        "Page",
                        min_value=1,
                        max_value=total_pages,
                        step=1,
                        key=editor_page_key,
                    )

                page_size = max(1, int(st.session_state.get(editor_page_size_key, DEFAULT_EDITOR_PAGE_SIZE)))
                page_number = max(1, int(st.session_state.get(editor_page_key, 1)))
                total_pages = max(1, (len(filtered_frame) + page_size - 1) // page_size)
                if page_number > total_pages:
                    page_number = total_pages
                    st.session_state[editor_page_key] = page_number
                page_start = (page_number - 1) * page_size
                page_end = page_start + page_size
                page_frame = filtered_frame.iloc[page_start:page_end].copy()

                start_display = page_start + 1
                end_display = min(page_end, len(filtered_frame))
                st.caption(f"Showing rows {start_display:,}-{end_display:,} of {len(filtered_frame):,}")

                excel_row_column = "__excel_row__"
                editable_frame = page_frame.copy()
                editable_frame.insert(
                    0,
                    excel_row_column,
                    [row_numbers[int(row_index)] for row_index in page_frame.index],
                )
                editable_frame = editable_frame.reset_index(drop=True)

                editor_col, save_col = st.columns([4.4, 1.2], gap="small")
                with editor_col:
                    st.caption(
                        "Edit cells directly, add new rows, or delete rows here. "
                        "Click Save Edits to write changes."
                    )
                with save_col:
                    st.markdown("<div style='height: 1.85rem;'></div>", unsafe_allow_html=True)
                    save_edits_clicked = st.button(
                        "Save Edits",
                        use_container_width=True,
                        type="primary",
                        key=f"manual_save_edits::{entry_scope_token}",
                    )

                edited_frame = st.data_editor(
                    editable_frame,
                    use_container_width=True,
                    hide_index=True,
                    height=520,
                    key=f"manual_editor::{entry_scope_token}",
                    num_rows="dynamic",
                    disabled=[excel_row_column],
                    column_config={
                        excel_row_column: st.column_config.NumberColumn(
                            "Workbook Row",
                            help="Excel row number for this record.",
                            width="small",
                        ),
                    },
                )

                if save_edits_clicked:
                    original_row_numbers: set[int] = set()
                    original_by_row: dict[int, dict[str, Any]] = {}
                    for _, original_row in editable_frame.iterrows():
                        try:
                            row_number = int(original_row.get(excel_row_column, 0))
                        except Exception:
                            continue
                        original_row_numbers.add(row_number)
                        original_by_row[row_number] = {
                            header: normalize_editor_cell_value(original_row.get(header, ""))
                            for header in headers
                        }

                    edited_existing_row_numbers: set[int] = set()
                    row_updates: list[tuple[int, dict[str, Any]]] = []
                    new_rows: list[dict[str, Any]] = []
                    changed_cells = 0
                    for _, edited_row in edited_frame.iterrows():
                        updated_values: dict[str, Any] = {}
                        for header in headers:
                            normalized_value = normalize_editor_cell_value(edited_row.get(header, ""))
                            updated_values[header] = normalized_value

                        row_number: int | None = None
                        row_number_raw = edited_row.get(excel_row_column, "")
                        try:
                            parsed_row_number = int(row_number_raw)
                            if parsed_row_number > 1:
                                row_number = parsed_row_number
                        except Exception:
                            row_number = None

                        if row_number is None:
                            if any(has_non_empty_editor_value(updated_values.get(header, "")) for header in headers):
                                new_rows.append(updated_values)
                            continue

                        edited_existing_row_numbers.add(row_number)
                        original_values = original_by_row.get(row_number, {})
                        row_changed = False
                        for header in headers:
                            if updated_values.get(header, "") != original_values.get(header, ""):
                                row_changed = True
                                changed_cells += 1
                        if row_changed:
                            row_updates.append((row_number, updated_values))

                    row_deletes = sorted(original_row_numbers - edited_existing_row_numbers, reverse=True)
                    inserted_row_count = len(new_rows)

                    if not row_updates and not row_deletes and inserted_row_count == 0:
                        st.info("No edits detected.")
                    else:
                        try:
                            base_signature = str(st.session_state.get(editor_signature_key, "")).strip()
                            latest_signature = get_workbook_signature(workbook_path)
                            if base_signature and latest_signature != base_signature:
                                st.error(
                                    "Workbook changed since this editor view loaded. "
                                    "Reload latest data before saving edits."
                                )
                                if st.button(
                                    "Reload Latest Data",
                                    key=f"manual_editor_reload_conflict::{entry_scope_token}",
                                    use_container_width=False,
                                ):
                                    load_sheet_data.clear()
                                    st.session_state[editor_signature_key] = latest_signature
                                    st.session_state[workbook_signature_state_key] = latest_signature
                                    st.session_state[workbook_last_sync_state_key] = datetime.now().strftime("%H:%M:%S")
                                    st.session_state[editor_page_key] = 1
                                    st.rerun()
                            else:
                                with workbook_write_lock(workbook_path):
                                    backup_path = update_sheet_rows(
                                        path=workbook_path,
                                        sheet_name=sheet_name,
                                        headers=headers,
                                        row_updates=row_updates,
                                        backup_dir=backup_dir,
                                        keep_backups=int(keep_backups),
                                        row_deletes=row_deletes,
                                        new_rows=new_rows,
                                    )
                                load_sheet_data.clear()
                                latest_saved_signature = get_workbook_signature(workbook_path)
                                st.session_state[editor_signature_key] = latest_saved_signature
                                st.session_state[workbook_signature_state_key] = latest_saved_signature
                                st.session_state[workbook_last_sync_state_key] = datetime.now().strftime("%H:%M:%S")
                                st.session_state[editor_page_key] = 1

                                status_parts = [
                                    f"updated {len(row_updates)} row(s)",
                                    f"deleted {len(row_deletes)} row(s)",
                                    f"inserted {inserted_row_count} row(s)",
                                ]
                                status_text = ", ".join(status_parts)
                                if backup_path is None:
                                    st.success(
                                        f"Saved edits ({status_text}; {changed_cells} cell change(s))."
                                    )
                                else:
                                    st.success(
                                        f"Saved edits ({status_text}; {changed_cells} cell change(s)). "
                                        f"Backup created: `{backup_path.name}`"
                                    )
                                st.rerun()
                        except TimeoutError as exc:
                            log_runtime_error("view_tab.save_edits.timeout", exc)
                            st.error(str(exc))
                        except PermissionError:
                            append_runtime_log(
                                "ERROR",
                                "view_tab.save_edits.permission",
                                "Could not write edits to workbook because the file is locked or read-only.",
                            )
                            st.error("Could not write edits to the workbook. Close Excel and try again.")
                        except Exception as exc:
                            log_runtime_error("view_tab.save_edits", exc)
                            st.error(f"Edit save failed: {exc}")

    with reports_tab:
        st.subheader("Reports")
        st.caption(f"Workbook sync: {st.session_state.get(workbook_last_sync_state_key, '--:--:--')}")
        report_scope_state_key = "reports_scope_mode"
        if report_scope_state_key not in st.session_state:
            st.session_state[report_scope_state_key] = "All Worksheets"

        scope_control_col, scope_info_col = st.columns([2.2, 2.4], gap="small")
        with scope_control_col:
            report_scope_mode = st.radio(
                "Report Scope",
                options=["All Worksheets", "Current Worksheet"],
                key=report_scope_state_key,
                horizontal=True,
            )
        with scope_info_col:
            if report_scope_mode == "All Worksheets":
                st.caption(f"Including {len(sheet_names)} worksheets.")
            else:
                st.caption(f"Including only `{sheet_name}`.")

        report_sheet_names = list(sheet_names) if report_scope_mode == "All Worksheets" else [sheet_name]
        report_scope_text = "all worksheets" if report_scope_mode == "All Worksheets" else f"worksheet `{sheet_name}`"

        report_error_message = ""
        try:
            po_report_frame = build_reporting_frame_for_sheets(
                str(workbook_path),
                report_sheet_names,
                location_options=location_options,
            )
        except InvalidFileException:
            report_error_message = (
                "The selected workbook path is not a supported Excel workbook. "
                "Use .xlsx, .xlsm, .xltx, or .xltm."
            )
        except Exception as exc:
            log_runtime_error("reports_tab.load_data", exc)
            if is_network_path(workbook_path):
                report_error_message = (
                    "Could not read report data from the network path. "
                    "Check connectivity and permissions, then retry."
                )
            else:
                report_error_message = f"Could not read report data: {exc}"

        if report_error_message:
            st.error(report_error_message)
        elif po_report_frame.empty:
            st.info(f"No report data yet in {report_scope_text}.")
        else:
            total_po_count = int(len(po_report_frame))
            total_spend = float(po_report_frame["Total"].sum())
            average_po = round(total_spend / total_po_count, 2) if total_po_count else 0.0

            def build_ranked_totals(group_column: str) -> pd.DataFrame:
                grouped_frame = (
                    po_report_frame.groupby(group_column, as_index=False)["Total"]
                    .sum()
                    .reset_index(drop=True)
                )
                grouped_frame[group_column] = grouped_frame[group_column].astype(str).str.strip()
                grouped_frame = grouped_frame[grouped_frame[group_column] != ""].reset_index(drop=True)
                return grouped_frame.sort_values("Total", ascending=False).reset_index(drop=True)

            vendor_totals_all = build_ranked_totals("Vendor/Store")
            dept_loc_totals_all = build_ranked_totals("Department/Loc")
            location_totals_all = build_ranked_totals("Location")
            top_vendor_name = "N/A"
            top_vendor_spend = 0.0
            if not vendor_totals_all.empty:
                top_vendor_name = str(vendor_totals_all.loc[0, "Vendor/Store"]).strip() or "Unknown"
                top_vendor_spend = float(vendor_totals_all.loc[0, "Total"])

            def render_report_metric_card(label: str, value_text: str, note_text: str) -> None:
                st.markdown(
                    (
                        '<div class="potrol-report-metric">'
                        f'<div class="potrol-report-metric-label">{label}</div>'
                        f'<div class="potrol-report-metric-value">{value_text}</div>'
                        f'<div class="potrol-report-metric-note">{note_text}</div>'
                        "</div>"
                    ),
                    unsafe_allow_html=True,
                )

            metric_col_1, metric_col_2, metric_col_3, metric_col_4 = st.columns(4, gap="small")
            with metric_col_1:
                render_report_metric_card("Total POs", f"{total_po_count:,}", "Purchase orders in selected scope")
            with metric_col_2:
                render_report_metric_card("Total Spend", f"${total_spend:,.2f}", "Combined spend value")
            with metric_col_3:
                render_report_metric_card("Average PO", f"${average_po:,.2f}", "Average spend per order")
            with metric_col_4:
                render_report_metric_card("Top Vendor", f"${top_vendor_spend:,.2f}", top_vendor_name)

            max_ranked_items = max(len(vendor_totals_all), len(dept_loc_totals_all), len(location_totals_all))
            top_n_max = max(1, min(20, int(max_ranked_items)))
            top_n_default = min(10, top_n_max)
            top_n_state_key = "reports_top_n_items"
            stored_top_n_value: Any = st.session_state.get(top_n_state_key, top_n_default)
            try:
                stored_top_n = int(stored_top_n_value)
            except (TypeError, ValueError):
                stored_top_n = top_n_default
            stored_top_n = max(1, min(stored_top_n, top_n_max))
            st.session_state[top_n_state_key] = stored_top_n

            control_spacer_col, control_col = st.columns([3.0, 1.8], gap="small")
            with control_col:
                if top_n_max > 1:
                    top_n_items = st.slider(
                        "Top Results",
                        min_value=1,
                        max_value=top_n_max,
                        key=top_n_state_key,
                        help="Applies to Top Vendors, Department/Loc, and Locations.",
                    )
                else:
                    top_n_items = 1
                    st.caption("Top Results: 1")

            current_month_period = pd.Timestamp.now().to_period("M")
            trailing_month_periods = pd.period_range(end=current_month_period, periods=12, freq="M")
            monthly_trend_base = pd.DataFrame({"Month Period": trailing_month_periods})
            dated_report_frame = po_report_frame[po_report_frame["Date Parsed"].notna()].copy()
            if not dated_report_frame.empty:
                dated_report_frame["Month Period"] = dated_report_frame["Date Parsed"].dt.to_period("M")
                dated_report_frame = dated_report_frame[
                    dated_report_frame["Month Period"].isin(trailing_month_periods)
                ].copy()

            grouped_monthly_spend = pd.DataFrame(columns=["Month Period", "Total"])
            grouped_monthly_count = pd.DataFrame(columns=["Month Period", "PO Count"])
            grouped_monthly_average = pd.DataFrame(columns=["Month Period", "Average PO"])
            if not dated_report_frame.empty:
                grouped_monthly_spend = (
                    dated_report_frame.groupby("Month Period", as_index=False)["Total"].sum().reset_index(drop=True)
                )
                grouped_monthly_count = (
                    dated_report_frame.groupby("Month Period", as_index=False)["PO Number"]
                    .count()
                    .rename(columns={"PO Number": "PO Count"})
                    .reset_index(drop=True)
                )
                grouped_monthly_average = (
                    dated_report_frame.groupby("Month Period", as_index=False)["Total"]
                    .mean()
                    .rename(columns={"Total": "Average PO"})
                    .reset_index(drop=True)
                )

            monthly_spend = (
                monthly_trend_base.merge(grouped_monthly_spend, on="Month Period", how="left")
                .fillna({"Total": 0.0})
                .reset_index(drop=True)
            )
            monthly_po_count = (
                monthly_trend_base.merge(grouped_monthly_count, on="Month Period", how="left")
                .fillna({"PO Count": 0.0})
                .reset_index(drop=True)
            )
            monthly_average_po = (
                monthly_trend_base.merge(grouped_monthly_average, on="Month Period", how="left")
                .fillna({"Average PO": 0.0})
                .reset_index(drop=True)
            )

            for trend_frame in [monthly_spend, monthly_po_count, monthly_average_po]:
                trend_frame["Month Sort"] = trend_frame["Month Period"].dt.to_timestamp()
                trend_frame.sort_values("Month Sort", inplace=True)
                trend_frame["Month"] = trend_frame["Month Period"].apply(
                    lambda period_value: period_value.strftime("%b %Y")
                )

            monthly_spend = monthly_spend[["Month", "Total"]].reset_index(drop=True)
            monthly_po_count = monthly_po_count[["Month", "PO Count"]].reset_index(drop=True)
            monthly_average_po = monthly_average_po[["Month", "Average PO"]].reset_index(drop=True)
            monthly_cumulative_spend = monthly_spend.copy()
            monthly_cumulative_spend["Cumulative Spend"] = monthly_cumulative_spend["Total"].cumsum()
            monthly_cumulative_spend = monthly_cumulative_spend[["Month", "Cumulative Spend"]]

            vendor_spend = vendor_totals_all.head(int(top_n_items)).reset_index(drop=True)
            dept_loc_spend = dept_loc_totals_all.head(int(top_n_items)).reset_index(drop=True)
            location_spend = location_totals_all.head(int(top_n_items)).reset_index(drop=True)

            def render_report_line_chart(
                source_frame: pd.DataFrame,
                index_column: str,
                title: str,
                caption_text: str,
                tooltip_title: str,
                label_angle: int = -24,
                value_column: str = "Total",
                y_axis_title: str = "Spend",
                y_axis_format: str = "$,.0f",
                tooltip_value_title: str = "Total Spend",
                tooltip_value_format: str = "$,.2f",
                chart_height: int = 286,
            ) -> None:
                with st.container(border=True):
                    st.markdown(f"<div class='potrol-report-card-title'>{title}</div>", unsafe_allow_html=True)
                    st.markdown(f"<div class='potrol-report-card-sub'>{caption_text}</div>", unsafe_allow_html=True)
                    if source_frame.empty:
                        st.caption("No data available yet.")
                        return

                    if index_column not in source_frame.columns or value_column not in source_frame.columns:
                        st.caption("No data available yet.")
                        return

                    chart_frame = source_frame[[index_column, value_column]].copy()
                    chart_frame[index_column] = chart_frame[index_column].astype(str).str.strip()
                    chart_frame[value_column] = pd.to_numeric(chart_frame[value_column], errors="coerce")
                    chart_frame = chart_frame[chart_frame[index_column] != ""].reset_index(drop=True)
                    chart_frame = chart_frame[chart_frame[value_column].notna()].reset_index(drop=True)
                    if chart_frame.empty:
                        st.caption("No data available yet.")
                        return

                    sort_order = chart_frame[index_column].tolist()
                    chart_spec = {
                        "$schema": "https://vega.github.io/schema/vega-lite/v5.json",
                        "layer": [
                            {
                                "mark": {
                                    "type": "area",
                                    "interpolate": "monotone",
                                    "color": theme_palette["accent"],
                                    "opacity": 0.16,
                                }
                            },
                            {
                                "mark": {
                                    "type": "line",
                                    "interpolate": "monotone",
                                    "strokeWidth": 3,
                                    "color": theme_palette["accent"],
                                }
                            },
                            {
                                "mark": {
                                    "type": "point",
                                    "filled": True,
                                    "size": 62,
                                    "fill": theme_palette["surface"],
                                    "stroke": theme_palette["accent_strong"],
                                    "strokeWidth": 2,
                                }
                            },
                        ],
                        "encoding": {
                            "x": {
                                "field": index_column,
                                "type": "ordinal",
                                "sort": sort_order,
                                "axis": {
                                    "title": "",
                                    "labelAngle": label_angle,
                                    "labelColor": theme_palette["muted"],
                                    "labelLimit": 170,
                                    "domainColor": theme_palette["border"],
                                    "tickColor": theme_palette["border"],
                                },
                            },
                            "y": {
                                "field": value_column,
                                "type": "quantitative",
                                "axis": {
                                    "title": y_axis_title,
                                    "format": y_axis_format,
                                    "labelColor": theme_palette["muted"],
                                    "titleColor": theme_palette["text"],
                                    "domainColor": theme_palette["border"],
                                    "tickColor": theme_palette["border"],
                                    "gridColor": theme_palette["border"],
                                    "gridOpacity": 0.45,
                                },
                            },
                            "tooltip": [
                                {
                                    "field": index_column,
                                    "type": "nominal",
                                    "title": tooltip_title,
                                },
                                {
                                    "field": value_column,
                                    "type": "quantitative",
                                    "title": tooltip_value_title,
                                    "format": tooltip_value_format,
                                },
                            ],
                        },
                        "height": chart_height,
                        "config": {
                            "background": "transparent",
                            "view": {
                                "stroke": theme_palette["border"],
                                "strokeOpacity": 0.95,
                                "fill": theme_palette["surface_soft"],
                            },
                            "axis": {
                                "labelFontSize": 11,
                                "titleFontSize": 12,
                            },
                        },
                    }
                    st.vega_lite_chart(chart_frame, chart_spec, use_container_width=True)

            render_report_line_chart(
                source_frame=monthly_spend,
                index_column="Month",
                title="Monthly Spend Trend",
                caption_text="Month-over-month purchase spend for the last 12 months.",
                tooltip_title="Month",
                label_angle=-22,
            )

            trend_col_1, trend_col_2 = st.columns(2, gap="small")
            with trend_col_1:
                render_report_line_chart(
                    source_frame=monthly_po_count,
                    index_column="Month",
                    value_column="PO Count",
                    title="PO Volume Trend",
                    caption_text="Number of purchase orders submitted each month (last 12 months).",
                    tooltip_title="Month",
                    label_angle=-22,
                    y_axis_title="PO Count",
                    y_axis_format=",.0f",
                    tooltip_value_title="PO Count",
                    tooltip_value_format=",.0f",
                    chart_height=258,
                )
            with trend_col_2:
                render_report_line_chart(
                    source_frame=monthly_average_po,
                    index_column="Month",
                    value_column="Average PO",
                    title="Average PO Value Trend",
                    caption_text="Average spend per order by month (last 12 months).",
                    tooltip_title="Month",
                    label_angle=-22,
                    y_axis_title="Average PO Value",
                    y_axis_format="$,.0f",
                    tooltip_value_title="Average PO Value",
                    tooltip_value_format="$,.2f",
                    chart_height=258,
                )

            render_report_line_chart(
                source_frame=monthly_cumulative_spend,
                index_column="Month",
                value_column="Cumulative Spend",
                title="Cumulative Spend (Last 12 Months)",
                caption_text="Running spend total across the trailing 12 months.",
                tooltip_title="Month",
                label_angle=-22,
                y_axis_title="Cumulative Spend",
                y_axis_format="$,.0f",
                tooltip_value_title="Cumulative Spend",
                tooltip_value_format="$,.2f",
                chart_height=258,
            )

            chart_col_1, chart_col_2 = st.columns(2, gap="small")
            with chart_col_1:
                render_report_line_chart(
                    source_frame=vendor_spend,
                    index_column="Vendor/Store",
                    title=f"Top Vendors (Top {int(top_n_items)})",
                    caption_text=f"Highest-spend vendors in selected scope (Top {int(top_n_items)}).",
                    tooltip_title="Vendor",
                    label_angle=-28,
                )
            with chart_col_2:
                render_report_line_chart(
                    source_frame=dept_loc_spend,
                    index_column="Department/Loc",
                    title=f"Top Department/Loc (Top {int(top_n_items)})",
                    caption_text=f"Highest-spend department/location pairs (Top {int(top_n_items)}).",
                    tooltip_title="Department/Loc",
                    label_angle=-28,
                )

            render_report_line_chart(
                source_frame=location_spend,
                index_column="Location",
                title=f"Top Locations (Top {int(top_n_items)})",
                caption_text=f"Highest-spend locations in selected scope (Top {int(top_n_items)}).",
                tooltip_title="Location",
                label_angle=-18,
            )

            with st.container(border=True):
                st.markdown("<div class='potrol-report-card-title'>Export Report Data</div>", unsafe_allow_html=True)
                st.markdown(
                    "<div class='potrol-report-card-sub'>Download normalized report rows as CSV.</div>",
                    unsafe_allow_html=True,
                )
                report_csv = po_report_frame.drop(columns=["Date Parsed"], errors="ignore").to_csv(index=False)
                report_scope_slug = (
                    "all_worksheets"
                    if report_scope_mode == "All Worksheets"
                    else (re.sub(r"[^a-z0-9]+", "_", sheet_name.casefold()).strip("_") or "worksheet")
                )
                st.download_button(
                    "Download Report CSV",
                    data=report_csv.encode("utf-8"),
                    file_name=f"{report_scope_slug}_po_report.csv",
                    mime="text/csv",
                    use_container_width=True,
                )


if __name__ == "__main__":
    main()
