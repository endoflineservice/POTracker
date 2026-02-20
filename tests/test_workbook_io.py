import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

import potrol


def build_row(po_number: str, item_name: str, sub_total: float) -> dict[str, object]:
    return {
        "PO Number": po_number,
        "Date": "2026-02-19",
        "Vendor/Store": "Test Vendor",
        "Department": "IT",
        "Location": "GLN",
        "Items Being Purchased": item_name,
        "Price Per Item": float(sub_total),
        "Quantity": 1,
        "Sub Total": float(sub_total),
        "Shipping Cost": 0.0,
        "Sales Tax": 0.0,
        "Grand Total": float(sub_total),
    }


class WorkbookIoTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.workbook_path = self.root / "IT POs.xlsx"
        self.backup_dir = self.root / "PO_Backups"
        self.sheet_name = "PO Log"
        potrol.create_workbook(self.workbook_path, self.sheet_name, potrol.DEFAULT_HEADERS.copy())
        potrol.load_sheet_data.clear()

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_append_record_writes_row_and_creates_backup(self) -> None:
        backup_path = potrol.append_record(
            path=self.workbook_path,
            sheet_name=self.sheet_name,
            headers=potrol.DEFAULT_HEADERS.copy(),
            values=build_row("IT579", "Dock", 120.0),
            backup_dir=self.backup_dir,
            keep_backups=3,
        )
        self.assertIsNotNone(backup_path)
        self.assertTrue(Path(backup_path).exists())

        workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
        worksheet = workbook[self.sheet_name]
        self.assertEqual(str(worksheet.cell(row=2, column=1).value), "IT579")
        self.assertEqual(str(worksheet.cell(row=2, column=6).value), "Dock")
        workbook.close()

    def test_create_backup_respects_keep_latest_limit(self) -> None:
        for _ in range(3):
            backup_path = potrol.create_backup(self.workbook_path, self.backup_dir, keep_latest=2)
            self.assertIsNotNone(backup_path)

        backups = potrol.list_backups(self.workbook_path, self.backup_dir)
        self.assertLessEqual(len(backups), 2)

    def test_update_sheet_rows_supports_update_delete_insert(self) -> None:
        potrol.append_record(
            path=self.workbook_path,
            sheet_name=self.sheet_name,
            headers=potrol.DEFAULT_HEADERS.copy(),
            values=[
                build_row("IT580", "Monitor", 250.0),
                build_row("IT581", "Keyboard", 75.0),
            ],
            backup_dir=self.backup_dir,
            keep_backups=5,
        )

        backup_path = potrol.update_sheet_rows(
            path=self.workbook_path,
            sheet_name=self.sheet_name,
            headers=potrol.DEFAULT_HEADERS.copy(),
            row_updates=[(2, build_row("IT580", "Monitor Updated", 255.0))],
            row_deletes=[3],
            new_rows=[build_row("IT582", "Mouse", 35.0)],
            backup_dir=self.backup_dir,
            keep_backups=5,
        )
        self.assertIsNotNone(backup_path)

        workbook = load_workbook(self.workbook_path, read_only=True, data_only=True)
        worksheet = workbook[self.sheet_name]
        self.assertEqual(str(worksheet.cell(row=2, column=6).value), "Monitor Updated")
        self.assertEqual(str(worksheet.cell(row=3, column=1).value), "IT582")
        self.assertEqual(str(worksheet.cell(row=3, column=6).value), "Mouse")
        workbook.close()

    def test_workbook_write_lock_times_out_when_lock_exists(self) -> None:
        lock_path = potrol.get_workbook_lock_path(self.workbook_path)
        lock_path.write_text("lock", encoding="utf-8")
        try:
            with self.assertRaises(TimeoutError):
                with potrol.workbook_write_lock(
                    self.workbook_path,
                    timeout_seconds=0.25,
                    stale_seconds=9999.0,
                ):
                    pass
        finally:
            lock_path.unlink(missing_ok=True)


if __name__ == "__main__":
    unittest.main()

